import argparse
import base64
import binascii
import html
import hmac
import gzip
import io
import ipaddress
import json
import os
import re
import smtplib
import socket
import ssl
import urllib.parse
import urllib.request
import urllib.error
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from http import cookies
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from email.message import EmailMessage
from io import BytesIO
from cryptography.fernet import Fernet, InvalidToken
import hashlib
import threading
import time
import secrets

import qrcode

try:
    import sentry_sdk
except ImportError:
    sentry_sdk = None

from db import DEFAULT_DB, init_db, ensure_configured_staff, connect, create_session, get_session, audit, insert_id, begin_write, integrity_errors, fetchone_for_update
from security import verify_password, hash_password, random_token, now_ts, password_is_strong, generate_totp_secret, verify_totp, encrypt_pii, decrypt_pii, pii_lookup_hash, pii_key_configured
from antifraud import validate_stamp, FraudError
from wallet import wallet_status, apple_pass_link, google_wallet_link, build_apple_pkpass, google_save_url, google_update_object, apple_auth_token, apple_push_update
from platform_features import has_permission, session_permissions, active_multiplier, add_point_lot, consume_point_lots, expire_points_once, record_purchase
from integrations import platform_order
from intelligence import customer_intelligence, campaign_intelligence

BASE = Path(__file__).resolve().parent
STATIC = BASE / 'static'
DB_PATH = os.environ.get('DATABASE_URL') or os.environ.get('CLUBE_DB_PATH', DEFAULT_DB)
SESSION_COOKIE = 'clube_session'
VERSION='v151'
TERMS_VERSION='1.1'
PRIVACY_VERSION='1.1'
DUMMY_PASSWORD_HASH=hash_password('Fidelizae-Dummy-Password-Only-For-Timing-Protection-2026')


def init_sentry():
    """Inicializa monitoramento de erros somente quando SENTRY_DSN estiver configurado."""
    dsn=(os.environ.get('SENTRY_DSN') or '').strip()
    if not dsn:
        return False
    if sentry_sdk is None:
        print('[SENTRY] SENTRY_DSN configurado, mas sentry-sdk não está instalado')
        return False
    sentry_sdk.init(
        dsn=dsn,
        environment=(os.environ.get('APP_ENV') or 'production').strip() or 'production',
        release=f'fidelizae@{VERSION}',
        send_default_pii=False,
        traces_sample_rate=0.0,
    )
    print(f'[SENTRY] error monitoring ativo release=fidelizae@{VERSION}')
    return True


SENTRY_ENABLED=init_sentry()


def jdump(obj):
    return json.dumps(obj, ensure_ascii=False, separators=(',', ':')).encode('utf-8')


def rowdict(row):
    return dict(row) if row else None

def customer_rowdict(row):
    """Converte linha de cliente para saída segura, descriptografando CPF/telefone só em memória."""
    d=rowdict(row)
    if not d:
        return d
    if 'phone_enc' in d:
        d['phone']=decrypt_pii(d.get('phone_enc'),'phone') or str(d.get('phone') or '')
    if 'cpf_enc' in d:
        d['cpf']=decrypt_pii(d.get('cpf_enc'),'cpf') or str(d.get('cpf') or '')
    d.pop('phone_enc',None); d.pop('cpf_enc',None); d.pop('phone_hash',None); d.pop('cpf_hash',None)
    return d

def protected_customer_pii(phone='', cpf=''):
    """Retorna ciphertext + HMAC pesquisável. Exige CLUBE_ENCRYPTION_KEY quando houver PII."""
    return {
        'phone_enc': encrypt_pii(phone,'phone') if phone else None,
        'phone_hash': pii_lookup_hash(phone,'phone') if phone else None,
        'cpf_enc': encrypt_pii(cpf,'cpf') if cpf else None,
        'cpf_hash': pii_lookup_hash(cpf,'cpf') if cpf else None,
    }

def queue_rowdict(row):
    d=rowdict(row)
    if d and d.get('kind')=='whatsapp':
        d['recipient']=decrypt_pii(d.get('recipient'),'phone')
    if d: d.pop('recipient_hash',None)
    return d

def now_iso():
    return datetime.now(ZoneInfo('America/Sao_Paulo')).isoformat(timespec='seconds')


def legal_context():
    return {
        'company_name': (os.environ.get('CLUBE_LEGAL_COMPANY_NAME') or 'Agência Taboo').strip(),
        'cnpj': (os.environ.get('CLUBE_LEGAL_CNPJ') or '10.995.977/0001-40').strip(),
        'email': (os.environ.get('CLUBE_LEGAL_EMAIL') or 'contato@fidelizae.com.br').strip(),
        'lgpd_email': (os.environ.get('CLUBE_LEGAL_LGPD_EMAIL') or os.environ.get('CLUBE_LEGAL_EMAIL') or 'contato@fidelizae.com.br').strip(),
    }


def render_legal_template(name):
    ctx=legal_context()
    parts=[f'<b>{html.escape(ctx["company_name"])}</b>']
    if ctx['cnpj']: parts.append('CNPJ '+html.escape(ctx['cnpj']))
    parts.append('Contato: <a href="mailto:'+html.escape(ctx['email'],quote=True)+'">'+html.escape(ctx['email'])+'</a>')
    if ctx['lgpd_email'] and ctx['lgpd_email'].lower()!=ctx['email'].lower():
        parts.append('Canal LGPD: <a href="mailto:'+html.escape(ctx['lgpd_email'],quote=True)+'">'+html.escape(ctx['lgpd_email'])+'</a>')
    block='<br>'.join(parts)
    return (STATIC/name).read_text(encoding='utf-8').replace('{{VERSION}}',VERSION).replace('{{LEGAL_ENTITY_BLOCK}}',block).replace('{{TERMS_VERSION}}',TERMS_VERSION).replace('{{PRIVACY_VERSION}}',PRIVACY_VERSION)


BACKUP_FORMAT='fidelizae-platform-backup-v1'
BACKUP_EXCLUDED_TABLES={'sessions','auth_challenges','password_reset_tokens','security_rate_limits'}

def _backup_table_names(conn):
    if str(DB_PATH).startswith(('postgres://','postgresql://')):
        rows=conn.execute("SELECT table_name FROM information_schema.tables WHERE table_schema='public' AND table_type='BASE TABLE' ORDER BY table_name").fetchall()
        names=[r['table_name'] for r in rows]
    else:
        rows=conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name").fetchall()
        names=[r['name'] for r in rows]
    return [n for n in names if n not in BACKUP_EXCLUDED_TABLES and re.fullmatch(r'[A-Za-z_][A-Za-z0-9_]*',n)]


def build_platform_backup(conn):
    tables={}
    for table in _backup_table_names(conn):
        tables[table]=[rowdict(r) for r in conn.execute(f'SELECT * FROM {table}').fetchall()]
    payload={
        'format':BACKUP_FORMAT,
        'generated_at':now_iso(),
        'version':VERSION,
        'database':'postgresql' if str(DB_PATH).startswith(('postgres://','postgresql://')) else 'sqlite',
        'excluded_transient_tables':sorted(BACKUP_EXCLUDED_TABLES),
        'tables':tables,
        'counts':{k:len(v) for k,v in tables.items()},
        'restore_notes':'Restaure somente em banco vazio/inicializado e preserve a mesma CLUBE_ENCRYPTION_KEY para dados criptografados.',
    }
    raw=json.dumps(payload,ensure_ascii=False,sort_keys=True,separators=(',',':')).encode('utf-8')
    payload['sha256']=hashlib.sha256(raw).hexdigest()
    return payload


def verify_platform_backup(payload):
    if not isinstance(payload,dict) or payload.get('format')!=BACKUP_FORMAT or not isinstance(payload.get('tables'),dict):
        return False,'invalid_backup_format'
    expected=str(payload.get('sha256') or '')
    unsigned=dict(payload); unsigned.pop('sha256',None)
    raw=json.dumps(unsigned,ensure_ascii=False,sort_keys=True,separators=(',',':')).encode('utf-8')
    if not expected or not hmac.compare_digest(expected,hashlib.sha256(raw).hexdigest()): return False,'backup_checksum_invalid'
    counts=payload.get('counts') or {}
    for name,rows in payload['tables'].items():
        if not isinstance(rows,list) or int(counts.get(name,-1))!=len(rows): return False,'backup_count_mismatch'
    return True,'ok'


# Cloudflare R2 (S3-compatible) automatic private backups.
_R2_BACKUP_STATE={'configured':False,'last_success':None,'last_error':None,'last_key':None,'last_size':0}
_R2_BACKUP_LOCK=threading.Lock()

def r2_backup_config():
    endpoint=(os.environ.get('R2_ENDPOINT') or '').strip().rstrip('/')
    bucket=(os.environ.get('R2_BUCKET') or '').strip()
    access=(os.environ.get('R2_ACCESS_KEY_ID') or '').strip()
    secret=(os.environ.get('R2_SECRET_ACCESS_KEY') or '').strip()
    account=(os.environ.get('R2_ACCOUNT_ID') or '').strip()
    if not endpoint and account: endpoint=f'https://{account}.r2.cloudflarestorage.com'
    return {'endpoint':endpoint,'bucket':bucket,'access_key':access,'secret_key':secret,'account_id':account}

def r2_backup_configured():
    c=r2_backup_config()
    return bool(c['endpoint'] and c['bucket'] and c['access_key'] and c['secret_key'])

def _aws_sigv4_headers(method,url,body,access_key,secret_key,content_type='application/gzip'):
    parsed=urllib.parse.urlsplit(url); host=parsed.netloc
    now=datetime.now(ZoneInfo('UTC')); amz_date=now.strftime('%Y%m%dT%H%M%SZ'); date_stamp=now.strftime('%Y%m%d')
    payload_hash=hashlib.sha256(body).hexdigest()
    canonical_uri=urllib.parse.quote(urllib.parse.unquote(parsed.path or '/'),safe='/-_.~')
    canonical_query=parsed.query
    headers={'content-type':content_type,'host':host,'x-amz-content-sha256':payload_hash,'x-amz-date':amz_date}
    signed_headers=';'.join(sorted(headers))
    canonical_headers=''.join(f'{k}:{headers[k].strip()}\n' for k in sorted(headers))
    canonical_request='\n'.join([method,canonical_uri,canonical_query,canonical_headers,signed_headers,payload_hash])
    scope=f'{date_stamp}/auto/s3/aws4_request'
    string_to_sign='AWS4-HMAC-SHA256\n'+amz_date+'\n'+scope+'\n'+hashlib.sha256(canonical_request.encode()).hexdigest()
    def sign(key,msg): return hmac.new(key,msg.encode(),hashlib.sha256).digest()
    k_date=sign(('AWS4'+secret_key).encode(),date_stamp); k_region=sign(k_date,'auto'); k_service=sign(k_region,'s3'); k_signing=sign(k_service,'aws4_request')
    signature=hmac.new(k_signing,string_to_sign.encode(),hashlib.sha256).hexdigest()
    headers['authorization']=f'AWS4-HMAC-SHA256 Credential={access_key}/{scope}, SignedHeaders={signed_headers}, Signature={signature}'
    return {('Authorization' if k=='authorization' else '-'.join(x.capitalize() for x in k.split('-'))):v for k,v in headers.items() if k!='host'}

def r2_put_object(key,data,content_type='application/gzip'):
    c=r2_backup_config(); key=str(key).lstrip('/')
    url=f"{c['endpoint']}/{urllib.parse.quote(c['bucket'],safe='')}/{urllib.parse.quote(key,safe='/')}"
    headers=_aws_sigv4_headers('PUT',url,data,c['access_key'],c['secret_key'],content_type)
    req=urllib.request.Request(url,data=data,method='PUT',headers=headers)
    with urllib.request.urlopen(req,timeout=45) as resp:
        if int(resp.status) not in (200,201,204): raise RuntimeError(f'r2_http_{resp.status}')
    return True

def create_r2_backup(kind='daily'):
    if not r2_backup_configured(): raise RuntimeError('r2_not_configured')
    with _R2_BACKUP_LOCK:
        with connect(DB_PATH) as conn: payload=build_platform_backup(conn)
        ok,reason=verify_platform_backup(payload)
        if not ok: raise RuntimeError(reason)
        raw=json.dumps(payload,ensure_ascii=False,separators=(',',':')).encode('utf-8')
        compressed=gzip.compress(raw,compresslevel=6,mtime=0)
        now=datetime.now(ZoneInfo('America/Sao_Paulo'))
        if kind=='monthly': key=f'monthly/{now:%Y-%m}/fidelizae-backup.json.gz'
        else: key=f'daily/{now:%Y-%m-%d}/fidelizae-backup.json.gz'
        r2_put_object(key,compressed)
        _R2_BACKUP_STATE.update({'configured':True,'last_success':now.isoformat(timespec='seconds'),'last_error':None,'last_key':key,'last_size':len(compressed)})
        print(f'[R2_BACKUP] success kind={kind} key={key} bytes={len(compressed)} sha256={payload["sha256"][:12]}')
        return {'key':key,'bytes':len(compressed),'sha256':payload['sha256']}

def r2_backup_status():
    out=dict(_R2_BACKUP_STATE); out['configured']=r2_backup_configured(); out['bucket']=r2_backup_config().get('bucket') if out['configured'] else None
    return out

def run_scheduled_r2_backup_once():
    if not r2_backup_configured(): return
    now=datetime.now(ZoneInfo('America/Sao_Paulo'))
    # Deterministic object keys make restarts idempotent. Run after 03:00 local time.
    if now.hour < 3: return
    today=now.strftime('%Y-%m-%d')
    if getattr(run_scheduled_r2_backup_once,'_daily',None)!=today:
        create_r2_backup('daily'); run_scheduled_r2_backup_once._daily=today
    month=now.strftime('%Y-%m')
    if now.day==1 and getattr(run_scheduled_r2_backup_once,'_monthly',None)!=month:
        create_r2_backup('monthly'); run_scheduled_r2_backup_once._monthly=month


def _safe_ip(value, fallback='0.0.0.0'):
    try:
        return str(ipaddress.ip_address(str(value or '').strip()))
    except ValueError:
        try: return str(ipaddress.ip_address(str(fallback or '').strip()))
        except ValueError: return '0.0.0.0'


def _email_tag(email):
    normalized=str(email or '').strip().lower()
    return hashlib.sha256(normalized.encode('utf-8')).hexdigest()[:12] if normalized else 'none'


def persistent_rate_allow(rate_key, limit, window, block_seconds=None):
    """Rate limit compartilhado por todas as instâncias e preservado em restarts."""
    ts=now_ts(); limit=max(1,int(limit)); window=max(1,int(window)); block_seconds=max(window,int(block_seconds or window))
    with connect(DB_PATH) as conn:
        begin_write(conn)
        # Cria o estado de forma idempotente antes de bloquear a linha; evita corrida entre instâncias.
        conn.execute('INSERT INTO security_rate_limits(rate_key,window_start,count,blocked_until,updated_at) VALUES(?,?,?,?,?) ON CONFLICT(rate_key) DO NOTHING',(rate_key,ts,0,0,ts))
        row=fetchone_for_update(conn,'SELECT rate_key,window_start,count,blocked_until FROM security_rate_limits WHERE rate_key=?',(rate_key,))
        if int(row['blocked_until'] or 0)>ts:
            return False, int(row['blocked_until'])-ts
        if ts-int(row['window_start'] or 0)>=window:
            conn.execute('UPDATE security_rate_limits SET window_start=?,count=1,blocked_until=0,updated_at=? WHERE rate_key=?',(ts,ts,rate_key))
            return True, 0
        count=int(row['count'] or 0)+1
        if count>limit:
            blocked_until=ts+block_seconds
            conn.execute('UPDATE security_rate_limits SET count=?,blocked_until=?,updated_at=? WHERE rate_key=?',(count,blocked_until,ts,rate_key))
            return False, block_seconds
        conn.execute('UPDATE security_rate_limits SET count=?,updated_at=? WHERE rate_key=?',(count,ts,rate_key))
        return True, 0


def persistent_rate_reset(rate_key):
    try:
        with connect(DB_PATH) as conn: conn.execute('DELETE FROM security_rate_limits WHERE rate_key=?',(rate_key,))
    except Exception: pass


def _cookie_secure():
    default_secure = '1' if (str(DB_PATH).startswith(('postgres://','postgresql://')) or (os.environ.get('APP_ENV') or '').lower()=='production') else '0'
    return os.environ.get('CLUBE_SECURE_COOKIE',default_secure)=='1'


def _session_cookie(token,max_age=28800):
    value=f'{SESSION_COOKIE}={token}; Path=/; HttpOnly; SameSite=Strict; Max-Age={int(max_age)}'
    if _cookie_secure(): value+='; Secure'
    return value


def _challenge_cookie(raw,max_age=300):
    value=f'clube_2fa_challenge={raw}; Path=/; HttpOnly; SameSite=Strict; Max-Age={int(max_age)}'
    if _cookie_secure(): value+='; Secure'
    return value


def _clear_cookie(name):
    value=f'{name}=deleted; Path=/; HttpOnly; SameSite=Strict; Max-Age=0'
    if _cookie_secure(): value+='; Secure'
    return value


def create_2fa_challenge(conn,user_id,ttl=300):
    raw=random_token(32); th=hashlib.sha256(raw.encode()).hexdigest(); ts=now_ts()
    conn.execute('DELETE FROM auth_challenges WHERE user_id=? OR expires_at<?',(user_id,ts))
    conn.execute('INSERT INTO auth_challenges(token_hash,user_id,expires_at,attempts,created_at) VALUES(?,?,?,?,?)',(th,user_id,ts+ttl,0,ts))
    return raw


def _totp_uri(secret,email):
    label=urllib.parse.quote(f'Fidelizaê!:{email}',safe='')
    issuer=urllib.parse.quote('Fidelizaê!',safe='')
    return f'otpauth://totp/{label}?secret={urllib.parse.quote(secret)}&issuer={issuer}&algorithm=SHA1&digits=6&period=30'


def _totp_qr_data(uri):
    img=qrcode.make(uri); out=BytesIO(); img.save(out,format='PNG')
    return 'data:image/png;base64,'+base64.b64encode(out.getvalue()).decode('ascii')


def current_branch_id(conn, user_id, campaign_id=None):
    if not user_id:
        return None
    if campaign_id:
        row=conn.execute('SELECT branch_id FROM users WHERE id=? AND campaign_id=?',(user_id,campaign_id)).fetchone()
    else:
        row=conn.execute('SELECT branch_id FROM users WHERE id=?',(user_id,)).fetchone()
    return row['branch_id'] if row and row['branch_id'] else None


ECOMMERCE_PLATFORMS={'none','woocommerce','nuvemshop','shopify','tray','vtex','loja_integrada','custom'}

def normalize_ecommerce_platform(value):
    value=str(value or 'none').strip().lower()
    return value if value in ECOMMERCE_PLATFORMS else 'none'

def ecommerce_extract(payload, platform):
    data=payload if isinstance(payload,dict) else {}
    order_id=data.get('order_id') or data.get('id') or data.get('number') or data.get('order_number')
    status=str(data.get('payment_status') or data.get('financial_status') or data.get('status') or '').strip().lower()
    total=data.get('total_cents')
    if total is None:
        total=data.get('total_price', data.get('total', data.get('amount',0)))
        try: total=int(round(float(str(total).replace(',','.'))*100))
        except: total=0
    else:
        try: total=int(total)
        except: total=0
    customer=data.get('customer') if isinstance(data.get('customer'),dict) else {}
    billing=data.get('billing') if isinstance(data.get('billing'),dict) else {}
    billing_address=data.get('billing_address') if isinstance(data.get('billing_address'),dict) else {}
    email=data.get('email') or customer.get('email') or billing.get('email')
    phone=data.get('phone') or customer.get('phone') or billing.get('phone') or billing_address.get('phone')
    cpf=data.get('cpf') or data.get('document') or customer.get('cpf') or customer.get('document')
    if not cpf and isinstance(data.get('meta_data'),list):
        for item in data['meta_data']:
            if isinstance(item,dict) and str(item.get('key','')).lower() in ('cpf','billing_cpf','_billing_cpf'):
                cpf=item.get('value'); break
    return {'order_id':str(order_id or '').strip()[:120],'status':status[:40],'total_cents':max(0,total),
            'email':normalize_email(email),'phone':normalize_phone(phone),'cpf':normalize_cpf(cpf)}

def ecommerce_find_membership(conn,campaign_id,info):
    cpf=info.get('cpf'); email=info.get('email'); phone=info.get('phone')
    if cpf:
        row=conn.execute('SELECT m.*,cu.name customer_name FROM memberships m JOIN customers cu ON cu.id=m.customer_id WHERE m.campaign_id=? AND cu.cpf_hash=? LIMIT 1',(campaign_id,pii_lookup_hash(cpf,'cpf'))).fetchone()
        if row:return row
    if email:
        row=conn.execute('SELECT m.*,cu.name customer_name FROM memberships m JOIN customers cu ON cu.id=m.customer_id WHERE m.campaign_id=? AND lower(cu.email)=lower(?) LIMIT 1',(campaign_id,email)).fetchone()
        if row:return row
    if phone:
        row=conn.execute('SELECT m.*,cu.name customer_name FROM memberships m JOIN customers cu ON cu.id=m.customer_id WHERE m.campaign_id=? AND cu.phone_hash=? LIMIT 1',(campaign_id,pii_lookup_hash(phone,'phone'))).fetchone()
        if row:return row
    return None

EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')


def normalize_email(value):
    value = str(value or '').strip().lower()
    return value if len(value) <= 160 and EMAIL_RE.fullmatch(value) else None


def normalize_phone(value):
    digits = re.sub(r'\D', '', str(value or ''))
    if digits.startswith('55') and len(digits) == 13:
        local = digits[2:]
    elif len(digits) == 11:
        local = digits
    else:
        return None
    if len(local) != 11 or local[0] == '0' or local[2] != '9':
        return None
    return '55' + local


def normalize_cpf(value):
    digits = re.sub(r'\D', '', str(value or ''))
    if len(digits) != 11 or digits == digits[0] * 11:
        return None
    nums = [int(x) for x in digits]
    d1 = (sum(nums[i] * (10 - i) for i in range(9)) * 10) % 11
    d1 = 0 if d1 == 10 else d1
    d2 = (sum(nums[i] * (11 - i) for i in range(10)) * 10) % 11
    d2 = 0 if d2 == 10 else d2
    return digits if nums[9] == d1 and nums[10] == d2 else None


def device_os_from_user_agent(user_agent):
    ua=(user_agent or '').lower()
    if 'android' in ua: return 'android'
    if any(x in ua for x in ('iphone','ipad','ipod')): return 'ios'
    return 'other'

def normalize_birth_date(value):
    value = str(value or '').strip()
    try:
        born = date.fromisoformat(value)
    except ValueError:
        return None
    today = datetime.now(ZoneInfo('America/Sao_Paulo')).date()
    if born > today or born.year < 1900:
        return None
    return born.isoformat()





def _secret_box():
    master=os.environ.get('CLUBE_ENCRYPTION_KEY','').strip()
    if not master:
        return None
    key=base64.urlsafe_b64encode(hashlib.sha256(master.encode('utf-8')).digest())
    return Fernet(key)

def encrypt_secret(value):
    value=str(value or '')
    if not value: return None
    box=_secret_box()
    if not box: raise RuntimeError('encryption_key_not_configured')
    return box.encrypt(value.encode('utf-8')).decode('ascii')

def decrypt_secret(value):
    if not value: return ''
    box=_secret_box()
    if not box: return ''
    try: return box.decrypt(str(value).encode('ascii')).decode('utf-8')
    except (InvalidToken,ValueError): return ''

def client_integrations(conn, campaign_id):
    row=conn.execute("""SELECT id,smtp_host,smtp_port,smtp_user,smtp_password_enc,smtp_from,smtp_from_name,smtp_security,email_provider,brevo_api_key_enc,brevo_sender_email,brevo_sender_name,brevo_reply_to,
        whatsapp_phone_number_id,whatsapp_waba_id,whatsapp_access_token_enc,whatsapp_api_version
        FROM campaigns WHERE id=?""",(campaign_id,)).fetchone()
    if not row: return None
    x=rowdict(row)
    x['smtp_password']=decrypt_secret(x.pop('smtp_password_enc',None))
    x['brevo_api_key']=decrypt_secret(x.pop('brevo_api_key_enc',None))
    x['whatsapp_access_token']=decrypt_secret(x.pop('whatsapp_access_token_enc',None))
    return x

def global_smtp_config():
    return {
        'host':os.environ.get('CLUBE_SMTP_HOST','').strip(),
        'port':os.environ.get('CLUBE_SMTP_PORT','587').strip(),
        'user':os.environ.get('CLUBE_SMTP_USER','').strip(),
        'password':os.environ.get('CLUBE_SMTP_PASSWORD',''),
        'from_addr':os.environ.get('CLUBE_SMTP_FROM','').strip(),
        'from_name':os.environ.get('CLUBE_SMTP_FROM_NAME','Fidelizaê!').strip(),
        'security':os.environ.get('CLUBE_SMTP_SECURITY','starttls').strip().lower(),
        'source':'global'
    }

def email_config_for_client(conn=None,campaign_id=None):
    # Integração promocional sempre isolada por cliente, sem fallback global da Taboo.
    if conn is not None and campaign_id:
        x=client_integrations(conn,campaign_id)
        if x:
            provider=(x.get('email_provider') or 'smtp').strip().lower()
            if provider=='brevo':
                return {'provider':'brevo','api_key':x.get('brevo_api_key') or '',
                        'sender_email':x.get('brevo_sender_email') or '',
                        'sender_name':x.get('brevo_sender_name') or '',
                        'reply_to':x.get('brevo_reply_to') or '','source':'client'}
            return {'provider':'smtp','host':x.get('smtp_host') or '',
                    'port':str(x.get('smtp_port') or 587),'user':x.get('smtp_user') or '',
                    'password':x.get('smtp_password') or '','from_addr':x.get('smtp_from') or '',
                    'from_name':x.get('smtp_from_name') or '',
                    'security':x.get('smtp_security') or 'starttls','source':'client'}
    return {'provider':'smtp','host':'','port':'587','user':'','password':'','from_addr':'','from_name':'','security':'starttls','source':'client'}

def smtp_configured(config=None):
    c=config or global_smtp_config()
    return bool(c.get('host') and c.get('from_addr'))


def brevo_api_config():
    return {
        'api_key':os.environ.get('BREVO_API_KEY','').strip(),
        'sender_email':os.environ.get('BREVO_SENDER_EMAIL','contato@fidelizae.com.br').strip(),
        'sender_name':os.environ.get('BREVO_SENDER_NAME','Fidelizaê!').strip(),
        'reply_to':os.environ.get('BREVO_REPLY_TO','contato@fidelizae.com.br').strip(),
    }

def brevo_api_configured():
    c=brevo_api_config()
    return bool(c.get('api_key') and c.get('sender_email'))

def global_email_config():
    # Infraestrutura institucional do Fidelizaê!, usada apenas para mensagens da própria plataforma.
    # Nunca é reutilizada como configuração promocional das empresas clientes.
    if brevo_api_configured():
        c=brevo_api_config(); c.update({'provider':'brevo','source':'global'}); return c
    c=global_smtp_config(); c.update({'provider':'smtp','source':'global'}); return c

def email_configured(config=None):
    c=config or {}
    if c.get('provider')=='brevo': return bool(c.get('api_key') and c.get('sender_email'))
    if c.get('provider')=='smtp': return smtp_configured(c)
    return brevo_api_configured() or smtp_configured(c or None)

def _brevo_payload_from_message(msg, cfg=None):
    cfg=cfg or brevo_api_config()
    to=[]
    for addr in msg.get_all('To',[]):
        for item in str(addr).split(','):
            email=item.strip()
            if email: to.append({'email':email})
    if not to:
        raise ValueError('email_recipient_missing')
    text=''; html_content=''; attachments=[]
    if msg.is_multipart():
        for part in msg.walk():
            ctype=part.get_content_type(); disp=part.get_content_disposition()
            if disp=='attachment':
                raw=part.get_payload(decode=True) or b''
                attachments.append({'name':part.get_filename() or 'anexo','content':base64.b64encode(raw).decode('ascii')})
            elif ctype=='text/plain' and not text:
                try: text=part.get_content()
                except Exception: pass
            elif ctype=='text/html' and not html_content:
                try: html_content=part.get_content()
                except Exception: pass
    else:
        ctype=msg.get_content_type()
        try: content=msg.get_content()
        except Exception: content=str(msg.get_payload() or '')
        if ctype=='text/html': html_content=content
        else: text=content
    payload={
        'sender':{'name':cfg.get('sender_name') or 'Fidelizaê!','email':cfg['sender_email']},
        'to':to,
        'subject':str(msg.get('Subject') or 'Fidelizaê!'),
    }
    if html_content: payload['htmlContent']=html_content
    else: payload['textContent']=text or 'Fidelizaê!'
    if attachments: payload['attachment']=attachments
    # Respeita o Reply-To específico da mensagem (ex.: formulário comercial).
    # Só usa o Reply-To global como fallback.
    reply_to=str(msg.get('Reply-To') or cfg.get('reply_to') or '').strip()
    if reply_to: payload['replyTo']={'email':reply_to}
    return payload

def _brevo_blocked_ip_details(raw):
    """Detecta bloqueio de IP da Brevo sem expor credenciais.

    Retorna (bloqueado, ip). A Brevo pode usar grafias "unrecognised"
    ou "unrecognized" dependendo da resposta/idioma.
    """
    text=str(raw or '')
    try:
        data=json.loads(text)
        if isinstance(data,dict):
            text=str(data.get('message') or data.get('error') or text)
    except Exception:
        pass
    low=text.lower()
    blocked=('unrecognised ip address' in low or 'unrecognized ip address' in low or
             ('ip address' in low and ('not authorised' in low or 'not authorized' in low)))
    if not blocked:
        return False,''
    match=re.search(r'(?<!\d)(?:\d{1,3}\.){3}\d{1,3}(?!\d)',text)
    return True,(match.group(0) if match else '')

def send_email_brevo_api(msg, cfg=None):
    cfg=cfg or brevo_api_config()
    if not (cfg.get('api_key') and cfg.get('sender_email')):
        return {'sent':False,'reason':'brevo_api_not_configured'}
    try:
        payload=_brevo_payload_from_message(msg,cfg)
        req=urllib.request.Request('https://api.brevo.com/v3/smtp/email',data=json.dumps(payload,ensure_ascii=False).encode('utf-8'),method='POST',headers={
            'api-key':cfg['api_key'],'Content-Type':'application/json','Accept':'application/json'
        })
        with urllib.request.urlopen(req,timeout=20) as resp:
            data=json.loads(resp.read().decode('utf-8') or '{}')
        return {'sent':True,'source':'brevo_api','message_id':data.get('messageId')}
    except urllib.error.HTTPError as exc:
        raw=exc.read().decode('utf-8',errors='replace')
        ip_blocked,blocked_ip=_brevo_blocked_ip_details(raw)
        if exc.code in (401,403) and ip_blocked:
            # Mensagem operacional explícita para o Railway. Não imprime API Key,
            # payload do lead nem outras credenciais; apenas o IP recusado.
            ip_label=blocked_ip or 'nao_informado'
            print(f'[EMAIL] BREVO_IP_BLOCKED status={exc.code} ip={ip_label} action=brevo_security_authorized_ips_or_automatic_authorization')
            return {'sent':False,'reason':'brevo_ip_blocked','status':exc.code,'source':'brevo_api','blocked_ip':blocked_ip}
        print(f'[EMAIL] BREVO_HTTP_ERROR status={exc.code} body={raw[:500]}')
        reason='brevo_auth_failed' if exc.code in (401,403) else ('brevo_sender_invalid' if exc.code==400 else 'brevo_api_failed')
        return {'sent':False,'reason':reason,'status':exc.code,'source':'brevo_api'}
    except (TimeoutError, socket.timeout):
        print('[EMAIL] BREVO_TIMEOUT')
        return {'sent':False,'reason':'brevo_api_timeout','source':'brevo_api'}
    except urllib.error.URLError as exc:
        print(f'[EMAIL] BREVO_CONNECTION_FAILED error={type(exc.reason).__name__ if getattr(exc,"reason",None) else type(exc).__name__}')
        return {'sent':False,'reason':'brevo_api_connection_failed','source':'brevo_api'}
    except Exception as exc:
        print(f'[EMAIL] BREVO_SEND_FAILED error={type(exc).__name__}')
        return {'sent':False,'reason':'brevo_api_failed','source':'brevo_api'}

def send_email_message(msg, config=None):
    c=config or {}
    if c.get('provider')=='brevo': return send_email_brevo_api(msg,c)
    if c.get('provider')=='smtp': pass
    elif brevo_api_configured(): return send_email_brevo_api(msg)
    c=c or global_smtp_config()
    host=c.get('host','').strip(); from_addr=c.get('from_addr','').strip()
    if not host or not from_addr: return {'sent':False,'reason':'smtp_not_configured'}
    try: port=int(c.get('port') or 587)
    except (ValueError,TypeError): port=587
    user=c.get('user','').strip(); smtp_password=c.get('password','')
    security=(c.get('security') or 'starttls').strip().lower()
    from_name=(c.get('from_name') or '').strip()
    msg['From']=f'{from_name} <{from_addr}>' if from_name else from_addr
    context=ssl.create_default_context()
    try:
        if security=='ssl':
            with smtplib.SMTP_SSL(host,port,timeout=15,context=context) as smtp:
                if user: smtp.login(user,smtp_password)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(host,port,timeout=15) as smtp:
                smtp.ehlo()
                if security!='none':
                    smtp.starttls(context=context); smtp.ehlo()
                if user: smtp.login(user,smtp_password)
                smtp.send_message(msg)
        return {'sent':True,'source':c.get('source','global')}
    except smtplib.SMTPAuthenticationError as exc:
        print(f'[EMAIL] AUTH_FAILED to={msg.get("To","")} code={getattr(exc,"smtp_code","")}')
        return {'sent':False,'reason':'smtp_auth_failed','source':c.get('source','global')}
    except smtplib.SMTPSenderRefused as exc:
        print(f'[EMAIL] SENDER_REFUSED to={msg.get("To","")} code={getattr(exc,"smtp_code","")}')
        return {'sent':False,'reason':'smtp_sender_refused','source':c.get('source','global')}
    except smtplib.SMTPRecipientsRefused as exc:
        print(f'[EMAIL] RECIPIENT_REFUSED to={msg.get("To","")}')
        return {'sent':False,'reason':'smtp_recipient_refused','source':c.get('source','global')}
    except (TimeoutError, socket.timeout) as exc:
        print(f'[EMAIL] TIMEOUT to={msg.get("To","")}')
        return {'sent':False,'reason':'smtp_timeout','source':c.get('source','global')}
    except (ConnectionError, OSError, smtplib.SMTPConnectError) as exc:
        print(f'[EMAIL] CONNECTION_FAILED to={msg.get("To","")} error={type(exc).__name__}')
        return {'sent':False,'reason':'smtp_connection_failed','source':c.get('source','global')}
    except Exception as exc:
        print(f'[EMAIL] SEND_FAILED to={msg.get("To","")} error={type(exc).__name__}')
        return {'sent':False,'reason':'smtp_send_failed','source':c.get('source','global')}

def validate_logo_data(value):
    """Valida e normaliza a logo enviada pelo Painel Taboo.

    Mantém o data URL original para persistência, mas valida MIME, base64,
    tamanho máximo de 500 KB e a assinatura real do arquivo.
    """
    if not value:
        return None
    text=str(value).strip()
    match=re.fullmatch(r'data:(image/(?:png|jpeg|webp));base64,([A-Za-z0-9+/=\s]+)',text)
    if not match:
        raise ValueError('invalid_logo_format')
    try:
        raw=base64.b64decode(re.sub(r'\s+','',match.group(2)),validate=True)
    except (binascii.Error,ValueError):
        raise ValueError('invalid_logo_format')
    if not raw:
        raise ValueError('invalid_logo_format')
    if len(raw)>500_000:
        raise ValueError('logo_too_large')
    mime=match.group(1)
    valid=(
        (mime=='image/png' and raw.startswith(b'\x89PNG\r\n\x1a\n')) or
        (mime=='image/jpeg' and raw.startswith(b'\xff\xd8\xff')) or
        (mime=='image/webp' and len(raw)>=12 and raw[:4]==b'RIFF' and raw[8:12]==b'WEBP')
    )
    if not valid:
        raise ValueError('invalid_logo_format')
    return f'data:{mime};base64,'+base64.b64encode(raw).decode('ascii')


def decode_image_data(value, max_bytes=700_000):
    if not value:
        return None
    text=str(value)
    match=re.fullmatch(r'data:(image/(?:png|jpeg|webp));base64,([A-Za-z0-9+/=\s]+)',text)
    if not match:
        raise ValueError('invalid_image_format')
    try:
        raw=base64.b64decode(match.group(2),validate=True)
    except (binascii.Error,ValueError):
        raise ValueError('invalid_image_format')
    if not raw or len(raw)>max_bytes:
        raise ValueError('image_too_large')
    subtype={'image/png':'png','image/jpeg':'jpeg','image/webp':'webp'}[match.group(1)]
    return raw, subtype


def send_campaign_email(to_email, to_name, message, image_data=None, subject='Mensagem do Fidelizaê!', smtp_config=None):
    if not email_configured(smtp_config):
        return {'sent':False,'reason':'smtp_not_configured'}
    msg=EmailMessage()
    msg['Subject']=subject
    msg['To']=to_email
    text=(str(message or '').strip() or 'Você recebeu uma nova mensagem do seu programa de fidelidade.')
    msg.set_content(text)
    if image_data:
        raw,subtype=decode_image_data(image_data)
        msg.add_attachment(raw,maintype='image',subtype=subtype,filename='clube-fidelidade.'+('jpg' if subtype=='jpeg' else subtype))
    return send_email_message(msg, smtp_config)


def send_password_recovery_email(email, reset_token, smtp_config=None):
    if not email_configured(smtp_config):
        return {'sent':False,'reason':'smtp_not_configured'}
    base=(os.environ.get('PUBLIC_BASE_URL') or 'https://app.fidelizae.com.br').rstrip('/')
    reset_url=base+'/reset-password?token='+urllib.parse.quote(reset_token)
    msg=EmailMessage(); msg['Subject']='Redefinição de senha • Fidelizaê!'; msg['To']=email
    msg.set_content('Recebemos uma solicitação para redefinir sua senha no Fidelizaê!.\n\nAbra o link abaixo (válido por 30 minutos):\n'+reset_url+'\n\nSe você não solicitou a alteração, ignore esta mensagem.')
    return send_email_message(msg, smtp_config)



PLAN_FEATURES={
 'beginner':{'client_limit':50,'staff_limit':1,'points':False,'communications':False,'advanced':False,'coupons':False,'reports':True,'complete_reports':False,'advanced_reports':False,'nps':False,'vip_tiers':False,'multipliers':False,'gift_cards':False,'automations':False,'customer_area':False},
 'intermediate':{'client_limit':0,'staff_limit':5,'points':True,'communications':False,'advanced':False,'coupons':True,'reports':True,'complete_reports':True,'advanced_reports':False,'nps':False,'vip_tiers':False,'multipliers':False,'gift_cards':False,'automations':False,'customer_area':False},
 'pro':{'client_limit':0,'staff_limit':0,'points':True,'communications':True,'advanced':True,'coupons':True,'reports':True,'complete_reports':True,'advanced_reports':True,'nps':True,'vip_tiers':True,'multipliers':True,'gift_cards':True,'automations':True,'customer_area':True},
}
def normalize_plan(v):
    v=str(v or 'beginner').strip().lower()
    return v if v in PLAN_FEATURES else 'beginner'
def campaign_plan(conn,campaign_id):
    try:
        r=reconcile_campaign_billing(conn,campaign_id)
    except NameError:
        r=conn.execute('SELECT plan FROM campaigns WHERE id=?',(campaign_id,)).fetchone()
    return normalize_plan(r['plan'] if r else 'beginner')
def plan_allows(conn,campaign_id,feature):
    return bool(PLAN_FEATURES[campaign_plan(conn,campaign_id)].get(feature))

PLAN_PRICES={'beginner':0.0,'intermediate':49.90,'pro':99.90}
BILLING_OPTIONS={
    'beginner': {'free': {'amount':0.0,'frequency':None,'frequency_type':None,'commitment_days':0,'label':'Grátis'}},
    'intermediate': {
        'monthly': {'amount':49.90,'frequency':1,'frequency_type':'months','commitment_days':0,'label':'Mensal — R$ 49,90/mês'},
        'annual_monthly': {'amount':44.90,'frequency':1,'frequency_type':'months','commitment_days':365,'label':'Anual — R$ 44,90/mês por 12 meses'},
        'annual_upfront': {'amount':515.00,'frequency':12,'frequency_type':'months','commitment_days':365,'label':'Anual à vista — R$ 515,00/ano'},
    },
    'pro': {
        'monthly': {'amount':99.90,'frequency':1,'frequency_type':'months','commitment_days':0,'label':'Mensal — R$ 99,90/mês'},
        'annual_monthly': {'amount':89.90,'frequency':1,'frequency_type':'months','commitment_days':365,'label':'Anual — R$ 89,90/mês por 12 meses'},
        'annual_upfront': {'amount':1020.00,'frequency':12,'frequency_type':'months','commitment_days':365,'label':'Anual à vista — R$ 1.020,00/ano'},
    },
}

def normalize_billing_option(plan,value):
    plan=normalize_plan(plan)
    if plan=='beginner': return 'free'
    value=str(value or 'monthly').strip().lower()
    return value if value in BILLING_OPTIONS.get(plan,{}) else 'monthly'

def billing_config(plan,billing_option):
    plan=normalize_plan(plan); option=normalize_billing_option(plan,billing_option)
    return option,BILLING_OPTIONS[plan][option]



def mp_request(method,path,payload=None,extra_headers=None):
    token=os.environ.get('MERCADOPAGO_ACCESS_TOKEN','').strip()
    if not token: raise RuntimeError('mercadopago_not_configured')
    data=json.dumps(payload).encode() if payload is not None else None
    headers={'Authorization':'Bearer '+token,'Content-Type':'application/json'}
    if extra_headers:
        for key,value in dict(extra_headers).items():
            if value is not None and str(value).strip(): headers[str(key)]=str(value).strip()
    req=urllib.request.Request('https://api.mercadopago.com'+path,data=data,method=method,headers=headers)
    try:
        with urllib.request.urlopen(req,timeout=20) as r:return json.loads(r.read().decode() or '{}')
    except urllib.error.HTTPError as e:
        body=e.read().decode(errors='replace')[:1000]; print('[BILLING] MP_ERROR',e.code,body); raise RuntimeError('mercadopago_api_error')


def _mp_subscription_diagnostic(sub, event='MP_SUBSCRIPTION'):
    """Registra somente metadados não secretos úteis ao diagnóstico de Assinaturas.

    Nunca inclui payer_email, access token, dados do cartão, URL completa do checkout ou
    qualquer segredo. IDs de conta/cartão são reduzidos a flags de presença.
    """
    sub=sub if isinstance(sub,dict) else {}
    recurring=sub.get('auto_recurring') if isinstance(sub.get('auto_recurring'),dict) else {}
    safe={
        'id': str(sub.get('id') or '') or None,
        'status': str(sub.get('status') or '') or None,
        'reason': str(sub.get('reason') or '')[:120] or None,
        'external_reference': str(sub.get('external_reference') or '')[:160] or None,
        'init_point_present': bool(sub.get('init_point')),
        'payer_id_present': bool(sub.get('payer_id')),
        'card_id_present': bool(sub.get('card_id')),
        'payment_method_id': str(sub.get('payment_method_id') or '')[:60] or None,
        'frequency': recurring.get('frequency'),
        'frequency_type': recurring.get('frequency_type'),
        'transaction_amount': recurring.get('transaction_amount'),
        'currency_id': recurring.get('currency_id'),
        'date_created': sub.get('date_created'),
        'last_modified': sub.get('last_modified'),
        'next_payment_date': sub.get('next_payment_date'),
    }
    print('[BILLING] '+event+' '+json.dumps(safe,ensure_ascii=False,default=str),flush=True)
    return safe


def create_mp_subscription(email,plan,reference,device_id=None,billing_option='monthly'):
    billing_option,cfg=billing_config(plan,billing_option)
    amount=cfg['amount']; base=(os.environ.get('PUBLIC_BASE_URL') or 'https://app.fidelizae.com.br').rstrip('/')
    # O payer de teste só pode substituir o e-mail real fora de produção.
    # Isso evita que MERCADOPAGO_TEST_PAYER_EMAIL, deixado por engano no Railway,
    # faça uma venda real tentar cobrar uma conta de teste do Mercado Pago.
    environment=(os.environ.get('APP_ENV') or 'production').strip().lower()
    test_payer=(os.environ.get('MERCADOPAGO_TEST_PAYER_EMAIL') or '').strip()
    test_environments={'test','testing','development','dev','staging','sandbox'}
    if test_payer and environment in test_environments:
        payer_email=test_payer
    else:
        payer_email=str(email or '').strip()
        if test_payer and environment not in test_environments:
            print(f'[BILLING] TEST_PAYER_IGNORED environment={environment}')
    # O Device ID é coletado no navegador pelo security.js oficial do Mercado Pago e
    # encaminhado somente como header de risco. Nunca é persistido nem registrado em log.
    risk_headers={'X-meli-session-id':str(device_id or '').strip()} if str(device_id or '').strip() else None
    recurring={'frequency':cfg['frequency'],'frequency_type':cfg['frequency_type'],'transaction_amount':amount,'currency_id':'BRL'}
    sub=mp_request('POST','/preapproval',{'reason':f"Fidelizaê! {plan.title()} • {cfg['label']}",'external_reference':reference,'payer_email':payer_email,'auto_recurring':recurring,'back_url':base+'/signup/payment-return','status':'pending'},extra_headers=risk_headers)
    _mp_subscription_diagnostic(sub,'MP_CREATE_RESPONSE')
    # Uma leitura imediata do recurso ajuda a identificar diferenças entre a resposta do POST
    # e o estado efetivamente persistido pelo Mercado Pago. Falhas aqui não quebram o checkout.
    sub_id=str(sub.get('id') or '').strip()
    if sub_id:
        try:
            fresh=mp_request('GET','/preapproval/'+urllib.parse.quote(sub_id,safe=''))
            _mp_subscription_diagnostic(fresh,'MP_CREATE_STATE')
        except Exception as exc:
            print('[BILLING] MP_CREATE_STATE_UNAVAILABLE type=%s' % type(exc).__name__,flush=True)
    return sub


def _mp_timestamp(value):
    if not value:return None
    try:return int(datetime.fromisoformat(str(value).replace('Z','+00:00')).timestamp())
    except Exception:return None


def _mp_status(value):
    status=str(value or '').strip().lower()
    return {'authorized':'active','pending':'pending','paused':'past_due','cancelled':'cancelled'}.get(status,status or 'pending')


def validate_mp_webhook_signature(headers,query):
    """Valida x-signature quando MERCADOPAGO_WEBHOOK_SECRET estiver configurada.

    O Mercado Pago assina o manifesto id:<data.id>;request-id:<x-request-id>;ts:<ts>;.
    Sem secret configurada mantemos compatibilidade com instalações já publicadas; o payload
    ainda é verificado consultando o recurso diretamente na API antes de qualquer ativação.
    """
    secret=os.environ.get('MERCADOPAGO_WEBHOOK_SECRET','').strip()
    if not secret:return True
    signature=str(headers.get('x-signature') or headers.get('X-Signature') or '')
    request_id=str(headers.get('x-request-id') or headers.get('X-Request-Id') or '')
    parts={}
    for item in signature.split(','):
        if '=' in item:
            k,v=item.split('=',1);parts[k.strip()]=v.strip()
    ts=parts.get('ts'); received=parts.get('v1')
    data_id=(query.get('data.id') or query.get('data_id') or [''])[0]
    if not ts or not received:return False
    manifest=''
    if data_id:manifest+='id:'+str(data_id).lower()+';'
    if request_id:manifest+='request-id:'+request_id+';'
    manifest+='ts:'+ts+';'
    expected=hmac.new(secret.encode('utf-8'),manifest.encode('utf-8'),hashlib.sha256).hexdigest()
    return hmac.compare_digest(expected,received)


def send_subscription_welcome(name,email,company,plan,billing_option=None):
    cfg=global_email_config()
    if not email_configured(cfg): return {'sent':False,'reason':'email_not_configured'}
    base=(os.environ.get('PUBLIC_BASE_URL') or 'https://app.fidelizae.com.br').rstrip('/')
    _,bcfg=billing_config(plan,billing_option or ('free' if normalize_plan(plan)=='beginner' else 'monthly'))
    plan_name={'beginner':'Iniciante','intermediate':'Intermediário','pro':'PRO'}.get(normalize_plan(plan),str(plan))
    label=f'{plan_name} — {bcfg["label"]}'
    msg=EmailMessage();msg['Subject']='Seu Fidelizaê! está pronto 🎉';msg['To']=email
    msg.set_content(f'Olá, {name}!\n\nA empresa {company} foi ativada no plano {label}.\nAcesse: {base}/login\nUsuário: {email}\nUse a senha criada no cadastro.\n\nFidelizaê! — Fidelidade que marca pontos.')
    return send_email_message(msg,cfg)


def provision_signup(conn,row,subscription=None):
    if row['status']=='active': return None
    company=conn.execute("SELECT id FROM companies ORDER BY id LIMIT 1").fetchone()
    if not company: raise RuntimeError('base_company_missing')
    company_id=company['id']; code=('AUTO'+secrets.token_hex(4)).upper()
    plan=normalize_plan(row['plan']); loyalty='stamps' if plan=='beginner' else (row['loyalty_type'] or 'stamps')
    sub=subscription or {}; now=now_ts(); next_ts=_mp_timestamp(sub.get('next_payment_date'))
    billing_option=normalize_billing_option(plan,row['billing_option'] if 'billing_option' in row.keys() else ('free' if plan=='beginner' else 'monthly'))
    _,bcfg=billing_config(plan,billing_option); commitment_until=(now+bcfg['commitment_days']*86400) if bcfg['commitment_days'] else None
    cid=insert_id(conn,"INSERT INTO campaigns(company_id,code,name,reward_name,goal,icon,card_theme,plan,loyalty_type,points_spend_cents,logo_image,subscription_provider,subscription_id,subscription_status,subscription_started_at,subscription_current_period_end,subscription_next_payment_at,subscription_status_updated_at,billing_option,billing_amount,commitment_until,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(company_id,code,row['company_name'],'Recompensa do programa',5,'__LOGO__','orange',plan,loyalty,200,row['logo_image'] if 'logo_image' in row.keys() else None,'mercadopago' if plan!='beginner' else 'free',row['subscription_id'], 'active',now,next_ts,next_ts,now,billing_option,bcfg['amount'],commitment_until,now))
    uid=insert_id(conn,"INSERT INTO users(company_id,name,email,password_hash,role,active,is_client_admin,campaign_id,created_at) VALUES(?,?,?,?,?,?,?,?,?)",(company_id,row['responsible_name'],row['email'],row['password_hash'],'attendant',1,1,cid,now))
    conn.execute("UPDATE subscription_signups SET status='active',provisioned_at=? WHERE id=?",(now,row['id']))
    conn.execute("UPDATE legal_acceptances SET campaign_id=COALESCE(campaign_id,?),user_id=COALESCE(user_id,?),email=COALESCE(email,?) WHERE signup_id=?",(cid,uid,row['email'],row['id']))
    audit(conn,company_id,uid,'subscription_signup','campaign',cid,details=plan)
    try: send_subscription_welcome(row['responsible_name'],row['email'],row['company_name'],plan,billing_option)
    except Exception as exc: print('[BILLING] welcome email failed',exc)
    return cid


def _best_effort_cancel_subscription(subscription_id):
    if not subscription_id:return True
    try:
        mp_request('PUT','/preapproval/'+urllib.parse.quote(str(subscription_id),safe=''),{'status':'cancelled'})
        return True
    except Exception as exc:
        print('[BILLING] cancel previous subscription failed',subscription_id,exc)
        return False


def _approved_subscription_invoice_count(subscription_id):
    """Conta somente faturas da assinatura cujo pagamento foi efetivamente aprovado."""
    if not subscription_id:return 0
    try:
        data=mp_request('GET','/authorized_payments/search?preapproval_id='+urllib.parse.quote(str(subscription_id),safe=''))
    except Exception as exc:
        print('[BILLING] invoice count unavailable type=%s' % type(exc).__name__,flush=True)
        return None
    count=0
    for invoice in (data.get('results') or []):
        payment=invoice.get('payment') if isinstance(invoice,dict) else None
        if isinstance(payment,dict) and str(payment.get('status') or '').lower()=='approved':count+=1
    return count


def _refresh_annual_commitment_after_payment(conn,campaign):
    """Abre um novo compromisso anual somente após uma cobrança aprovada já no novo ciclo."""
    if not campaign or campaign['subscription_cancel_at_period_end']: return False
    option=normalize_billing_option(campaign['plan'],campaign['billing_option'])
    if option not in ('annual_monthly','annual_upfront'): return False
    end=int(campaign['commitment_until'] or 0)
    now=now_ts()
    if not end or now < end: return False
    new_end=end
    while new_end<=now: new_end+=365*86400
    conn.execute('UPDATE campaigns SET commitment_until=?,subscription_status_updated_at=? WHERE id=?',(new_end,now,campaign['id']))
    audit(conn,campaign['company_id'],None,'subscription_annual_renewed','campaign',campaign['id'],details=f'{option};commitment_until={new_end}')
    return True


def _expire_non_renewing_campaign(conn,campaign_id):
    """Encerra o acesso quando termina o período já pago, sem apagar dados históricos."""
    now=now_ts()
    conn.execute("UPDATE campaigns SET active=0,subscription_status='expired',subscription_next_payment_at=NULL,subscription_status_updated_at=? WHERE id=?",(now,campaign_id))
    conn.execute('DELETE FROM sessions WHERE user_id IN (SELECT id FROM users WHERE campaign_id=?)',(campaign_id,))


def reconcile_campaign_billing(conn,campaign_id,allow_remote=True):
    """Aplica mudanças agendadas somente quando o ciclo pago realmente terminou.

    Upgrades usam uma nova assinatura pendente e são concluídos pelo webhook após autorização.
    Downgrades pagos mantêm o plano atual até a próxima cobrança; no vencimento, confirmamos
    na API que a assinatura segue autorizada e que o próximo vencimento avançou. Downgrade
    para o plano grátis é aplicado no fim do período já pago, mesmo com a assinatura cancelada.
    """
    c=conn.execute('SELECT * FROM campaigns WHERE id=?',(campaign_id,)).fetchone()
    if not c:return None
    now=now_ts()
    # Cancelamento de renovação preserva o acesso pelo período já pago/contratado.
    if c['subscription_cancel_at_period_end']:
        option=normalize_billing_option(normalize_plan(c['plan']),c['billing_option'] if 'billing_option' in c.keys() else 'monthly')
        access_until=int((c['commitment_until'] if option in ('annual_monthly','annual_upfront') else c['subscription_current_period_end']) or 0)
        if option=='annual_monthly' and allow_remote and c['subscription_id']:
            paid=_approved_subscription_invoice_count(c['subscription_id'])
            if paid is not None and paid>=12:
                if _best_effort_cancel_subscription(c['subscription_id']):
                    conn.execute("UPDATE campaigns SET subscription_status='non_renewing',subscription_next_payment_at=NULL,subscription_status_updated_at=? WHERE id=?",(now,campaign_id))
        if access_until and now>=access_until:
            _expire_non_renewing_campaign(conn,campaign_id)
            return conn.execute('SELECT * FROM campaigns WHERE id=?',(campaign_id,)).fetchone()
        c=conn.execute('SELECT * FROM campaigns WHERE id=?',(campaign_id,)).fetchone()
    # Limpeza de uma assinatura anterior que não pôde ser cancelada no instante do upgrade.
    if allow_remote and c['previous_subscription_id']:
        if _best_effort_cancel_subscription(c['previous_subscription_id']):
            conn.execute('UPDATE campaigns SET previous_subscription_id=NULL WHERE id=?',(campaign_id,))
            c=conn.execute('SELECT * FROM campaigns WHERE id=?',(campaign_id,)).fetchone()
    pending=normalize_plan(c['pending_plan']) if c['pending_plan'] else None
    effective=int(c['subscription_current_period_end'] or 0)
    if not pending or c['pending_subscription_id'] or not effective or now < effective:return c
    if pending=='beginner':
        conn.execute("UPDATE campaigns SET plan='beginner',pending_plan=NULL,subscription_provider='free',subscription_id=NULL,subscription_status='active',subscription_cancel_at_period_end=0,subscription_current_period_end=NULL,subscription_next_payment_at=NULL,subscription_status_updated_at=? WHERE id=?",(now,campaign_id))
        return conn.execute('SELECT * FROM campaigns WHERE id=?',(campaign_id,)).fetchone()
    if not allow_remote or not c['subscription_id']:return c
    try: sub=mp_request('GET','/preapproval/'+urllib.parse.quote(c['subscription_id'],safe=''))
    except Exception:return c
    mapped=_mp_status(sub.get('status')); next_ts=_mp_timestamp(sub.get('next_payment_date'))
    conn.execute('UPDATE campaigns SET subscription_status=?,subscription_next_payment_at=?,subscription_status_updated_at=? WHERE id=?',(mapped,next_ts,now,campaign_id))
    if mapped=='active' and next_ts and next_ts>effective:
        conn.execute('UPDATE campaigns SET plan=?,pending_plan=NULL,subscription_cancel_at_period_end=0,subscription_current_period_end=?,subscription_next_payment_at=? WHERE id=?',(pending,next_ts,next_ts,campaign_id))
    return conn.execute('SELECT * FROM campaigns WHERE id=?',(campaign_id,)).fetchone()

CARD_THEMES={
    'green':('#174f3f','#082b25'),
    'orange':('#d18a1f','#8a4907'),
    'blue':('#183f6d','#091f3b'),
    'red':('#7c2028','#3b0b11'),
    'black':('#292929','#070707'),
}

def card_theme_colors(theme):
    return CARD_THEMES.get(str(theme or 'green').strip().lower(), CARD_THEMES['green'])


def send_customer_welcome_email(name, email, client_name, public_id, campaign, email_config=None):
    """Envia o cartão recém-criado usando exclusivamente a integração do cliente."""
    if not email_configured(email_config):
        return {'sent':False,'reason':'email_provider_not_configured','skipped':True}
    base_url=(os.environ.get('CLUBE_PUBLIC_URL') or 'https://app.fidelizae.com.br').strip().rstrip('/')
    card_url=f'{base_url}/card?id={urllib.parse.quote(public_id)}'
    card_code=f'CLUBE:{public_id}'
    qr_url=f'{base_url}/api/qr?data={urllib.parse.quote(card_code, safe="")}'
    client=html.escape(str(client_name or ''))
    customer=html.escape(str(name or ''))
    safe_card_url=html.escape(card_url, quote=True)
    safe_code=html.escape(card_code)
    goal=int(campaign.get('goal') or 5) if hasattr(campaign,'get') else int(campaign['goal'] or 5)
    icon=(campaign.get('icon') if hasattr(campaign,'get') else campaign['icon']) or '●'
    stamp='●' if icon=='__LOGO__' else html.escape(str(icon))
    cols=3 if goal==3 else (4 if goal==8 else 5)
    theme=(campaign.get('card_theme') if hasattr(campaign,'get') else campaign['card_theme']) or 'green'
    theme_start,theme_end=card_theme_colors(theme)
    cells=[f'<td style="padding:5px"><div style="width:42px;height:42px;border:2px solid #d7c6bb;border-radius:50%;display:flex;align-items:center;justify-content:center;filter:grayscale(1);opacity:.5;font-size:20px">{stamp}</div></td>' for _ in range(goal)]
    rows=['<tr>'+''.join(cells[i:i+cols])+'</tr>' for i in range(0,len(cells),cols)]
    wallet_buttons=[]
    apple=apple_pass_link(public_id)
    google=google_wallet_link(public_id)
    if apple: wallet_buttons.append(f'<a href="{html.escape(base_url+apple,quote=True)}" style="display:inline-block;padding:11px 16px;margin:4px;background:#231a16;color:#fff;text-decoration:none;border-radius:10px;font-weight:700">Apple Wallet</a>')
    if google: wallet_buttons.append(f'<a href="{html.escape(base_url+google,quote=True)}" style="display:inline-block;padding:11px 16px;margin:4px;background:#231a16;color:#fff;text-decoration:none;border-radius:10px;font-weight:700">Google Wallet</a>')
    if not wallet_buttons:
        wallet_buttons.append(f'<a href="{safe_card_url}" style="display:inline-block;padding:11px 16px;margin:4px;background:#231a16;color:#fff;text-decoration:none;border-radius:10px;font-weight:700">Abrir cartão / Wallet</a>')
    text=(f'Agora você faz parte do programa de fidelidade {client_name}.\n'
          'Para ter acesso às nossas vantagens, apresente o seu cartão com o QR code aos nossos atendentes toda vez que vier efetuar uma compra.\n\n'
          f'Link de acesso: {card_url}\nCódigo do cartão: {card_code}\n')
    html_body=(
        '<!doctype html><html><body style="margin:0;background:#f7f3ef;font-family:Arial,sans-serif;color:#231a16">'
        '<div style="max-width:620px;margin:0 auto;padding:28px 18px">'
        f'<h2 style="margin:0 0 12px">Agora você faz parte do programa de fidelidade {client}.</h2>'
        '<p style="line-height:1.55;margin:0 0 24px">Para ter acesso às nossas vantagens, apresente o seu cartão com o QR code aos nossos atendentes toda vez que vier efetuar uma compra.</p>'
        f'<div style="background:linear-gradient(145deg,{theme_start},{theme_end});color:#fff;border-radius:26px;padding:26px;text-align:center">'
        '<div style="font-weight:900;letter-spacing:.08em;font-size:14px">CLUBE DE FIDELIDADE</div>'
        f'<h2 style="margin:18px 0 6px">{client}</h2><p style="margin:0 0 12px">{customer}</p>'
        f'<table role="presentation" align="center" cellspacing="0" cellpadding="0" style="margin:10px auto 16px">{"".join(rows)}</table>'
        f'<div style="margin:16px auto;padding:10px 12px;border:1px solid rgba(255,255,255,.25);border-radius:12px"><div style="font-size:11px;opacity:.75;text-transform:uppercase">Código do cartão</div><strong>{safe_code}</strong></div>'
        f'<img src="{html.escape(qr_url,quote=True)}" alt="QR code do cartão" width="170" height="170" style="display:block;background:#fff;border-radius:14px;padding:10px;margin:16px auto"></div>'
        f'<div style="text-align:center;margin:18px 0">{"".join(wallet_buttons)}</div>'
        f'<p style="line-height:1.55"><strong>Link de acesso:</strong><br><a href="{safe_card_url}">{html.escape(card_url)}</a></p>'
        f'<p style="line-height:1.55"><strong>Código do cartão:</strong><br>{safe_code}</p></div></body></html>'
    )
    msg=EmailMessage()
    msg['Subject']=f'Bem-vindo ao programa de fidelidade {client_name}'
    msg['To']=email
    msg.set_content(text)
    msg.add_alternative(html_body, subtype='html')
    result=send_email_message(msg,email_config)
    if not result.get('sent'):
        print(f'[EMAIL] CUSTOMER_WELCOME_FAILED email={email} campaign={client_name!r} reason={result.get("reason")}')
    return result


def send_attendant_welcome_email(name, email, client_name, smtp_config=None):
    if not email_configured(smtp_config):
        return {'sent':False,'reason':'smtp_not_configured'}
    login_url=os.environ.get('CLUBE_LOGIN_URL','https://app.fidelizae.com.br/login').strip()
    msg=EmailMessage()
    msg['Subject']='Acesso ao Fidelizaê!'
    msg['To']=email
    msg.set_content(
        'Cadastro realizado com sucesso! Agora é só acessar o link abaixo, inserir seu e-mail e senha para ter acesso ao painel do seu Fidelizaê!.\n\n'
        f'{login_url}\n\n'
        f'E-mail: {email}\n'
        'Use a senha inicial fornecida pelo administrador. Por segurança, ela não é enviada por e-mail.\n'
        f'Cliente: {client_name}\n'
    )
    result=send_email_message(msg, smtp_config)
    if not result.get('sent'):
        print(f'[EMAIL] ATTENDANT_WELCOME_FAILED user={_email_tag(email)} reason={result.get("reason")}')
    return result


def whatsapp_config_for_client(conn=None,campaign_id=None):
    # WhatsApp sempre isolado por cliente. Vazio = integração não configurada.
    if conn is not None and campaign_id:
        x=client_integrations(conn,campaign_id)
        if x:
            return {'phone_number_id':x.get('whatsapp_phone_number_id') or '',
                    'waba_id':x.get('whatsapp_waba_id') or '',
                    'token':x.get('whatsapp_access_token') or '',
                    'version':x.get('whatsapp_api_version') or 'v24.0','source':'client'}
    return {'phone_number_id':'','waba_id':'','token':'','version':'v24.0','source':'client'}

def whatsapp_cloud_configured(config=None):
    c=config or whatsapp_config_for_client()
    return all(c.get(k,'') for k in ('token','phone_number_id','version'))

def _normalize_phone(value):
    return re.sub(r'\D','',str(value or ''))[:20]

def whatsapp_meta_test_config():
    # Credenciais exclusivas do numero de teste fornecido pela Meta.
    # Nunca sao usadas por automacoes/campanhas de producao; apenas pelo endpoint ENVIAR TESTE.
    return {
        'phone_number_id':(os.environ.get('META_TEST_WHATSAPP_PHONE_NUMBER_ID') or '').strip(),
        'waba_id':(os.environ.get('META_TEST_WHATSAPP_WABA_ID') or '').strip(),
        'token':(os.environ.get('META_TEST_WHATSAPP_ACCESS_TOKEN') or '').strip(),
        'version':(os.environ.get('META_GRAPH_VERSION') or 'v24.0').strip() or 'v24.0',
        'source':'meta_test'
    }

def whatsapp_meta_test_recipients():
    raw=(os.environ.get('META_TEST_WHATSAPP_RECIPIENTS') or '').strip()
    return {_normalize_phone(x) for x in re.split(r'[,;\n]+',raw) if _normalize_phone(x)}

def whatsapp_meta_test_configured():
    return whatsapp_cloud_configured(whatsapp_meta_test_config()) and bool(whatsapp_meta_test_recipients())

def whatsapp_test_delivery_config(conn,campaign_id,phone):
    client_cfg=whatsapp_config_for_client(conn,campaign_id)
    if whatsapp_cloud_configured(client_cfg):
        return client_cfg,'production'
    test_cfg=whatsapp_meta_test_config()
    if whatsapp_cloud_configured(test_cfg) and _normalize_phone(phone) in whatsapp_meta_test_recipients():
        return test_cfg,'meta_test'
    return None,'unavailable'

def send_whatsapp_cloud(phone, message, config=None):
    c=config or whatsapp_config_for_client()
    version=c.get('version',''); phone_number_id=c.get('phone_number_id',''); token=c.get('token','')
    if not (version and phone_number_id and token): raise RuntimeError('whatsapp_not_configured')
    url=f'https://graph.facebook.com/{urllib.parse.quote(version)}/{urllib.parse.quote(phone_number_id)}/messages'
    body=json.dumps({'messaging_product':'whatsapp','recipient_type':'individual','to':str(phone),'type':'text',
        'text':{'preview_url':False,'body':str(message)}}).encode('utf-8')
    req=urllib.request.Request(url,data=body,method='POST',headers={'Authorization':'Bearer '+token,'Content-Type':'application/json'})
    try:
        with urllib.request.urlopen(req,timeout=15) as resp: return json.loads(resp.read().decode('utf-8') or '{}')
    except urllib.error.HTTPError as exc:
        raw=exc.read().decode('utf-8',errors='replace')
        try: detail=json.loads(raw)
        except Exception: detail={'error':raw[:500]}
        raise RuntimeError(json.dumps(detail,ensure_ascii=False)) from exc

def meta_public_base_url():
    base=(os.environ.get('PUBLIC_BASE_URL') or '').strip().rstrip('/')
    if not base:
        domain=(os.environ.get('RAILWAY_PUBLIC_DOMAIN') or '').strip().strip('/')
        if domain:
            base='https://'+domain
    return base

def meta_callback_url():
    base=meta_public_base_url()
    return (base+'/auth/meta/callback') if base else ''

def meta_embedded_signup_configured():
    return bool(os.environ.get('META_APP_ID','').strip() and os.environ.get('META_APP_SECRET','').strip() and os.environ.get('META_CONFIG_ID','').strip())

def meta_exchange_code(code):
    app_id=os.environ.get('META_APP_ID','').strip(); secret=os.environ.get('META_APP_SECRET','').strip()
    if not (app_id and secret and code): raise RuntimeError('meta_embedded_signup_not_configured')
    qs=urllib.parse.urlencode({'client_id':app_id,'client_secret':secret,'code':code})
    req=urllib.request.Request('https://graph.facebook.com/v24.0/oauth/access_token?'+qs,method='GET')
    with urllib.request.urlopen(req,timeout=20) as resp: return json.loads(resp.read().decode('utf-8') or '{}')

def meta_phone_details(phone_id,token):
    req=urllib.request.Request(f'https://graph.facebook.com/v24.0/{urllib.parse.quote(str(phone_id))}?fields=id,display_phone_number,verified_name',headers={'Authorization':'Bearer '+token})
    with urllib.request.urlopen(req,timeout=20) as resp: return json.loads(resp.read().decode('utf-8') or '{}')



def _qr_secret():
    secret=(os.environ.get('CLUBE_QR_SECRET') or os.environ.get('CLUBE_ENCRYPTION_KEY') or '').strip()
    env=(os.environ.get('APP_ENV') or 'production').strip().lower()
    if not secret:
        if env in {'development','dev','test','testing','staging','sandbox'}: secret='development-only-change-me'
        else: raise RuntimeError('qr_secret_not_configured')
    return secret.encode('utf-8')

def make_dynamic_qr(public_id, ttl=60):
    exp=now_ts()+max(30,min(int(ttl),120)); nonce=secrets.token_urlsafe(6)
    body=f'{public_id}.{exp}.{nonce}'
    sig=hmac.new(_qr_secret(),body.encode(),hashlib.sha256).hexdigest()[:32]
    return f'DYN.{body}.{sig}',exp

def resolve_member_token(raw):
    token=str(raw or '').strip()
    if token.startswith('CLUBE:'): token=token[6:]
    if token.startswith('DYN.'):
        parts=token.split('.')
        if len(parts)!=5:return None,'invalid_dynamic_qr'
        _,public_id,exp,nonce,sig=parts
        try:exp=int(exp)
        except ValueError:return None,'invalid_dynamic_qr'
        if exp<now_ts():return None,'qr_expired'
        body=f'{public_id}.{exp}.{nonce}'
        expected=hmac.new(_qr_secret(),body.encode(),hashlib.sha256).hexdigest()[:32]
        if not hmac.compare_digest(sig,expected):return None,'invalid_dynamic_qr'
        return public_id,None
    return token,None

def enqueue_message(conn, campaign_id, kind, recipient, payload, delay=0):
    stored_recipient=recipient; recipient_hash=None
    if kind=='whatsapp' and recipient:
        stored_recipient=encrypt_pii(recipient,'phone')
        recipient_hash=pii_lookup_hash(recipient,'phone')
    return insert_id(conn,"INSERT INTO message_queue(campaign_id,kind,recipient,recipient_hash,payload_json,status,attempts,available_at,created_at) VALUES(?,?,?,?,?,?,?,?,?)",(campaign_id,kind,stored_recipient,recipient_hash,json.dumps(payload,ensure_ascii=False),'pending',0,now_ts()+delay,now_ts()))

def _queue_send(item, conn):
    payload=json.loads(item['payload_json'] or '{}'); kind=item['kind']; campaign_id=item['campaign_id']
    if kind=='customer_welcome':
        c=conn.execute('SELECT * FROM campaigns WHERE id=?',(campaign_id,)).fetchone();
        return send_customer_welcome_email(payload['name'],payload['email'],c['name'],payload['public_id'],rowdict(c),email_config_for_client(conn,campaign_id))
    if kind=='campaign_email':
        return send_campaign_email(item['recipient'],payload.get('name',''),payload.get('message',''),payload.get('image_data'),payload.get('subject','Mensagem do Fidelizaê!'),email_config_for_client(conn,campaign_id))
    if kind=='whatsapp':
        try:
            response=send_whatsapp_cloud(item['recipient'],payload.get('message',''),whatsapp_config_for_client(conn,campaign_id))
            return {'sent':True,'message_id':((response.get('messages') or [{}])[0]).get('id')}
        except Exception as exc:return {'sent':False,'reason':str(exc)[:500]}
    if kind=='attendant_welcome':
        return send_attendant_welcome_email(payload['name'],item['recipient'],payload['client_name'],global_email_config())
    if kind=='password_recovery':
        return send_password_recovery_email(item['recipient'],payload['token'],global_email_config())
    if kind=='outbound_webhook':
        delivery_id=int(payload.get('delivery_id') or 0); webhook_id=int(payload.get('webhook_id') or 0)
        hook=conn.execute("SELECT secret_enc,active FROM webhook_subscriptions WHERE id=? AND campaign_id=?",(webhook_id,campaign_id)).fetchone()
        if not hook or not hook['active']:
            return {'sent':False,'reason':'webhook_inactive'}
        secret=decrypt_secret(hook['secret_enc'])
        if not secret:
            return {'sent':False,'reason':'webhook_secret_unavailable'}
        body=jdump(payload.get('payload') or {})
        req=urllib.request.Request(item['recipient'],data=body,method='POST',headers={
            'Content-Type':'application/json','User-Agent':'Fidelizae-Webhooks/1.0',
            'X-Fidelizae-Event':str((payload.get('payload') or {}).get('event') or ''),
            'X-Fidelizae-Signature':'sha256='+_webhook_signature(secret,body)
        })
        try:
            with urllib.request.urlopen(req,timeout=12) as resp:
                status=int(getattr(resp,'status',200) or 200)
            conn.execute("UPDATE webhook_deliveries SET status='delivered',http_status=?,attempts=attempts+1,delivered_at=? WHERE id=?",(status,now_ts(),delivery_id))
            return {'sent':True,'http_status':status}
        except urllib.error.HTTPError as exc:
            conn.execute("UPDATE webhook_deliveries SET status='failed',http_status=?,attempts=attempts+1,last_error=? WHERE id=?",(int(exc.code),str(exc)[:500],delivery_id))
            return {'sent':False,'reason':'http_'+str(exc.code)}
        except Exception as exc:
            conn.execute("UPDATE webhook_deliveries SET status='failed',attempts=attempts+1,last_error=? WHERE id=?",(str(exc)[:500],delivery_id))
            return {'sent':False,'reason':str(exc)[:500]}
    return {'sent':False,'reason':'unknown_queue_kind'}

def process_message_queue_once(limit=15):
    with connect(DB_PATH) as conn:
        now=now_ts()
        conn.execute("UPDATE message_queue SET status='retry',available_at=? WHERE status='processing' AND available_at<=?",(now,now))
        rows=conn.execute("SELECT id FROM message_queue WHERE status IN ('pending','retry') AND available_at<=? ORDER BY id LIMIT ?",(now,limit)).fetchall()
        for candidate in rows:
            item_id=candidate['id']; lease_until=now_ts()+120
            cur=conn.execute("UPDATE message_queue SET status='processing',attempts=attempts+1,available_at=? WHERE id=? AND status IN ('pending','retry') AND available_at<=?",(lease_until,item_id,now_ts()))
            if getattr(cur,'rowcount',0)!=1:
                conn.rollback(); continue
            conn.commit()
            row=conn.execute('SELECT * FROM message_queue WHERE id=?',(item_id,)).fetchone()
            if not row: continue
            item=queue_rowdict(row)
            try: result=_queue_send(item,conn)
            except Exception as exc: result={'sent':False,'reason':type(exc).__name__+':'+str(exc)[:300]}
            if result.get('sent'):
                conn.execute("UPDATE message_queue SET status='sent',sent_at=?,last_error=NULL WHERE id=? AND status='processing'",(now_ts(),item_id))
            else:
                attempts=int(item.get('attempts') or 0); status='failed' if attempts>=4 else 'retry'; delay=min(900,30*(2**max(0,attempts-1)))
                conn.execute("UPDATE message_queue SET status=?,last_error=?,available_at=? WHERE id=? AND status='processing'",(status,result.get('reason','failed'),now_ts()+delay,item_id))
            conn.commit()

def customer_segment(last_activity,created_at,visits,reward_ready,almost_reward,now=None):
    now=now or now_ts(); age=max(0,now-int(last_activity or created_at or now))
    if reward_ready: return 'reward_ready'
    if almost_reward: return 'almost_reward'
    if age>=90*86400: return 'inactive90'
    if age>=60*86400: return 'inactive60'
    if age>=30*86400: return 'at_risk'
    if int(visits or 0)>=12: return 'vip'
    if int(visits or 0)<=1 and now-int(created_at or now)<=30*86400: return 'new'
    return 'active'

def segment_sql(segment,cid,loyalty_type='stamps'):
    now=now_ts(); extra=''; params=[]
    last="COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at)"
    visits="(SELECT COUNT(*) FROM transactions t WHERE t.membership_id=m.id AND ((c.loyalty_type='points' AND t.type='adjustment' AND t.value>0) OR (c.loyalty_type='stamps' AND t.type='stamp' AND t.value>0)))"
    if segment=='birthdays': extra=" AND substr(cu.birth_date,6,2)=?"; params=[datetime.now(ZoneInfo('America/Sao_Paulo')).strftime('%m')]
    elif segment=='at_risk': extra=f" AND {last}<? AND {last}>=?"; params=[now-30*86400,now-60*86400]
    elif segment=='inactive60': extra=f" AND {last}<? AND {last}>=?"; params=[now-60*86400,now-90*86400]
    elif segment=='inactive90': extra=f" AND {last}<?"; params=[now-90*86400]
    elif segment=='vip': extra=f" AND {visits}>=12"
    elif segment=='recurrent': extra=f" AND {visits}>=3 AND {visits}<12 AND {last}>=?"; params=[now-45*86400]
    elif segment=='inactive': extra=f" AND {last}<?"; params=[now-90*86400]
    elif segment=='new': extra=" AND m.created_at>=?"; params=[now-30*86400]
    elif segment=='reward_ready': extra=" AND (m.rewards_available>0 OR (c.loyalty_type='points' AND m.points_balance>=(SELECT COALESCE(MIN(points_cost),999999999) FROM reward_catalog WHERE campaign_id=m.campaign_id AND active=1)))"
    elif segment=='almost_reward': extra=" AND ((c.loyalty_type='stamps' AND m.progress=c.goal-1) OR (c.loyalty_type='points' AND EXISTS(SELECT 1 FROM reward_catalog r WHERE r.campaign_id=m.campaign_id AND r.active=1 AND r.points_cost>m.points_balance AND r.points_cost-m.points_balance<=GREATEST(1,CAST(r.points_cost*0.15 AS INTEGER)))))" if str(DB_PATH).startswith(('postgres://','postgresql://')) else " AND ((c.loyalty_type='stamps' AND m.progress=c.goal-1) OR (c.loyalty_type='points' AND EXISTS(SELECT 1 FROM reward_catalog r WHERE r.campaign_id=m.campaign_id AND r.active=1 AND r.points_cost>m.points_balance AND r.points_cost-m.points_balance<=MAX(1,CAST(r.points_cost*0.15 AS INTEGER)))))"
    return extra,params

def campaign_recipient_rows(conn,cid,segment):
    if segment in ('all','birthdays','inactive60','inactive90'):
        extra,params=segment_sql(segment,cid)
        return [customer_rowdict(r) for r in conn.execute("""SELECT m.id membership_id,m.public_id,cu.id customer_id,cu.name,cu.email,cu.phone,cu.phone_enc,cu.marketing_email,cu.marketing_whatsapp
          FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id
          WHERE m.campaign_id=? AND m.status='active' """+extra+' ORDER BY cu.name',(cid,*params)).fetchall()]
    rows=conn.execute("""SELECT m.id membership_id,m.public_id,m.campaign_id,m.created_at,m.progress,m.points_balance,m.rewards_available,
                               c.goal,c.loyalty_type,cu.id customer_id,cu.name,cu.email,cu.phone,cu.phone_enc,cu.marketing_email,cu.marketing_whatsapp
                        FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id
                        WHERE m.campaign_id=? AND m.status='active' ORDER BY cu.name""",(cid,)).fetchall()
    out=[]
    for row in rows:
        d=customer_rowdict(row)
        intel=customer_intelligence(conn,{**d,'id':d['membership_id']},d)
        if intel['segment']==segment:
            out.append(d)
    return out

def ensure_automation_defaults(conn,campaign_id):
    try:
        campaign=conn.execute('SELECT loyalty_type FROM campaigns WHERE id=?',(campaign_id,)).fetchone()
        loyalty_type=(campaign['loyalty_type'] if campaign else 'stamps') or 'stamps'
    except Exception:
        # Mantém compatibilidade com testes/bancos legados onde apenas
        # automation_rules existe neste contexto isolado.
        loyalty_type='stamps'
    defaults={
      'birthday':('both','Feliz aniversário, {nome}! O {cliente} deseja um dia especial para você.'),
      'inactive30':('both','Sentimos sua falta, {nome}! Volte ao {cliente} e continue acumulando no seu programa.'),
      'inactive60':('both','Já faz um tempo, {nome}. Temos saudades de você no {cliente}. Volte e continue aproveitando seus benefícios.'),
      'one_to_reward':('both','Falta só 1 selo, {nome}! Sua recompensa no {cliente} está quase lá.'),
      'reward_available':('both','Parabéns, {nome}! Você já tem uma recompensa disponível no {cliente}.')}
    if loyalty_type!='stamps':
        defaults.pop('one_to_reward',None)
        conn.execute("UPDATE automation_rules SET enabled=0 WHERE campaign_id=? AND rule_type='one_to_reward'",(campaign_id,))
    for rule,(channel,msg) in defaults.items():
        # Compatibilidade com bancos de produção antigos: algumas instalações
        # podem ter automation_rules sem a constraint UNIQUE adicionada depois.
        # Evitamos ON CONFLICT(campaign_id,rule_type), que falha no PostgreSQL
        # quando essa constraint não existe e deixa a transação abortada.
        exists=conn.execute('SELECT id FROM automation_rules WHERE campaign_id=? AND rule_type=? LIMIT 1',(campaign_id,rule)).fetchone()
        if not exists:
            conn.execute('INSERT INTO automation_rules(campaign_id,rule_type,channel,enabled,message,created_at) VALUES(?,?,?,?,?,?)',(campaign_id,rule,channel,0,msg,now_ts()))

def render_test_template(body, campaign, customer_name='Cliente Teste'):
    values={
      '{nome}':customer_name,
      '{empresa}':campaign.get('name') or 'Empresa',
      '{cliente}':campaign.get('name') or 'Empresa',
      '{selos}':'3',
      '{meta}':str(campaign.get('goal') or 10),
      '{recompensa}':campaign.get('reward_name') or 'sua recompensa'
    }
    out=str(body or '')
    for key,value in values.items(): out=out.replace(key,str(value))
    return out

def run_automations_once():
    today=datetime.now(ZoneInfo('America/Sao_Paulo')).date(); now=now_ts()
    with connect(DB_PATH) as conn:
        campaigns=conn.execute('SELECT id,name,loyalty_type FROM campaigns WHERE active=1').fetchall()
        for c in campaigns: ensure_automation_defaults(conn,c['id'])
        rules=conn.execute('SELECT r.*,c.name client_name,c.loyalty_type FROM automation_rules r JOIN campaigns c ON c.id=r.campaign_id WHERE r.enabled=1 AND c.active=1').fetchall()
        for rule in rules:
            rows=conn.execute('''SELECT m.id membership_id,m.progress,m.rewards_available,m.public_id,m.created_at membership_created,c.goal,cu.id customer_id,cu.name,cu.email,cu.phone,cu.phone_enc,cu.birth_date,cu.marketing_email,cu.marketing_whatsapp,
              COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at) last_activity
              FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND m.status='active' ''',(rule['campaign_id'],)).fetchall()
            for raw_x in rows:
                x=customer_rowdict(raw_x)
                match=False; period=''
                if rule['rule_type']=='birthday' and x['birth_date'] and x['birth_date'][5:10]==today.isoformat()[5:10]: match=True; period=str(today.year)
                elif rule['rule_type']=='inactive30' and x['last_activity']<=now-30*86400 and x['last_activity']>now-60*86400: match=True; period=today.strftime('%Y-%m')
                elif rule['rule_type']=='inactive60' and x['last_activity']<=now-60*86400: match=True; period=today.strftime('%Y-%m')
                elif rule['rule_type']=='one_to_reward' and rule['loyalty_type']=='stamps' and x['progress']==x['goal']-1: match=True; period=f"p{x['progress']}-{today.strftime('%Y-%m')}"
                elif rule['rule_type']=='reward_available' and x['rewards_available']>0: match=True; period=f"r{x['rewards_available']}-{today.strftime('%Y-%m')}"
                if not match: continue
                exists=conn.execute('SELECT id FROM automation_runs WHERE rule_id=? AND membership_id=? AND period_key=?',(rule['id'],x['membership_id'],period)).fetchone()
                if exists: continue
                msg=rule['message'].format(nome=x['name'],cliente=rule['client_name'])
                queued=False; channel=rule['channel']
                if channel in ('email','both') and x['email'] and x['marketing_email'] and email_configured(email_config_for_client(conn,rule['campaign_id'])):
                    enqueue_message(conn,rule['campaign_id'],'campaign_email',x['email'],{'name':x['name'],'message':msg,'subject':'Fidelizaê! • '+rule['client_name']}); queued=True
                if channel in ('whatsapp','both') and x['phone'] and x['marketing_whatsapp'] and whatsapp_cloud_configured(whatsapp_config_for_client(conn,rule['campaign_id'])):
                    enqueue_message(conn,rule['campaign_id'],'whatsapp',x['phone'],{'message':msg}); queued=True
                if queued:
                    conn.execute('INSERT INTO automation_runs(rule_id,membership_id,period_key,created_at) VALUES(?,?,?,?) ON CONFLICT(rule_id,membership_id,period_key) DO NOTHING',(rule['id'],x['membership_id'],period,now_ts()))

def background_loop():
    tick=299
    while True:
        try: process_message_queue_once()
        except Exception as exc: print('[QUEUE]',type(exc).__name__,str(exc)[:300])
        tick+=1
        if tick%30==0:
            try:
                with connect(DB_PATH) as _c: expire_points_once(_c,now_ts())
            except Exception as exc: print('[POINTS_EXPIRY]',type(exc).__name__,str(exc)[:300])
            try: run_automations_once()
            except Exception as exc: print('[AUTOMATION]',type(exc).__name__,str(exc)[:300])
        if tick%300==0:
            try: run_scheduled_r2_backup_once()
            except Exception as exc:
                _R2_BACKUP_STATE.update({'configured':r2_backup_configured(),'last_error':str(exc)[:300]})
                print('[R2_BACKUP]',type(exc).__name__,str(exc)[:300])
        time.sleep(2)

def card_record(conn,public_id):
    row=conn.execute('''SELECT m.public_id,m.progress,m.points_balance,m.rewards_available,m.status,m.created_at,cu.name customer_name,c.name campaign_name,c.code campaign_code,c.reward_name,c.goal,c.icon,c.logo_image,c.card_theme,c.loyalty_type,c.points_spend_cents
      FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=?''',(public_id,)).fetchone()
    return rowdict(row)

def notify_wallet_updates(conn,public_id):
    card=card_record(conn,public_id)
    if not card:return
    try: google_update_object(card)
    except Exception: pass
    regs=conn.execute('''SELECT wr.push_token FROM wallet_registrations wr JOIN memberships m ON m.id=wr.membership_id WHERE m.public_id=?''',(public_id,)).fetchall()
    for r in regs:
        try: apple_push_update(r['push_token'])
        except Exception: pass


def _webhook_signature(secret, body_bytes):
    return hmac.new(secret.encode('utf-8'), body_bytes, hashlib.sha256).hexdigest()

def queue_webhook_event(conn, campaign_id, event_type, data):
    rows=conn.execute("SELECT id,url,secret_enc,events_json FROM webhook_subscriptions WHERE campaign_id=? AND active=1",(campaign_id,)).fetchall()
    for r in rows:
        try: events=json.loads(r['events_json'] or '[]')
        except Exception: events=[]
        if event_type not in events and '*' not in events:
            continue
        payload={'event':event_type,'created_at':now_iso(),'data':data}
        delivery_id=insert_id(conn,"INSERT INTO webhook_deliveries(campaign_id,webhook_id,event_type,status,created_at) VALUES(?,?,?,?,?)",(campaign_id,r['id'],event_type,'pending',now_ts()))
        enqueue_message(conn,campaign_id,'outbound_webhook',r['url'],{'delivery_id':delivery_id,'webhook_id':r['id'],'payload':payload})

def _parse_import_file(filename, data_b64):
    raw=base64.b64decode(data_b64 or '',validate=True)
    if len(raw)>4_000_000: raise ValueError('file_too_large')
    name=str(filename or '').lower()
    rows=[]
    if name.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError('xlsx_support_unavailable') from exc
        wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        ws=wb.active
        values=list(ws.iter_rows(values_only=True))
        if not values:return []
        headers=[str(x or '').strip().lower() for x in values[0]]
        for vals in values[1:]:
            rows.append(dict(zip(headers,vals)))
    else:
        import csv
        text=raw.decode('utf-8-sig',errors='strict')
        sample=text[:2048]
        try: dialect=csv.Sniffer().sniff(sample,delimiters=',;\\t')
        except Exception: dialect=csv.excel
        rows=list(csv.DictReader(io.StringIO(text),dialect=dialect))
    def pick(d,*names):
        norm={re.sub(r'[^a-z0-9]','',str(k).lower()):v for k,v in d.items()}
        for n in names:
            key=re.sub(r'[^a-z0-9]','',n.lower())
            if key in norm:return norm[key]
        return ''
    out=[]
    for i,r in enumerate(rows[:10000],start=2):
        name=str(pick(r,'nome','name') or '').strip()[:120]
        raw_email=str(pick(r,'email','e-mail') or '').strip(); email=normalize_email(raw_email) if raw_email else ''
        phone=str(pick(r,'celular','telefone','phone') or '').strip()[:40]
        raw_cpf=str(pick(r,'cpf') or '').strip(); cpf=normalize_cpf(raw_cpf) if raw_cpf else ''
        birth=str(pick(r,'nascimento','data de nascimento','birth_date','birthdate') or '').strip()[:10]
        initial=str(pick(r,'saldo inicial','saldo','pontos','selos','initial_balance') or '').strip()
        try: initial_balance=max(0,int(float(initial.replace(',','.')))) if initial else 0
        except Exception: initial_balance=0
        errors=[]
        if len(name)<2: errors.append('nome')
        if raw_email and not email: errors.append('email')
        if raw_cpf and not cpf: errors.append('cpf')
        if birth and not re.fullmatch(r'\d{4}-\d{2}-\d{2}',birth): errors.append('nascimento')
        out.append({'row':i,'name':name,'email':email,'phone':phone,'cpf':cpf,'birth_date':birth,'initial_balance':initial_balance,'errors':errors})
    return out


class Handler(BaseHTTPRequestHandler):
    sys_version = ''
    def _need_permission(self,sess,key):
        if has_permission(rowdict(sess),key): return True
        self.send_json({'ok':False,'error':'permission_denied','permission':key},403); return False

    def _rate_ok(self,bucket,limit,window,subject=None,block_seconds=None):
        subject=str(subject or self._ip()).strip().lower()[:180]
        digest=hashlib.sha256(subject.encode('utf-8')).hexdigest()[:32]
        ok,retry=persistent_rate_allow(f'{bucket}:{digest}',limit,window,block_seconds)
        if ok: return True
        self.send_json({'ok':False,'error':'rate_limited','retry_after':retry},429,{'Retry-After':str(max(1,retry))}); return False

    def _api_context(self,conn):
        auth=(self.headers.get('Authorization') or '').strip()
        if not auth.lower().startswith('bearer '):
            self.send_json({'ok':False,'error':'api_key_required'},401); return None
        token=auth[7:].strip()
        if len(token)<24:
            self.send_json({'ok':False,'error':'invalid_api_key'},401); return None
        digest=hashlib.sha256(token.encode()).hexdigest()
        row=conn.execute('''SELECT ak.id key_id,ak.campaign_id,c.company_id,c.name campaign_name,c.loyalty_type,c.goal,c.points_spend_cents,c.points_expiry_days
                            FROM api_keys ak JOIN campaigns c ON c.id=ak.campaign_id
                            WHERE ak.token_hash=? AND ak.active=1 AND c.active=1''',(digest,)).fetchone()
        if not row:
            self.send_json({'ok':False,'error':'invalid_api_key'},401); return None
        if not plan_allows(conn,row['campaign_id'],'advanced'):
            self.send_json({'ok':False,'error':'api_feature_not_available'},403); return None
        conn.execute('UPDATE api_keys SET last_used_at=? WHERE id=?',(now_ts(),row['key_id']))
        return rowdict(row)

    server_version = 'Fidelizae/20.0'

    def log_message(self, fmt, *args):
        print(f'[{self.log_date_time_string()}] {self.address_string()} - {fmt % args}')

    def _cookies(self):
        c = cookies.SimpleCookie()
        c.load(self.headers.get('Cookie', ''))
        return c

    def _session_token(self):
        c = self._cookies().get(SESSION_COOKIE)
        return c.value if c else None

    def _session(self, conn):
        return get_session(conn, self._session_token())

    def _body_json(self):
        n = int(self.headers.get('Content-Length', '0') or 0)
        if n > 8_000_000:
            raise ValueError('body_too_large')
        raw = self.rfile.read(n) if n else b'{}'
        return json.loads(raw.decode('utf-8') or '{}')

    def _body_payload(self):
        n = int(self.headers.get('Content-Length', '0') or 0)
        if n > 8_000_000:
            raise ValueError('body_too_large')
        raw = self.rfile.read(n) if n else b''
        ctype = (self.headers.get('Content-Type') or '').split(';', 1)[0].strip().lower()
        if ctype == 'application/x-www-form-urlencoded':
            parsed = urllib.parse.parse_qs(raw.decode('utf-8'), keep_blank_values=True)
            return {k: (v[-1] if v else '') for k, v in parsed.items()}, 'form'
        if ctype == 'multipart/form-data':
            raise ValueError('multipart_not_supported')
        return json.loads(raw.decode('utf-8') or '{}'), 'json'

    def _security_headers(self, html_response=False):
        self.send_header('X-Content-Type-Options','nosniff')
        current_path = urllib.parse.urlparse(getattr(self, 'path', '') or '').path
        same_origin_embed = current_path in ('/security', '/loyalty360')
        self.send_header('X-Frame-Options','SAMEORIGIN' if same_origin_embed else 'DENY')
        self.send_header('Referrer-Policy','strict-origin-when-cross-origin')
        self.send_header('Permissions-Policy','camera=(self), microphone=(), geolocation=(), payment=()')
        self.send_header('X-Permitted-Cross-Domain-Policies','none')
        # HSTS é opt-in no aplicativo; enquanto o Cloudflare estiver sem HSTS, mantenha desligado.
        if _cookie_secure() and os.environ.get('CLUBE_HSTS_ENABLED','0')=='1':
            try: max_age=max(0,min(int(os.environ.get('CLUBE_HSTS_MAX_AGE','31536000') or 31536000),63072000))
            except ValueError: max_age=31536000
            value=f'max-age={max_age}'
            if os.environ.get('CLUBE_HSTS_INCLUDE_SUBDOMAINS','1')=='1': value+='; includeSubDomains'
            self.send_header('Strict-Transport-Security',value)
        if html_response:
            frame_ancestors = "'self'" if same_origin_embed else "'none'"
            self.send_header('Content-Security-Policy', f"default-src 'self'; base-uri 'self'; object-src 'none'; frame-ancestors {frame_ancestors}; form-action 'self'; script-src 'self' 'unsafe-inline' https://www.mercadopago.com https://cdn.jsdelivr.net https://connect.facebook.net; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; font-src 'self' https://fonts.gstatic.com data:; img-src 'self' data: blob: https:; connect-src 'self' https://api.mercadopago.com https://graph.facebook.com https://www.facebook.com; frame-src 'self' https://www.facebook.com https://web.facebook.com; upgrade-insecure-requests")

    def send_redirect(self, location, status=303, headers=None):
        self.send_response(status)
        self.send_header('Location', location)
        self.send_header('Cache-Control', 'no-store')
        self._security_headers(False)
        if headers:
            for k, v in headers.items():
                if isinstance(v,(list,tuple)):
                    for item in v:self.send_header(k,str(item))
                else:self.send_header(k,str(v))
        self.end_headers()

    def _ip(self):
        direct=_safe_ip(self.client_address[0])
        trust_proxy=os.environ.get('CLUBE_TRUST_PROXY','1' if os.environ.get('RAILWAY_ENVIRONMENT_NAME') or os.environ.get('RAILWAY_PROJECT_ID') else '0')=='1'
        if not trust_proxy:return direct
        forwarded=(self.headers.get('X-Forwarded-For') or '').split(',')[0].strip()
        return _safe_ip(forwarded,direct)

    def send_json(self, obj, status=200, headers=None):
        data = jdump(obj)
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Cache-Control', 'no-store')
        self._security_headers(False)
        if headers:
            for k,v in headers.items():
                if isinstance(v,(list,tuple)):
                    for item in v:self.send_header(k,str(item))
                else:self.send_header(k,str(v))
        self.end_headers(); self.wfile.write(data)

    def send_text(self, text, status=200, ctype='text/html; charset=utf-8'):
        data = text.encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', ctype)
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Cache-Control', 'no-store')
        self._security_headers(ctype.startswith('text/html'))
        self.end_headers(); self.wfile.write(data)

    def send_bytes(self, data, ctype='application/octet-stream', status=200, headers=None):
        self.send_response(status); self.send_header('Content-Type',ctype); self.send_header('Content-Length',str(len(data))); self.send_header('Cache-Control','no-store'); self._security_headers(str(ctype).startswith('text/html'))
        if headers:
            for k,v in headers.items():
                if isinstance(v,(list,tuple)):
                    for item in v:self.send_header(k,str(item))
                else:self.send_header(k,str(v))
        self.end_headers(); self.wfile.write(data)

    def _require_auth(self, conn, role=None):
        s = self._session(conn)
        if not s:
            self.send_json({'ok':False,'error':'unauthorized'},401); return None
        if role and s['role'] != role:
            self.send_json({'ok':False,'error':'forbidden'},403); return None
        return s

    def _require_csrf(self, session, payload):
        token = self.headers.get('X-CSRF-Token') or payload.get('csrf')
        return bool(token and session and hmac.compare_digest(str(token),str(session['csrf'])))

    def csrf_ok(self):
        token=self.headers.get('X-CSRF-Token')
        if not token:return False
        with connect(DB_PATH) as conn:
            session=self._session(conn)
            return bool(session and hmac.compare_digest(str(token),str(session['csrf'])))

    def do_GET(self):
        p = urllib.parse.urlparse(self.path)
        path = p.path
        qs = urllib.parse.parse_qs(p.query)
        if path == '/': return self.send_text((STATIC/'index.html').read_text(encoding='utf-8').replace('{{VERSION}}',VERSION))
        if path in ('/signup/payment-return','/signup') and (path == '/signup/payment-return' or (qs.get('payment') or [''])[0].startswith('return')):
            # O Mercado Pago acrescenta os parâmetros de retorno à back_url. Usamos uma rota
            # dedicada (sem query string prévia) para evitar URLs como ?payment=return?preapproval_id=...
            # e mantemos compatibilidade com retornos da v101 que já tenham sido gerados.
            preapproval_id=(qs.get('preapproval_id') or [''])[0].strip()
            if not preapproval_id and path == '/signup':
                raw_payment=(qs.get('payment') or [''])[0]
                if '?' in raw_payment:
                    legacy=urllib.parse.parse_qs(raw_payment.split('?',1)[1])
                    preapproval_id=(legacy.get('preapproval_id') or [''])[0].strip()
            if not preapproval_id:
                return self.send_redirect('/signup?payment=missing',302)
            try:
                sub=mp_request('GET','/preapproval/'+urllib.parse.quote(preapproval_id,safe=''))
            except Exception:
                return self.send_redirect('/signup?payment=error',302)
            status=str(sub.get('status') or '').strip().lower()
            reference=str(sub.get('external_reference') or '').strip()
            # Diagnóstico do retorno do Mercado Pago. Não registra e-mail, token de acesso
            # ou outros dados sensíveis; apenas os campos necessários para entender o fluxo.
            print('[BILLING] MP_RETURN', json.dumps({
                'preapproval_id': preapproval_id,
                'status': status or None,
                'external_reference': reference or None,
                'init_point_present': bool(sub.get('init_point')),
                'date_created': sub.get('date_created'),
                'last_modified': sub.get('last_modified'),
                'next_payment_date': sub.get('next_payment_date'),
            }, ensure_ascii=False, default=str), flush=True)
            if status != 'authorized' or not reference.startswith('signup:'):
                print('[BILLING] MP_RETURN_NOT_ACTIVE status=%s reference_ok=%s' % (status or 'missing', reference.startswith('signup:')), flush=True)
                # O Mercado Pago pode retornar ao back_url antes de a assinatura mudar de
                # pending para authorized. Em vez de perder o contexto e voltar ao formulário,
                # mantemos o usuário numa tela de confirmação que consulta o status novamente.
                if status == 'pending' and reference.startswith('signup:'):
                    return self.send_redirect('/signup/payment-pending?preapproval_id='+urllib.parse.quote(preapproval_id,safe=''),302)
                return self.send_redirect('/signup?payment='+urllib.parse.quote(status or 'pending'),302)
            token=reference.split(':',1)[1]
            with connect(DB_PATH) as conn:
                signup=conn.execute('SELECT * FROM subscription_signups WHERE token=? AND subscription_id=?',(token,preapproval_id)).fetchone()
                if not signup:
                    return self.send_redirect('/signup?payment=invalid',302)
                if signup['status'] != 'active': provision_signup(conn,signup,sub)
            return self.send_redirect('/login?subscription=active',302)
        if path == '/signup/payment-pending':
            preapproval_id=(qs.get('preapproval_id') or [''])[0].strip()
            if not preapproval_id:
                return self.send_redirect('/signup?payment=missing',302)
            page=(STATIC/'signup-payment-pending.html').read_text(encoding='utf-8').replace('{{VERSION}}',VERSION).replace('{{PREAPPROVAL_ID}}',html.escape(preapproval_id,quote=True))
            return self.send_text(page)
        if path == '/api/public/signup/payment-status':
            preapproval_id=(qs.get('preapproval_id') or [''])[0].strip()
            if not preapproval_id:
                return self.send_json({'ok':False,'error':'missing_preapproval_id'},400)
            try:
                sub=mp_request('GET','/preapproval/'+urllib.parse.quote(preapproval_id,safe=''))
            except Exception:
                return self.send_json({'ok':False,'error':'payment_status_unavailable'},503)
            status=str(sub.get('status') or '').strip().lower()
            reference=str(sub.get('external_reference') or '').strip()
            _mp_subscription_diagnostic(sub,'MP_STATUS_POLL')
            if not reference.startswith('signup:'):
                return self.send_json({'ok':False,'error':'invalid_reference'},400)
            token=reference.split(':',1)[1]
            with connect(DB_PATH) as conn:
                signup=conn.execute('SELECT * FROM subscription_signups WHERE token=? AND subscription_id=?',(token,preapproval_id)).fetchone()
                if not signup:
                    return self.send_json({'ok':False,'error':'signup_not_found'},404)
                if signup['status']=='active':
                    return self.send_json({'ok':True,'active':True,'status':'authorized','redirect':'/login?subscription=active'})
                if status=='authorized':
                    provision_signup(conn,signup,sub)
                    print('[BILLING] MP_POLL_ACTIVATED preapproval_id=%s' % preapproval_id, flush=True)
                    return self.send_json({'ok':True,'active':True,'status':status,'redirect':'/login?subscription=active'})
            rejection_code=None; retry_attempt=None; next_retry_date=None
            if status in ('pending','paused','cancelled'):
                try:
                    invoices=mp_request('GET','/authorized_payments/search?preapproval_id='+urllib.parse.quote(preapproval_id,safe=''))
                    results=invoices.get('results') if isinstance(invoices,dict) else []
                    if isinstance(results,list) and results:
                        latest=sorted(results,key=lambda x:str(x.get('last_modified') or x.get('date_created') or ''),reverse=True)[0]
                        rejection_code=str(latest.get('rejection_code') or '').strip().lower() or None
                        retry_attempt=latest.get('retry_attempt'); next_retry_date=latest.get('next_retry_date')
                        if not rejection_code and isinstance(latest.get('payment'),dict):
                            rejection_code=str(latest['payment'].get('status_detail') or '').strip().lower() or None
                except Exception as exc:
                    print('[BILLING] MP_STATUS_INVOICE_UNAVAILABLE type=%s' % type(exc).__name__,flush=True)
            return self.send_json({'ok':True,'active':False,'status':status or 'pending','rejection_code':rejection_code,'retry_attempt':retry_attempt,'next_retry_date':next_retry_date})
        if path == '/signup': return self.send_text((STATIC/'signup.html').read_text(encoding='utf-8').replace('{{VERSION}}',VERSION))
        if path == '/auth/meta/callback':
            code=(qs.get('code') or [''])[0].strip()
            error=(qs.get('error_description') or qs.get('error_message') or qs.get('error') or [''])[0].strip()
            payload=json.dumps({'type':'CLUBE_META_OAUTH_CALLBACK','code':code,'error':error},ensure_ascii=False).replace('</','<\\/')
            page=f'''<!doctype html><html lang="pt-BR"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Conexão Meta • Fidelizaê!</title><style>body{{font-family:Arial,sans-serif;background:#f7f4ef;color:#1e1713;display:grid;place-items:center;min-height:100vh;margin:0}}main{{max-width:520px;background:#fff;padding:32px;border-radius:22px;box-shadow:0 16px 50px #0002;text-align:center}}h1{{font-size:24px}}p{{line-height:1.5;color:#655b54}}</style></head><body><main><h1>{'Autorização recebida' if code else 'Não foi possível concluir'}</h1><p>{'Você pode voltar ao Fidelizaê!. Esta janela será fechada automaticamente.' if code else html.escape(error or 'A Meta não retornou uma autorização válida.')}</p></main><script>(function(){{const data={payload};try{{if(window.opener&&!window.opener.closed)window.opener.postMessage(data,window.location.origin)}}catch(e){{}}setTimeout(function(){{try{{window.close()}}catch(e){{}}}},900)}})();</script></body></html>'''
            return self.send_text(page)
        if path.startswith('/empresa/'):
            code=path.split('/empresa/',1)[1].strip().upper()
            return self.send_redirect('/join?campaign='+urllib.parse.quote(code),302)
        if path == '/terms': return self.send_text(render_legal_template('terms.html'))
        if path == '/privacy':
            code=(qs.get('campaign') or [''])[0].upper().strip(); client='seu estabelecimento'
            if code:
                with connect(DB_PATH) as conn:
                    c=conn.execute('SELECT name FROM campaigns WHERE code=? AND active=1',(code,)).fetchone()
                    if c:client=c['name']
            template=render_legal_template('privacy.html').replace('{{CLIENT_NAME}}',html.escape(str(client)))
            return self.send_text(template)
        if path == '/join':
            code=(qs.get('campaign') or ['CAFE5'])[0].upper().strip()
            with connect(DB_PATH) as conn:
                c=conn.execute('''SELECT c.name,c.reward_name,c.goal,c.logo_image,co.primary_color,co.logo_text FROM campaigns c JOIN companies co ON co.id=c.company_id WHERE c.code=? AND c.active=1''',(code,)).fetchone()
            template=(STATIC/'join.html').read_text(encoding='utf-8').replace('{{VERSION}}',VERSION)
            if c:
                if c['logo_image']:
                    logo_block = '<img class="campaign-logo" src="' + html.escape(str(c['logo_image']), quote=True) + '" alt="Logo do cliente">'
                else:
                    logo_block = '<div class="brand campaign-logo-fallback">' + html.escape(str(c['logo_text'])) + '</div>'
                template=template.replace('{{LOGO_BLOCK}}',logo_block).replace('{{CAMPAIGN_NAME}}',html.escape(str(c['name'])))
                template=template.replace('name="campaign_code" value="CAFE5"',f'name="campaign_code" value="{html.escape(code)}"')
                template=template.replace('href="/privacy"',f'href="/privacy?campaign={urllib.parse.quote(code)}"')
                template=template.replace('</head>',f"<style>:root{{--accent:{html.escape(str(c['primary_color']))}}}</style></head>")
                error_code=(qs.get('error') or [''])[0]
                if error_code:
                    messages={'invalid_name':'Preencha seu nome corretamente.','invalid_email':'Digite um e-mail válido.','invalid_phone':'Digite um celular válido com DDD.','invalid_birth_date':'Digite uma data de nascimento válida.','invalid_cpf':'Digite um CPF válido.','privacy_consent_required':'É necessário aceitar a Política de Privacidade para criar o cartão.','campaign_not_found':'Cliente não encontrado.'}
                    message=messages.get(error_code,'Não foi possível criar o cartão. Confira os dados e tente novamente.')
                    template=template.replace('<div id="msg"></div>','<div id="msg"><div class="notice error">'+html.escape(message)+'</div></div>')
            else:
                template=template.replace('{{LOGO_BLOCK}}','<div class="brand campaign-logo-fallback">CLUBE</div>').replace('{{CAMPAIGN_NAME}}','Cliente não encontrado').replace('<form id="f" class="form" method="post" action="/join">','<form id="f" class="form hidden" method="post" action="/join">')
            return self.send_text(template)
        if path == '/security':
            return self.send_text((STATIC/'security.html').read_text(encoding='utf-8').replace('{{VERSION}}',VERSION))
        if path == '/login/2fa':
            template=(STATIC/'two-factor.html').read_text(encoding='utf-8').replace('{{VERSION}}',VERSION)
            if (qs.get('error') or [''])[0]: template=template.replace('<div id="msg"></div>','<div id="msg"><div class="notice error">Código inválido ou expirado.</div></div>')
            return self.send_text(template)
        if path in ['/login','/manager','/attendant','/card','/rewards','/loyalty360','/help','/api-docs']:
            name = path.strip('/') + '.html'
            template=(STATIC/name).read_text(encoding='utf-8').replace('{{VERSION}}',VERSION)
            if path == '/login' and (qs.get('error') or [''])[0]:
                login_error=(qs.get('error') or [''])[0]
                message='Sua assinatura terminou e o acesso da empresa está encerrado. Entre em contato para reativar o plano.' if login_error=='subscription_expired' else 'E-mail ou senha inválidos.'
                template=template.replace('<div id="msg"></div>','<div id="msg"><div class="notice error">'+html.escape(message)+'</div></div>')
            return self.send_text(template)
        if path.startswith('/static/'):
            target = (STATIC / path[len('/static/'):]).resolve()
            if STATIC.resolve() not in target.parents or not target.exists() or not target.is_file():
                return self.send_text('Not found',404,'text/plain')
            import mimetypes
            ctype = mimetypes.guess_type(str(target))[0] or 'application/octet-stream'
            if ctype.startswith('text/') or ctype in ('application/javascript','application/json','image/svg+xml'):
                return self.send_text(target.read_text(encoding='utf-8'),200,ctype + ('; charset=utf-8' if ctype.startswith('text/') or ctype in ('application/javascript','application/json') else ''))
            return self.send_bytes(target.read_bytes(),ctype,200,{'Cache-Control':'public, max-age=3600'})
        if path == '/reset-password':
            target=STATIC/'reset-password.html'; return self.send_text(target.read_text(encoding='utf-8').replace('{{VERSION}}',VERSION),200,'text/html; charset=utf-8')
        if path == '/api/health': return self.send_json({'ok':True,'version':VERSION,'database':'postgresql' if str(DB_PATH).startswith(('postgres://','postgresql://')) else 'sqlite'})
        if path == '/api/session':
            with connect(DB_PATH) as conn:
                s=self._session(conn)
                if not s: return self.send_json({'ok':False,'authenticated':False})
                if s['campaign_id']:
                    reconcile_campaign_billing(conn,s['campaign_id'])
                    s=self._session(conn)
                    if not s:return self.send_json({'ok':False,'authenticated':False,'error':'subscription_expired'},401)
                return self.send_json({'ok':True,'authenticated':True,'user':{'id':s['user_id'],'name':s['name'],'email':s['email'],'role':s['role'],'campaign_id':s['campaign_id'],'client_name':s['client_name'],'client_logo_image':s['client_logo_image'],'client_plan':normalize_plan(s['client_plan']),'subscription_status':s['subscription_status'] or 'manual','subscription_next_payment_at':s['subscription_next_payment_at'],'subscription_current_period_end':s['subscription_current_period_end'],'subscription_cancel_at_period_end':bool(s['subscription_cancel_at_period_end']),'billing_option':s['billing_option'],'billing_amount':s['billing_amount'],'commitment_until':s['commitment_until'],'renewal_cancelled_at':s['renewal_cancelled_at'],'pending_plan':s['pending_plan'],'is_client_admin':bool(s['is_client_admin']),'permissions':session_permissions(rowdict(s)),'two_factor_enabled':bool(s['totp_enabled'])},'csrf':s['csrf']})
        if path == '/api/security/sessions':
            with connect(DB_PATH) as conn:
                s=self._require_auth(conn)
                if not s:return
                n=conn.execute('SELECT COUNT(*) n FROM sessions WHERE user_id=? AND expires_at>=?',(s['user_id'],now_ts())).fetchone()['n']
                return self.send_json({'ok':True,'active_sessions':int(n or 0)})
        if path == '/api/wallet/status': return self.send_json({'ok':True,**wallet_status()})
        if path == '/api/campaign/public':
            code=(qs.get('code') or [''])[0].upper().strip()
            with connect(DB_PATH) as conn:
                c=conn.execute('''SELECT c.code,c.name,c.reward_name,c.goal,c.icon,c.logo_image,c.card_theme,c.loyalty_type,c.points_spend_cents,c.plan,co.name company_name,co.primary_color,co.logo_text FROM campaigns c JOIN companies co ON co.id=c.company_id WHERE c.code=? AND c.active=1''',(code,)).fetchone()
                if not c:
                    return self.send_redirect('/join?campaign='+urllib.parse.quote(code or 'CAFE5')+'&error=1') if path=='/join' else self.send_json({'ok':False,'error':'campaign_not_found'},404)
                return self.send_json({'ok':True,'campaign':rowdict(c)})
        if path == '/api/card':
            public_id=(qs.get('id') or [''])[0]
            with connect(DB_PATH) as conn:
                m=conn.execute('''SELECT m.public_id,m.qr_token,m.progress,m.rewards_available,m.status,m.created_at,
                                  c.name campaign_name,c.reward_name,c.goal,c.icon,c.code,c.logo_image,c.card_theme,c.loyalty_type,c.points_spend_cents,c.plan,
                                  m.points_balance,m.id membership_id,cu.name customer_name,co.name company_name,co.primary_color,co.logo_text
                                  FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id JOIN companies co ON co.id=c.company_id
                                  WHERE m.public_id=?''',(public_id,)).fetchone()
                if not m: return self.send_json({'ok':False,'error':'card_not_found'},404)
                device_os=device_os_from_user_agent(self.headers.get('User-Agent'))
                conn.execute('UPDATE memberships SET last_device_os=? WHERE id=?',(device_os,m['membership_id']))
                data=rowdict(m)
                data['card_code']=f'CLUBE:{m["public_id"]}'
                data['qr_value']=data['card_code']
                data['apple_link']=apple_pass_link(public_id)
                data['google_link']=google_wallet_link(public_id)
                data['recent_history']=[rowdict(x) for x in conn.execute('SELECT type,value,note,created_at FROM transactions WHERE membership_id=? ORDER BY created_at DESC LIMIT 12',(m['membership_id'],)).fetchall()]
                plan=normalize_plan(m['plan'])
                data['available_coupons']=[rowdict(x) for x in conn.execute("SELECT name,code,benefit_type,benefit_value,ends_at FROM coupons WHERE campaign_id=(SELECT campaign_id FROM memberships WHERE id=?) AND active=1 AND (starts_at IS NULL OR starts_at<=?) AND (ends_at IS NULL OR ends_at>=?) ORDER BY id DESC LIMIT 5",(m['membership_id'],now_ts(),now_ts())).fetchall()] if PLAN_FEATURES[plan]['coupons'] else []
                tier=conn.execute('SELECT name,benefit FROM loyalty_tiers WHERE campaign_id=(SELECT campaign_id FROM memberships WHERE id=?) AND active=1 AND min_points<=? ORDER BY min_points DESC LIMIT 1',(m['membership_id'],m['points_balance'] or 0)).fetchone() if PLAN_FEATURES[plan]['vip_tiers'] else None; data['tier']=rowdict(tier)
                data['nps_due']=PLAN_FEATURES[plan]['nps'] and not bool(conn.execute('SELECT 1 FROM nps_responses WHERE membership_id=? AND created_at>=? LIMIT 1',(m['membership_id'],now_ts()-90*86400)).fetchone()) and bool(conn.execute('SELECT 1 FROM transactions WHERE membership_id=? LIMIT 1',(m['membership_id'],)).fetchone())
                return self.send_json({'ok':True,'card':data,'wallet':wallet_status()})
        if path == '/api/qr':
            value=(qs.get('data') or [''])[0]
            if not value or len(value)>300: return self.send_text('bad qr data',400,'text/plain')
            img=qrcode.make(value)
            bio=io.BytesIO(); img.save(bio,format='PNG'); data=bio.getvalue()
            self.send_response(200); self.send_header('Content-Type','image/png'); self.send_header('Content-Length',str(len(data))); self.send_header('Cache-Control','no-store'); self._security_headers(False); self.end_headers(); self.wfile.write(data); return
        if path == '/api/card/qr-token':
            public_id=(qs.get('id') or [''])[0].strip()
            with connect(DB_PATH) as conn:
                if not conn.execute('SELECT id FROM memberships WHERE public_id=? AND status=?',(public_id,'active')).fetchone(): return self.send_json({'ok':False,'error':'membership_not_found'},404)
            try: token,exp=make_dynamic_qr(public_id,60)
            except RuntimeError: return self.send_json({'ok':False,'error':'qr_security_not_configured'},503)
            return self.send_json({'ok':True,'token':token,'expires_at':exp})
        if path.startswith('/api/wallet/logo-v62/') or path.startswith('/api/wallet/logo/'):
            if path.startswith('/api/wallet/logo-v62/'):
                # Versioned path is intentionally immutable. Changing the path, not
                # only the query string, forces Google Wallet to fetch the new logo.
                tail=path.split('/api/wallet/logo-v62/',1)[1].strip('/')
                campaign_code=urllib.parse.unquote(tail.split('/',1)[0]).strip().upper()
            else:
                campaign_code=urllib.parse.unquote(path.split('/api/wallet/logo/',1)[1]).strip()
                if campaign_code.lower().endswith('.png'):
                    campaign_code=campaign_code[:-4]
                campaign_code=campaign_code.strip().upper()
            if not campaign_code:return self.send_text('not found',404,'text/plain')
            with connect(DB_PATH) as conn:
                c=conn.execute('SELECT logo_image FROM campaigns WHERE UPPER(code)=UPPER(?) AND active=1',(campaign_code,)).fetchone()
            if not c or not c['logo_image']:return self.send_text('not found',404,'text/plain')
            try:
                # Google Wallet renders programLogo inside a fixed circular slot.
                # Many uploaded logos are square images with a large white/solid
                # background around the real artwork, so simply resizing the whole
                # upload makes the brand mark look tiny. v62 removes the
                # background that is CONNECTED to the image borders, preserving
                # white details that belong to the logo itself.
                raw,_subtype=decode_image_data(c['logo_image'])
                from PIL import Image as PILImage, ImageChops, ImageDraw as PILImageDraw
                src=PILImage.open(io.BytesIO(raw)).convert('RGBA')

                # v62: crop the ACTUAL artwork, not the uploaded square.
                # Google masks programLogo as a circle and recommends a 15% safe
                # margin. We therefore isolate pixels that differ from the edge
                # background, crop tightly, then rebuild a square asset whose
                # artwork fills the whole safe area.
                from PIL import ImageStat
                alpha=src.getchannel('A')
                alpha_bbox=alpha.getbbox()
                if alpha_bbox:
                    src=src.crop(alpha_bbox)

                rgb=src.convert('RGB')
                w,h=rgb.size
                # Robust edge background estimate from many border samples.
                samples=[]
                step=max(1,min(w,h)//64)
                for x in range(0,w,step):
                    samples.append(rgb.getpixel((x,0))); samples.append(rgb.getpixel((x,h-1)))
                for y in range(0,h,step):
                    samples.append(rgb.getpixel((0,y))); samples.append(rgb.getpixel((w-1,y)))
                if samples:
                    bg=tuple(sorted(px[c] for px in samples)[len(samples)//2] for c in range(3))
                else:
                    bg=(255,255,255)

                bg_img=PILImage.new('RGB',rgb.size,bg)
                diff=ImageChops.difference(rgb,bg_img)
                dr,dg,db=diff.split()
                maxdiff=ImageChops.lighter(ImageChops.lighter(dr,dg),db)
                # Anything clearly different from the border is real artwork.
                fg=maxdiff.point(lambda v: 255 if v>24 else 0).convert('L')
                # Include non-transparent pixels when the source already has alpha.
                if src.getchannel('A').getextrema() != (255,255):
                    a=src.getchannel('A').point(lambda v:255 if v>20 else 0)
                    fg=ImageChops.lighter(fg,a)
                art_bbox=fg.getbbox()
                if art_bbox:
                    l,t,r,b=art_bbox
                    # Tiny antialiasing allowance only.
                    pad=max(1,round(max(r-l,b-t)*0.01))
                    art=src.crop((max(0,l-pad),max(0,t-pad),min(src.width,r+pad),min(src.height,b+pad)))
                else:
                    art=src

                size=840
                margin=round(size*0.15)
                safe=size-2*margin
                scale=min(safe/max(1,art.width),safe/max(1,art.height))
                target=(max(1,round(art.width*scale)),max(1,round(art.height*scale)))
                art=art.resize(target,PILImage.Resampling.LANCZOS)
                # Use the detected original edge colour as the full square
                # background, matching Google's logo guidelines.
                canvas=PILImage.new('RGBA',(size,size),(*bg,255))
                # Preserve artwork alpha over the solid background.
                canvas.alpha_composite(art,((size-art.width)//2,(size-art.height)//2))
                out=io.BytesIO(); canvas.save(out,format='PNG',optimize=True)
                return self.send_bytes(out.getvalue(),'image/png',200,{'Cache-Control':'public, max-age=31536000, immutable','X-Content-Type-Options':'nosniff','X-Wallet-Logo-Revision':'v62'})
            except Exception as exc:
                print('[GOOGLE_WALLET] logo render failed:',repr(exc))
                return self.send_text('invalid image',422,'text/plain')
        if path.startswith('/api/wallet/apple/'):
            public_id=urllib.parse.unquote(path.split('/api/wallet/apple/',1)[1])
            with connect(DB_PATH) as conn: card=card_record(conn,public_id)
            if not card:return self.send_json({'ok':False,'error':'membership_not_found'},404)
            try:return self.send_bytes(build_apple_pkpass(card),'application/vnd.apple.pkpass',200,{'Content-Disposition':f'attachment; filename="clube-{public_id}.pkpass"'})
            except Exception as exc:return self.send_json({'ok':False,'error':'apple_wallet_failed','detail':str(exc)[:400]},503)
        if path.startswith('/api/wallet/google/'):
            public_id=urllib.parse.unquote(path.split('/api/wallet/google/',1)[1])
            with connect(DB_PATH) as conn: card=card_record(conn,public_id)
            if not card:return self.send_json({'ok':False,'error':'membership_not_found'},404)
            try:return self.send_redirect(google_save_url(card),302)
            except Exception as exc:return self.send_json({'ok':False,'error':'google_wallet_failed','detail':str(exc)[:400]},503)
        if path.startswith('/api/apple-wallet/v1/passes/'):
            parts=path.split('/'); public_id=urllib.parse.unquote(parts[-1]); auth=(self.headers.get('Authorization') or '').replace('ApplePass ','').strip()
            if not hmac.compare_digest(auth,apple_auth_token(public_id)):return self.send_json({'ok':False,'error':'unauthorized'},401)
            with connect(DB_PATH) as conn: card=card_record(conn,public_id)
            if not card:return self.send_json({'ok':False,'error':'not_found'},404)
            try:return self.send_bytes(build_apple_pkpass(card),'application/vnd.apple.pkpass')
            except Exception as exc:return self.send_json({'ok':False,'error':'apple_wallet_failed','detail':str(exc)[:300]},503)
        if path.startswith('/api/apple-wallet/v1/devices/') and '/registrations/' in path and path.count('/')==7:
            # /api/apple-wallet/v1/devices/{deviceLibraryIdentifier}/registrations/{passTypeIdentifier}
            parts=path.split('/')
            if len(parts)>=8 and parts[4]=='devices' and parts[6]=='registrations':
                device=urllib.parse.unquote(parts[5]); pass_type=urllib.parse.unquote(parts[7])
                try: since=int((qs.get('passesUpdatedSince') or ['0'])[0] or 0)
                except (TypeError,ValueError): since=0
                with connect(DB_PATH) as conn:
                    rows=conn.execute('''SELECT m.public_id FROM wallet_registrations wr JOIN memberships m ON m.id=wr.membership_id WHERE wr.device_library_id=?''',(device,)).fetchall()
                return self.send_json({'serialNumbers':[r['public_id'] for r in rows],'lastUpdated':str(now_ts())})
        if path == '/api/attendant/customer/history':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                try: customer_id=int((qs.get('customer_id') or ['0'])[0])
                except: customer_id=0
                row=conn.execute("SELECT cu.id,cu.name,cu.email,cu.phone,cu.phone_enc,cu.birth_date,cu.gender,cu.cpf,cu.cpf_enc,cu.created_at,cu.marketing_email,cu.marketing_whatsapp,m.id membership_id,m.public_id,m.progress,m.points_balance,m.rewards_available,m.status,c.name campaign_name,c.goal,c.reward_name,c.loyalty_type,c.points_spend_cents FROM customers cu JOIN memberships m ON m.customer_id=cu.id JOIN campaigns c ON c.id=m.campaign_id WHERE cu.id=? AND m.campaign_id=?",(customer_id,sess['campaign_id'])).fetchone()
                if not row:return self.send_json({'ok':False,'error':'customer_not_found'},404)
                row=customer_rowdict(row)
                hist=[rowdict(x) for x in conn.execute("SELECT t.type,t.value,t.previous_progress,t.new_progress,t.rewards_delta,t.note,t.created_at,u.name user_name FROM transactions t LEFT JOIN users u ON u.id=t.user_id WHERE t.membership_id=? ORDER BY t.created_at DESC LIMIT 300",(row['membership_id'],)).fetchall()]
                stats=conn.execute("SELECT COUNT(*) visits,MAX(created_at) last_activity,MIN(created_at) first_activity,COALESCE(SUM(CASE WHEN value>0 THEN value ELSE 0 END),0) total_earned,COALESCE(SUM(CASE WHEN type='redeem' THEN 1 ELSE 0 END),0) total_redeems FROM transactions WHERE membership_id=?",(row['membership_id'],)).fetchone()
                intelligence=customer_intelligence(conn,{**rowdict(row),'id':row['membership_id'],'campaign_id':sess['campaign_id']},rowdict(row))
                notes=[rowdict(x) for x in conn.execute("SELECT n.note,n.created_at,u.name user_name FROM customer_notes n LEFT JOIN users u ON u.id=n.user_id WHERE n.membership_id=? ORDER BY n.id DESC LIMIT 30",(row['membership_id'],)).fetchall()]
                communications=[rowdict(x) for x in conn.execute("SELECT kind,status,created_at,sent_at FROM message_queue WHERE campaign_id=? AND (recipient=? OR recipient_hash=?) ORDER BY created_at DESC LIMIT 30",(sess['campaign_id'],row['email'] or '',pii_lookup_hash(row['phone'],'phone') if row['phone'] else None)).fetchall()]
                coupons=[rowdict(x) for x in conn.execute("SELECT cp.name,cp.code,cr.created_at FROM coupon_redemptions cr JOIN coupons cp ON cp.id=cr.coupon_id WHERE cr.membership_id=? ORDER BY cr.created_at DESC LIMIT 30",(row['membership_id'],)).fetchall()]
                redemptions=[rowdict(x) for x in conn.execute("SELECT COALESCE(rc.name,'Recompensa') name,rr.points_cost,rr.created_at FROM reward_redemptions rr LEFT JOIN reward_catalog rc ON rc.id=rr.reward_id WHERE rr.membership_id=? ORDER BY rr.created_at DESC LIMIT 30",(row['membership_id'],)).fetchall()]
                return self.send_json({'ok':True,'customer':rowdict(row),'stats':rowdict(stats),'notes':notes,'communications':communications,'coupons':coupons,'redemptions':redemptions,'history':hist,'intelligence':intelligence})
        if path == '/api/admin/report.csv':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,sess['campaign_id'],'reports'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                plan=campaign_plan(conn,sess['campaign_id'])
                import csv
                b=io.StringIO();w=csv.writer(b)
                if plan=='beginner':
                    rows=conn.execute("SELECT cu.name,cu.email,cu.phone,cu.phone_enc,m.public_id,m.progress,m.status FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? ORDER BY cu.name",(sess['campaign_id'],)).fetchall()
                    rows=[customer_rowdict(r) for r in rows]
                    w.writerow(['Nome','E-mail','Celular','Código','Selos','Status'])
                    [w.writerow([r['name'],r['email'],r['phone'],'CLUBE:'+r['public_id'],r['progress'],r['status']]) for r in rows]
                    filename='relatorio-basico-clientes.csv'
                else:
                    rows=conn.execute("SELECT cu.name,cu.email,cu.phone,cu.phone_enc,cu.birth_date,cu.gender,cu.cpf,cu.cpf_enc,m.public_id,m.progress,m.points_balance,m.rewards_available,m.status FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? ORDER BY cu.name",(sess['campaign_id'],)).fetchall()
                    rows=[customer_rowdict(r) for r in rows]
                    w.writerow(['Nome','E-mail','Celular','Nascimento','CPF','Código','Selos','Pontos','Recompensas','Status'])
                    [w.writerow([r['name'],r['email'],r['phone'],r['birth_date'],r['cpf'],'CLUBE:'+r['public_id'],r['progress'],r['points_balance'],r['rewards_available'],r['status']]) for r in rows]
                    filename='relatorio-completo-clientes.csv'
                return self.send_bytes(b.getvalue().encode('utf-8-sig'),'text/csv; charset=utf-8',200,{'Content-Disposition':'attachment; filename='+filename})
        if path == '/api/manager/report.csv':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'manager')
                if not sess:return
                rows=conn.execute("SELECT c.name empresa,cu.name cliente,cu.email,cu.phone,cu.phone_enc,m.public_id,m.progress,m.rewards_available,m.status FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id WHERE c.company_id=? ORDER BY c.name,cu.name",(sess['company_id'],)).fetchall()
                rows=[customer_rowdict(r) for r in rows]
                import csv
                b=io.StringIO();w=csv.writer(b);w.writerow(['Empresa','Cliente','E-mail','Celular','Código','Selos','Recompensas','Status'])
                [w.writerow([r['empresa'],r['cliente'],r['email'],r['phone'],'CLUBE:'+r['public_id'],r['progress'],r['rewards_available'],r['status']]) for r in rows]
                return self.send_bytes(b.getvalue().encode('utf-8-sig'),'text/csv; charset=utf-8',200,{'Content-Disposition':'attachment; filename=relatorio-geral.csv'})
        if path == '/api/manager/join-qr':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'manager')
                if not sess:return
                try: cid=int((qs.get('campaign_id') or ['0'])[0])
                except: cid=0
                c=conn.execute('SELECT code FROM campaigns WHERE id=? AND company_id=?',(cid,sess['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                base=(os.environ.get('CLUBE_PUBLIC_URL') or ('https://'+self.headers.get('Host',''))).rstrip('/')
                img=qrcode.make(base+'/join?campaign='+urllib.parse.quote(c['code'])); bio=BytesIO(); img.save(bio,format='PNG')
                return self.send_bytes(bio.getvalue(),'image/png',200,{'Cache-Control':'no-store'})
        if path == '/api/admin/client-qr':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                c=conn.execute('SELECT code FROM campaigns WHERE id=? AND company_id=? AND active=1',(sess['campaign_id'],sess['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                base=(os.environ.get('PUBLIC_BASE_URL') or ('https://'+self.headers.get('Host',''))).rstrip('/')
                img=qrcode.make(base+'/join?campaign='+urllib.parse.quote(c['code'])); bio=BytesIO(); img.save(bio,format='PNG')
                return self.send_bytes(bio.getvalue(),'image/png',200,{'Cache-Control':'no-store'})
        if path == '/api/admin/templates':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,sess['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                return self.send_json({'ok':True,'templates':[rowdict(r) for r in conn.execute('SELECT * FROM message_templates WHERE campaign_id=? ORDER BY name',(sess['campaign_id'],)).fetchall()]})
        if path in ('/api/admin/template/whatsapp-consented-customers','/api/admin/template/consented-customers'):
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,sess['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                channel='whatsapp' if path.endswith('whatsapp-consented-customers') else str((qs.get('channel') or [''])[0]).strip().lower()
                if channel not in ('email','whatsapp','both'):return self.send_json({'ok':False,'error':'invalid_test_channel'},400)
                if channel=='email':
                    rows=conn.execute('''SELECT DISTINCT cu.id,cu.name,cu.email,cu.phone,cu.phone_enc FROM customers cu JOIN memberships m ON m.customer_id=cu.id
                        WHERE m.campaign_id=? AND cu.marketing_email=1 AND cu.email IS NOT NULL AND cu.email<>? ORDER BY cu.name''',(sess['campaign_id'],'')).fetchall()
                elif channel=='whatsapp':
                    rows=conn.execute('''SELECT DISTINCT cu.id,cu.name,cu.email,cu.phone,cu.phone_enc FROM customers cu JOIN memberships m ON m.customer_id=cu.id
                        WHERE m.campaign_id=? AND cu.marketing_whatsapp=1 AND COALESCE(cu.phone_enc,cu.phone) IS NOT NULL AND COALESCE(cu.phone_enc,cu.phone)<>? ORDER BY cu.name''',(sess['campaign_id'],'')).fetchall()
                else:
                    rows=conn.execute('''SELECT DISTINCT cu.id,cu.name,cu.email,cu.phone,cu.phone_enc FROM customers cu JOIN memberships m ON m.customer_id=cu.id
                        WHERE m.campaign_id=? AND cu.marketing_email=1 AND cu.marketing_whatsapp=1
                        AND cu.email IS NOT NULL AND cu.email<>? AND COALESCE(cu.phone_enc,cu.phone) IS NOT NULL AND COALESCE(cu.phone_enc,cu.phone)<>? ORDER BY cu.name''',(sess['campaign_id'],'','')).fetchall()
                rows=[customer_rowdict(r) for r in rows]
                wa_mode='not_applicable'
                wa_available=True
                if channel in ('whatsapp','both'):
                    client_cfg=whatsapp_config_for_client(conn,sess['campaign_id'])
                    if whatsapp_cloud_configured(client_cfg):
                        wa_mode='production'
                    elif whatsapp_meta_test_configured():
                        wa_mode='meta_test'
                        allowed=whatsapp_meta_test_recipients()
                        rows=[r for r in rows if _normalize_phone(r['phone']) in allowed]
                    else:
                        wa_mode='unavailable'; wa_available=False; rows=[]
                return self.send_json({'ok':True,'channel':channel,'customers':[dict(r) for r in rows],
                    'whatsapp_mode':wa_mode,'whatsapp_available':wa_available})
        if path == '/api/manager/notifications':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'manager')
                if not sess:return
                rows=[]
                failed=conn.execute("SELECT c.id,c.name,COUNT(*) n FROM message_queue q JOIN campaigns c ON c.id=q.campaign_id WHERE c.company_id=? AND q.status='failed' GROUP BY c.id,c.name",(sess['company_id'],)).fetchall()
                for r in failed: rows.append({'kind':'error','campaign_id':r['id'],'title':'Falhas de envio • '+r['name'],'message':str(r['n'])+' mensagem(ns) com falha. Verifique a integração e a fila de comunicação.'})
                camps=conn.execute('SELECT * FROM campaigns WHERE company_id=? AND active=1',(sess['company_id'],)).fetchall(); now=now_ts()
                for c in camps:
                    cid=c['id']; ec=email_configured(email_config_for_client(conn,cid)); wc=whatsapp_cloud_configured(whatsapp_config_for_client(conn,cid))
                    if not ec: rows.append({'kind':'integration','campaign_id':cid,'title':'E-mail não configurado • '+c['name'],'message':'Configure o e-mail para ativar campanhas e automações.'})
                    if not wc: rows.append({'kind':'integration','campaign_id':cid,'title':'WhatsApp não configurado • '+c['name'],'message':'Conecte o WhatsApp para ativar mensagens e recuperação de clientes.'})
                    last=conn.execute("SELECT MAX(t.created_at) ts FROM transactions t JOIN memberships m ON m.id=t.membership_id WHERE m.campaign_id=?",(cid,)).fetchone()['ts']
                    cards=conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND status='active'",(cid,)).fetchone()['n']
                    if cards and (not last or last < now-7*86400): rows.append({'kind':'attention','campaign_id':cid,'title':'Sem movimentação recente • '+c['name'],'message':'Nenhum atendimento foi registrado nos últimos 7 dias. Confira se a equipe está utilizando o Fidelizaê!.'})
                    risk=conn.execute("SELECT COUNT(*) n FROM memberships m WHERE m.campaign_id=? AND m.status='active' AND COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at)<? AND COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at)>=?",(cid,now-30*86400,now-60*86400)).fetchone()['n']
                    if risk: rows.append({'kind':'opportunity','campaign_id':cid,'title':'Oportunidade de recuperação • '+c['name'],'message':str(risk)+' cliente(s) estão entre 30 e 60 dias sem atividade. Uma campanha de retorno pode ajudar.'})
                    if c['loyalty_type']=='stamps':
                        almost=conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND status='active' AND progress=?",(cid,max(int(c['goal'] or 1)-1,0))).fetchone()['n']
                        if almost: rows.append({'kind':'opportunity','campaign_id':cid,'title':'Clientes quase lá • '+c['name'],'message':str(almost)+' cliente(s) estão a apenas 1 selo da recompensa.'})
                    pending=conn.execute("SELECT COUNT(*) n FROM message_queue WHERE campaign_id=? AND status IN ('pending','retry','processing')",(cid,)).fetchone()['n']
                    if pending>=20: rows.append({'kind':'attention','campaign_id':cid,'title':'Fila de comunicação • '+c['name'],'message':str(pending)+' mensagens aguardam processamento.'})
                order={'error':0,'attention':1,'opportunity':2,'integration':3}; rows.sort(key=lambda x:order.get(x.get('kind'),9))
                for n in rows:
                    fp=hashlib.sha256((n.get('title','')+'|'+n.get('message','')).encode()).hexdigest()[:24]; st=conn.execute('SELECT status,priority FROM alert_states WHERE fingerprint=?',(fp,)).fetchone(); n['fingerprint']=fp; n['status']=(st['status'] if st else 'new'); n['priority']=(st['priority'] if st else ('high' if n.get('kind')=='error' else 'medium')); n['action']='campaign' if 'recupera' in n.get('title','').lower() or 'quase' in n.get('title','').lower() else ('integration' if n.get('kind')=='integration' else 'company')
                return self.send_json({'ok':True,'notifications':[x for x in rows[:40] if x.get('status')!='resolved']})
        if path == '/api/attendant/dashboard':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin'] and not self._need_permission(sess,'view_reports'): return
                cid=sess['campaign_id']; now=now_ts(); month_start=int(datetime.now(ZoneInfo('America/Sao_Paulo')).replace(day=1,hour=0,minute=0,second=0,microsecond=0).timestamp())
                metrics={}
                metrics['active_cards']=conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND status='active'",(cid,)).fetchone()['n']
                metrics['new_month']=conn.execute('SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND created_at>=?',(cid,month_start)).fetchone()['n']
                campaign=conn.execute("SELECT loyalty_type FROM campaigns WHERE id=?",(cid,)).fetchone(); loyalty_type=(campaign['loyalty_type'] if campaign else 'stamps') or 'stamps'; metrics['loyalty_type']=loyalty_type
                metrics['stamps_month']=conn.execute("SELECT COALESCE(SUM(t.value),0) n FROM transactions t JOIN memberships m ON m.id=t.membership_id WHERE m.campaign_id=? AND t.type='stamp' AND t.created_at>=?",(cid,month_start)).fetchone()['n']
                metrics['points_month']=conn.execute("SELECT COALESCE(SUM(CASE WHEN t.type='adjustment' AND t.value>0 THEN t.value ELSE 0 END),0) n FROM transactions t JOIN memberships m ON m.id=t.membership_id WHERE m.campaign_id=? AND t.created_at>=?",(cid,month_start)).fetchone()['n']
                metrics['redeems_month']=conn.execute("SELECT COUNT(*) n FROM transactions t JOIN memberships m ON m.id=t.membership_id WHERE m.campaign_id=? AND t.type='redeem' AND t.created_at>=?",(cid,month_start)).fetchone()['n']
                total=metrics['active_cards'] or 1
                if loyalty_type=='points':
                    cheapest=conn.execute("SELECT MIN(points_cost) n FROM reward_catalog WHERE campaign_id=? AND active=1",(cid,)).fetchone()['n']; completed=conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND points_balance>=?",(cid,int(cheapest or 999999999))).fetchone()['n'] if cheapest else 0
                else: completed=conn.execute('SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND rewards_available>0',(cid,)).fetchone()['n']
                metrics['completion_rate']=round(completed*100/total,1)
                for days in (30,60,90): metrics[f'inactive_{days}']=conn.execute('''SELECT COUNT(*) n FROM memberships m WHERE m.campaign_id=? AND COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at)<?''',(cid,now-days*86400)).fetchone()['n']
                pending=conn.execute("SELECT COUNT(*) n FROM message_queue WHERE campaign_id=? AND status IN ('pending','retry','processing')",(cid,)).fetchone()['n']; metrics['messages_pending']=pending
                metrics['birthdays_month']=conn.execute("SELECT COUNT(*) n FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND substr(cu.birth_date,6,2)=?",(cid,datetime.now(ZoneInfo('America/Sao_Paulo')).strftime('%m'))).fetchone()['n']
                returned=conn.execute("SELECT COUNT(*) n FROM memberships m WHERE m.campaign_id=? AND (SELECT COUNT(*) FROM transactions t WHERE t.membership_id=m.id AND ((?='points' AND t.type='adjustment' AND t.value>0) OR (?='stamps' AND t.type='stamp')))>=2",(cid,loyalty_type,loyalty_type)).fetchone()['n']; metrics['return_rate']=round(returned*100/total,1)
                # Retenção e resumo semanal/mensal.
                visit_expr="((c.loyalty_type='points' AND t.type='adjustment' AND t.value>0) OR (c.loyalty_type='stamps' AND t.type='stamp' AND t.value>0))"
                metrics['visits_month']=conn.execute("SELECT COUNT(*) n FROM transactions t JOIN memberships m ON m.id=t.membership_id JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND t.created_at>=? AND "+visit_expr,(cid,month_start)).fetchone()['n']
                metrics['returning_month']=conn.execute("SELECT COUNT(DISTINCT m.id) n FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND EXISTS(SELECT 1 FROM transactions t WHERE t.membership_id=m.id AND t.created_at>=? AND "+visit_expr+") AND EXISTS(SELECT 1 FROM transactions t WHERE t.membership_id=m.id AND t.created_at<? AND "+visit_expr+")",(cid,month_start,month_start)).fetchone()['n']
                metrics['reactivated_month']=conn.execute("SELECT COUNT(DISTINCT m.id) n FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND EXISTS(SELECT 1 FROM transactions t WHERE t.membership_id=m.id AND t.created_at>=? AND "+visit_expr+") AND COALESCE((SELECT MAX(t2.created_at) FROM transactions t2 WHERE t2.membership_id=m.id AND t2.created_at<?),m.created_at)<?",(cid,month_start,month_start,month_start-30*86400)).fetchone()['n']
                week_start=now-7*86400; prev_week=now-14*86400
                week_visits=conn.execute("SELECT COUNT(*) n FROM transactions t JOIN memberships m ON m.id=t.membership_id JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND t.created_at>=? AND "+visit_expr,(cid,week_start)).fetchone()['n']
                prev_visits=conn.execute("SELECT COUNT(*) n FROM transactions t JOIN memberships m ON m.id=t.membership_id JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND t.created_at>=? AND t.created_at<? AND "+visit_expr,(cid,prev_week,week_start)).fetchone()['n']
                metrics['weekly']={'visits':week_visits,'new':conn.execute('SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND created_at>=?',(cid,week_start)).fetchone()['n'],'redeems':conn.execute("SELECT COUNT(*) n FROM transactions t JOIN memberships m ON m.id=t.membership_id WHERE m.campaign_id=? AND t.type='redeem' AND t.created_at>=?",(cid,week_start)).fetchone()['n'],'change_pct':round((week_visits-prev_visits)*100/max(prev_visits,1),1)}
                seg_counts={}
                for seg in ('new','active','vip','at_risk','inactive60','inactive90','almost_reward','reward_ready'):
                    try:
                        rows=campaign_recipient_rows(conn,cid,seg) if seg!='active' else []
                        if seg=='active':
                            raw=conn.execute("""SELECT m.id,m.created_at,COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at) last_activity,
                              (SELECT COUNT(*) FROM transactions t WHERE t.membership_id=m.id AND t.value>0) visits,m.rewards_available,m.progress,c.goal,c.loyalty_type,m.points_balance
                              FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND m.status='active' """,(cid,)).fetchall()
                            n=0
                            for x in raw:
                                rr=bool(x['rewards_available']); ar=(x['loyalty_type']=='stamps' and x['progress']==x['goal']-1)
                                if customer_segment(x['last_activity'],x['created_at'],x['visits'],rr,ar,now)=='active': n+=1
                            seg_counts[seg]=n
                        else: seg_counts[seg]=len(rows)
                    except Exception: seg_counts[seg]=0
                metrics['segments']=seg_counts
                retention=[]
                local_now=datetime.now(ZoneInfo('America/Sao_Paulo'))
                for back in range(5,-1,-1):
                    y=local_now.year; mo=local_now.month-back
                    while mo<=0: mo+=12; y-=1
                    start=datetime(y,mo,1,tzinfo=ZoneInfo('America/Sao_Paulo')); nxt=datetime(y+1,1,1,tzinfo=start.tzinfo) if mo==12 else datetime(y,mo+1,1,tzinfo=start.tzinfo)
                    st,en=int(start.timestamp()),int(nxt.timestamp())
                    active_n=conn.execute("SELECT COUNT(DISTINCT m.id) n FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND EXISTS(SELECT 1 FROM transactions t WHERE t.membership_id=m.id AND t.created_at>=? AND t.created_at<? AND "+visit_expr+")",(cid,st,en)).fetchone()['n']
                    ret_n=conn.execute("SELECT COUNT(DISTINCT m.id) n FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.campaign_id=? AND EXISTS(SELECT 1 FROM transactions t WHERE t.membership_id=m.id AND t.created_at>=? AND t.created_at<? AND "+visit_expr+") AND EXISTS(SELECT 1 FROM transactions t2 WHERE t2.membership_id=m.id AND t2.created_at<? AND "+visit_expr.replace('t.','t2.')+")",(cid,st,en,st)).fetchone()['n']
                    retention.append({'month':f'{mo:02d}/{str(y)[2:]}','rate':round(ret_n*100/max(active_n,1),1)})
                metrics['retention_trend']=retention
                activity=[]
                for d in range(29,-1,-1):
                    day=(datetime.now(ZoneInfo('America/Sao_Paulo'))-timedelta(days=d)).date(); start=int(datetime.combine(day,datetime.min.time(),ZoneInfo('America/Sao_Paulo')).timestamp()); end=start+86400
                    n=conn.execute('SELECT COUNT(*) n FROM transactions t JOIN memberships m ON m.id=t.membership_id WHERE m.campaign_id=? AND t.created_at>=? AND t.created_at<?',(cid,start,end)).fetchone()['n']
                    activity.append({'date':day.isoformat(),'count':n})
                metrics['activity_30d']=activity
                # Financeiro: usa valores de compra registrados tanto em pontos quanto em selos.
                fin=conn.execute('SELECT COALESCE(SUM(pr.amount_cents),0) revenue,COUNT(*) purchases FROM purchase_records pr JOIN memberships m ON m.id=pr.membership_id WHERE m.campaign_id=? AND pr.created_at>=?',(cid,month_start)).fetchone()
                metrics['revenue_month_cents']=int(fin['revenue'] or 0); metrics['purchases_month']=int(fin['purchases'] or 0); metrics['avg_ticket_cents']=round(metrics['revenue_month_cents']/max(metrics['purchases_month'],1))
                intel_summary=campaign_intelligence(conn,cid)
                metrics['segments_smart']=intel_summary['segments']; metrics['smart_return_rate']=intel_summary['return_rate']
                metrics['avg_frequency_days']=intel_summary['avg_frequency_days']; metrics['avg_ltv_cents']=intel_summary['avg_ltv_cents']
                metrics['campaign_revenue_cents']=int(conn.execute('SELECT COALESCE(SUM(mcr.attributed_revenue_cents),0) n FROM marketing_campaign_recipients mcr JOIN marketing_campaigns mc ON mc.id=mcr.marketing_campaign_id WHERE mc.campaign_id=? AND mcr.returned_at>=?',(cid,month_start)).fetchone()['n'] or 0)
                finance=[]
                for back in range(5,-1,-1):
                    y=local_now.year; mo=local_now.month-back
                    while mo<=0: mo+=12; y-=1
                    start=datetime(y,mo,1,tzinfo=ZoneInfo('America/Sao_Paulo')); nxt=datetime(y+1,1,1,tzinfo=start.tzinfo) if mo==12 else datetime(y,mo+1,1,tzinfo=start.tzinfo)
                    st,en=int(start.timestamp()),int(nxt.timestamp())
                    rv=conn.execute('SELECT COALESCE(SUM(pr.amount_cents),0) n FROM purchase_records pr JOIN memberships m ON m.id=pr.membership_id WHERE m.campaign_id=? AND pr.created_at>=? AND pr.created_at<?',(cid,st,en)).fetchone()['n']
                    finance.append({'month':f'{mo:02d}/{str(y)[2:]}','revenue_cents':int(rv or 0)})
                metrics['finance_trend']=finance
                # Demografia: gênero informado no cadastro, idade calculada pela data de nascimento e dispositivo do último acesso ao cartão.
                gender_counts={'female':0,'male':0,'other':0,'prefer_not':0,'unknown':0}
                age_counts={'under18':0,'a18_24':0,'a25_34':0,'a35_44':0,'a45_59':0,'a60plus':0,'unknown':0}
                device_counts={'android':0,'ios':0,'other':0}
                demo_rows=conn.execute("SELECT cu.birth_date,cu.gender,m.last_device_os FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND m.status='active'",(cid,)).fetchall()
                today=datetime.now(ZoneInfo('America/Sao_Paulo')).date()
                for dr in demo_rows:
                    g=(dr['gender'] or '').lower(); gender_counts[g if g in gender_counts and g!='unknown' else 'unknown']+=1
                    try:
                        bd=datetime.strptime(dr['birth_date'] or '', '%Y-%m-%d').date(); age=today.year-bd.year-((today.month,today.day)<(bd.month,bd.day))
                        if age<18: age_counts['under18']+=1
                        elif age<25: age_counts['a18_24']+=1
                        elif age<35: age_counts['a25_34']+=1
                        elif age<45: age_counts['a35_44']+=1
                        elif age<60: age_counts['a45_59']+=1
                        else: age_counts['a60plus']+=1
                    except Exception: age_counts['unknown']+=1
                    dev=(dr['last_device_os'] or 'other').lower(); device_counts[dev if dev in ('android','ios') else 'other']+=1
                metrics['demographics']={'gender':gender_counts,'age':age_counts,'device':device_counts}
                return self.send_json({'ok':True,'metrics':metrics})
        if path == '/api/admin/engagement':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                cid=sess['campaign_id']; now=now_ts()
                # Atualiza conversões: uma nova movimentação depois do envio conta como retorno.
                recs=conn.execute("SELECT r.id,r.membership_id,r.sent_at FROM marketing_campaign_recipients r JOIN marketing_campaigns mc ON mc.id=r.marketing_campaign_id WHERE mc.campaign_id=? AND r.returned_at IS NULL",(cid,)).fetchall()
                for r in recs:
                    hit=conn.execute('SELECT MIN(created_at) ts FROM transactions WHERE membership_id=? AND created_at>?',(r['membership_id'],r['sent_at'])).fetchone()['ts']
                    if hit: conn.execute('UPDATE marketing_campaign_recipients SET returned_at=? WHERE id=?',(hit,r['id']))
                plan=campaign_plan(conn,cid)
                campaigns=[]
                for r in (conn.execute('SELECT * FROM marketing_campaigns WHERE campaign_id=? ORDER BY id DESC LIMIT 40',(cid,)).fetchall() if PLAN_FEATURES[plan]['communications'] else []):
                    d=rowdict(r); stats=conn.execute('SELECT COUNT(*) sent,SUM(CASE WHEN returned_at IS NOT NULL THEN 1 ELSE 0 END) returned FROM marketing_campaign_recipients WHERE marketing_campaign_id=?',(d['id'],)).fetchone(); d['sent_count']=stats['sent'] or 0; d['returned_count']=stats['returned'] or 0; d['conversion_rate']=round((d['returned_count'] or 0)*100/max(d['sent_count'] or 0,1),1); d['attributed_revenue_cents']=conn.execute('SELECT COALESCE(SUM(attributed_revenue_cents),0) n FROM marketing_campaign_recipients WHERE marketing_campaign_id=?',(d['id'],)).fetchone()['n'] or 0; campaigns.append(d)
                coupons=[]
                for r in (conn.execute('SELECT * FROM coupons WHERE campaign_id=? ORDER BY active DESC,id DESC',(cid,)).fetchall() if PLAN_FEATURES[plan]['coupons'] else []):
                    d=rowdict(r); d['uses']=conn.execute('SELECT COUNT(*) n FROM coupon_redemptions WHERE coupon_id=?',(d['id'],)).fetchone()['n']; coupons.append(d)
                camp=conn.execute('SELECT name,code,logo_image,card_theme,loyalty_type,goal FROM campaigns WHERE id=?',(cid,)).fetchone()
                base=(os.environ.get('PUBLIC_BASE_URL') or '').rstrip('/'); join_url=(base+'/join?campaign='+urllib.parse.quote(camp['code'])) if base else ('/join?campaign='+urllib.parse.quote(camp['code']))
                return self.send_json({'ok':True,'campaigns':campaigns,'coupons':coupons,'promo':{'join_url':join_url,'qr_url':'/api/admin/client-qr','client_name':camp['name'],'theme':camp['card_theme'],'has_logo':bool(camp['logo_image'])}})

        if path == '/api/attendant/automations':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not plan_allows(conn,sess['campaign_id'],'automations'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                ensure_automation_defaults(conn,sess['campaign_id']); campaign=conn.execute('SELECT loyalty_type FROM campaigns WHERE id=?',(sess['campaign_id'],)).fetchone(); loyalty_type=(campaign['loyalty_type'] if campaign else 'stamps') or 'stamps'; rows=[rowdict(r) for r in conn.execute("SELECT * FROM automation_rules WHERE campaign_id=? AND (?='stamps' OR rule_type<>'one_to_reward') ORDER BY id",(sess['campaign_id'],loyalty_type)).fetchall()]
                return self.send_json({'ok':True,'rules':rows,'can_edit':bool(sess['is_client_admin'])})
        if path == '/api/client-admin/staff':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                rows=[rowdict(r) for r in conn.execute("""SELECT u.id,u.name,u.email,u.active,u.is_client_admin,u.created_at,u.branch_id,b.name branch_name,b.code branch_code
                    FROM users u LEFT JOIN branches b ON b.id=u.branch_id AND b.campaign_id=u.campaign_id
                    WHERE u.campaign_id=? AND u.role='attendant' ORDER BY u.is_client_admin DESC,u.name""",(sess['campaign_id'],)).fetchall()]
                branches=[rowdict(r) for r in conn.execute('SELECT id,name,code FROM branches WHERE campaign_id=? AND active=1 ORDER BY name',(sess['campaign_id'],)).fetchall()]
                return self.send_json({'ok':True,'staff':rows,'branches':branches})

        if path == '/api/attendant/loyalty360-summary':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                cid=sess['campaign_id']; now=now_ts()
                camp=conn.execute("SELECT id,name,plan,loyalty_type,points_spend_cents,cashback_percent,points_expiry_days FROM campaigns WHERE id=? AND company_id=?",(cid,sess['company_id'])).fetchone()
                if not camp:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                plan=normalize_plan(camp['plan'])
                tiers=[rowdict(x) for x in conn.execute("SELECT name,min_points,benefit,active FROM loyalty_tiers WHERE campaign_id=? AND active=1 ORDER BY min_points",(cid,)).fetchall()] if PLAN_FEATURES[plan]['vip_tiers'] else []
                mult=[rowdict(x) for x in conn.execute("SELECT name,factor,weekday,start_hour,end_hour,active FROM point_multipliers WHERE campaign_id=? AND active=1 ORDER BY id DESC",(cid,)).fetchall()] if PLAN_FEATURES[plan]['multipliers'] else []
                rewards_count=conn.execute("SELECT COUNT(*) n FROM reward_catalog WHERE campaign_id=? AND active=1",(cid,)).fetchone()['n']
                metrics={}
                metrics['customers']=conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND status='active'",(cid,)).fetchone()['n']
                metrics['points_circulation']=conn.execute("SELECT COALESCE(SUM(points_balance),0) n FROM memberships WHERE campaign_id=?",(cid,)).fetchone()['n']
                metrics['cashback_cents']=conn.execute("SELECT COALESCE(SUM(cashback_balance_cents),0) n FROM memberships WHERE campaign_id=?",(cid,)).fetchone()['n']
                metrics['inactive30']=conn.execute("SELECT COUNT(*) n FROM memberships m WHERE campaign_id=? AND COALESCE((SELECT MAX(created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at)<?",(cid,now-30*86400)).fetchone()['n']
                return self.send_json({'ok':True,'can_edit':bool(sess['is_client_admin']),'program_source':'taboo','campaign':rowdict(camp),'tiers':tiers,'multipliers':mult,'rewards_count':rewards_count,'metrics':metrics})

        if path == '/api/admin/loyalty360':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                cid=sess['campaign_id']; now=now_ts()
                camp=rowdict(conn.execute("SELECT id,code,name,reward_name,goal,icon,card_theme,plan,loyalty_type,points_spend_cents,cashback_percent,points_expiry_days,logo_image,billing_option,billing_amount,commitment_until,subscription_status,subscription_next_payment_at,subscription_current_period_end,subscription_cancel_at_period_end FROM campaigns WHERE id=?",(cid,)).fetchone())
                plan=normalize_plan(camp.get('plan'))
                tiers=[rowdict(x) for x in conn.execute("SELECT * FROM loyalty_tiers WHERE campaign_id=? ORDER BY min_points",(cid,)).fetchall()] if PLAN_FEATURES[plan]['vip_tiers'] else []
                mult=[rowdict(x) for x in conn.execute("SELECT * FROM point_multipliers WHERE campaign_id=? ORDER BY id DESC",(cid,)).fetchall()] if PLAN_FEATURES[plan]['multipliers'] else []
                nps=[rowdict(x) for x in conn.execute("SELECT * FROM nps_responses WHERE campaign_id=? ORDER BY id DESC LIMIT 100",(cid,)).fetchall()] if PLAN_FEATURES[plan]['nps'] else []
                gifts=[rowdict(x) for x in conn.execute("SELECT * FROM gift_cards WHERE campaign_id=? ORDER BY id DESC LIMIT 100",(cid,)).fetchall()] if PLAN_FEATURES[plan]['gift_cards'] else []
                metrics={}
                metrics['customers']=conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND status='active'",(cid,)).fetchone()['n']
                metrics['points_circulation']=conn.execute("SELECT COALESCE(SUM(points_balance),0) n FROM memberships WHERE campaign_id=?",(cid,)).fetchone()['n']
                metrics['cashback_cents']=conn.execute("SELECT COALESCE(SUM(cashback_balance_cents),0) n FROM memberships WHERE campaign_id=?",(cid,)).fetchone()['n']
                metrics['inactive30']=conn.execute("SELECT COUNT(*) n FROM memberships m WHERE campaign_id=? AND COALESCE((SELECT MAX(created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at)<?",(cid,now-30*86400)).fetchone()['n']
                scores=[int(x['score']) for x in nps]
                metrics['nps']=round((sum(1 for x in scores if x>=9)-sum(1 for x in scores if x<=6))*100/len(scores)) if scores else None
                return self.send_json({'ok':True,'program_source':'taboo','campaign':camp,'tiers':tiers,'multipliers':mult,'nps':nps,'gift_cards':gifts,'metrics':metrics,'features':PLAN_FEATURES[plan]})

        if path == '/api/attendant/gift-card':
            code=(qs.get('code') or [''])[0].strip().upper()
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not plan_allows(conn,sess['campaign_id'],'gift_cards'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                if not self._need_permission(sess,'use_gift'): return
                if not code:return self.send_json({'ok':False,'error':'gift_code_required'},400)
                gift=conn.execute("SELECT id,code,value_cents,balance_cents,status,purchaser_name,beneficiary_name,created_at FROM gift_cards WHERE campaign_id=? AND upper(code)=upper(?)",(sess['campaign_id'],code)).fetchone()
                if not gift:return self.send_json({'ok':False,'error':'gift_not_found'},404)
                events=[rowdict(x) for x in conn.execute('SELECT event_type,amount_cents,balance_after_cents,note,created_at FROM gift_card_events WHERE gift_card_id=? ORDER BY id DESC LIMIT 20',(gift['id'],)).fetchall()]; return self.send_json({'ok':True,'gift':rowdict(gift),'events':events,'qr_url':'/api/qr?data='+urllib.parse.quote('GIFT:'+gift['code'])})

        if path == '/api/card/history360':
            public_id=(qs.get('id') or [''])[0].strip()
            with connect(DB_PATH) as conn:
                m=conn.execute("SELECT m.id,m.public_id,m.campaign_id,m.progress,m.points_balance,m.cashback_balance_cents,m.rewards_available,m.status,m.created_at,c.name campaign_name,c.loyalty_type,c.cashback_percent,c.points_expiry_days FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=?",(public_id,)).fetchone()
                if not m:return self.send_json({'ok':False,'error':'card_not_found'},404)
                hist=[rowdict(x) for x in conn.execute("SELECT type,value,note,created_at FROM transactions WHERE membership_id=? ORDER BY id DESC LIMIT 100",(m['id'],)).fetchall()]
                tier=conn.execute("SELECT name FROM loyalty_tiers WHERE campaign_id=? AND min_points<=? ORDER BY min_points DESC LIMIT 1",(m['campaign_id'],m['points_balance'])).fetchone()
                return self.send_json({'ok':True,'card':rowdict(m),'history':hist,'tier':tier['name'] if tier else None})
        if path == '/api/manager/meta-config':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'manager')
                if not sess:return
                app_id=os.environ.get('META_APP_ID','').strip()
                config_id=os.environ.get('META_CONFIG_ID','').strip()
                version=(os.environ.get('META_GRAPH_VERSION') or 'v24.0').strip() or 'v24.0'
                callback=meta_callback_url()
                return self.send_json({'ok':True,'configured':meta_embedded_signup_configured(),'app_id':app_id,'config_id':config_id,'graph_version':version,'redirect_uri':callback,'public_base_url':meta_public_base_url()})
        if path == '/api/manager/diagnostics':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'manager')
                if not sess:return
                pending=conn.execute("SELECT COUNT(*) n FROM message_queue WHERE status IN ('pending','retry','processing')").fetchone()['n']; failed=conn.execute("SELECT COUNT(*) n FROM message_queue WHERE status='failed'").fetchone()['n']
                return self.send_json({'ok':True,'version':VERSION,'database':'postgresql' if str(DB_PATH).startswith(('postgres://','postgresql://')) else 'sqlite','wallet':wallet_status(),'queue':{'pending':pending,'failed':failed},'encryption':bool(_secret_box()),'pii_encryption':pii_key_configured(),'meta':meta_embedded_signup_configured(),'global_email':email_configured(global_email_config()),'public_base_url':os.environ.get('PUBLIC_BASE_URL',''),'environment':os.environ.get('APP_ENV','production'),'backup':'available','r2_backup':r2_backup_status(),'sentry':{'configured':bool((os.environ.get('SENTRY_DSN') or '').strip()),'enabled':bool(SENTRY_ENABLED)}})
        if path == '/api/manager/backup':
            return self.send_json({'ok':False,'error':'backup_requires_reauthentication'},405)
        if path == '/api/privacy/export':
            if not self._rate_ok('privacy-export',8,900,self._ip(),1800): return
            public_id=(qs.get('id') or [''])[0].strip(); cpf=normalize_cpf((qs.get('cpf') or [''])[0])
            if not cpf:return self.send_json({'ok':False,'error':'invalid_cpf'},400)
            with connect(DB_PATH) as conn:
                row=conn.execute('''SELECT cu.name,cu.email,cu.phone,cu.phone_enc,cu.birth_date,cu.gender,cu.cpf,cu.cpf_enc,cu.privacy_accepted_at,cu.marketing_email,cu.marketing_whatsapp,m.public_id,m.progress,m.rewards_available,c.name client_name FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=? AND cu.cpf_hash=?''',(public_id,pii_lookup_hash(cpf,'cpf'))).fetchone()
                if not row:return self.send_json({'ok':False,'error':'not_found'},404)
                tx=[rowdict(r) for r in conn.execute('SELECT type,value,previous_progress,new_progress,rewards_delta,note,created_at FROM transactions WHERE membership_id=(SELECT id FROM memberships WHERE public_id=?) ORDER BY created_at',(public_id,)).fetchall()]
                return self.send_json({'ok':True,'data':customer_rowdict(row),'history':tx})
        if path == '/api/manager/admins':
            with connect(DB_PATH) as conn:
                s=self._require_auth(conn,'manager')
                if not s:return
                rows=conn.execute("SELECT id,name,email,active,created_at FROM users WHERE company_id=? AND role='manager' AND active=1 ORDER BY name,email",(s['company_id'],)).fetchall()
                admins=[]
                for r in rows:
                    item=rowdict(r); item['is_current']=int(item.get('id') or 0)==int(s['user_id']); admins.append(item)
                return self.send_json({'ok':True,'admins':admins})
        if path == '/api/manager/overview':
            with connect(DB_PATH) as conn:
                s=self._require_auth(conn,'manager');
                if not s: return
                cid=s['company_id']; now=now_ts()
                metrics={}
                metrics['customers']=conn.execute('''SELECT COUNT(DISTINCT m.customer_id) n FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE c.company_id=?''',(cid,)).fetchone()['n']
                metrics['stamps']=conn.execute('''SELECT COALESCE(SUM(t.value),0) n FROM transactions t JOIN memberships m ON m.id=t.membership_id JOIN campaigns c ON c.id=m.campaign_id WHERE c.company_id=? AND t.type='stamp' ''',(cid,)).fetchone()['n']
                metrics['redeems']=conn.execute('''SELECT COUNT(*) n FROM transactions t JOIN memberships m ON m.id=t.membership_id JOIN campaigns c ON c.id=m.campaign_id WHERE c.company_id=? AND t.type='redeem' ''',(cid,)).fetchone()['n']
                campaigns=[rowdict(r) for r in conn.execute("""SELECT c.*,
                    (SELECT COUNT(*) FROM memberships m WHERE m.campaign_id=c.id) card_count,
                    (SELECT COUNT(*) FROM users ux WHERE ux.campaign_id=c.id AND ux.role='attendant' AND ux.active=1) staff_count,
                    (SELECT COALESCE(SUM(CASE WHEN t.type='stamp' THEN t.value WHEN t.type='adjustment' THEN t.value ELSE 0 END),0)
                       FROM transactions t JOIN memberships m2 ON m2.id=t.membership_id WHERE m2.campaign_id=c.id) stamp_count
                    FROM campaigns c WHERE c.company_id=? ORDER BY c.id DESC""",(cid,)).fetchall()]
                for c in campaigns:
                    c['smtp_configured']=bool(c.get('smtp_host') and c.get('smtp_from'))
                    c['brevo_configured']=bool(c.get('brevo_api_key_enc') and c.get('brevo_sender_email'))
                    c['email_provider']=c.get('email_provider') or 'smtp'
                    c['whatsapp_configured']=bool(c.get('whatsapp_phone_number_id') and c.get('whatsapp_access_token_enc') and c.get('whatsapp_api_version'))
                    c['whatsapp_signup_status']=c.get('whatsapp_signup_status') or ('connected' if c['whatsapp_configured'] else 'not_connected')
                    c['whatsapp_integration_mode']=c.get('whatsapp_integration_mode') or 'manual'
                    c['ecommerce_platform']=normalize_ecommerce_platform(c.get('ecommerce_platform'))
                    c['ecommerce_status']=c.get('ecommerce_status') or ('awaiting_connection' if c['ecommerce_platform']!='none' else 'not_connected')
                    c['ecommerce_configured']=bool(c['ecommerce_platform']!='none' and c.get('ecommerce_webhook_secret'))
                    public_base=(os.environ.get('PUBLIC_BASE_URL') or 'https://app.fidelizae.com.br').rstrip('/')
                    c['ecommerce_webhook_url']=(public_base+f"/api/integrations/ecommerce/{c['id']}/{c.get('ecommerce_webhook_secret')}") if c.get('ecommerce_webhook_secret') else ''
                    c['ecommerce_webhook_secret']=None
                    c['email_configured']=bool(c['brevo_configured'] if c['email_provider']=='brevo' else c['smtp_configured']); c['reward_catalog_count']=conn.execute("SELECT COUNT(*) n FROM reward_catalog WHERE campaign_id=? AND active=1",(c['id'],)).fetchone()['n']; c['wallet_google']=wallet_status()['google']['ready']; c['wallet_apple']=wallet_status()['apple']['ready']
                    c['smtp_password_enc']=None
                    c['brevo_api_key_enc']=None
                    c['whatsapp_access_token_enc']=None
                staff=[rowdict(r) for r in conn.execute('''SELECT u.id,u.name,u.email,u.role,u.active,u.is_client_admin,u.created_at,u.campaign_id,c.name client_name FROM users u LEFT JOIN campaigns c ON c.id=u.campaign_id WHERE u.company_id=? ORDER BY u.role,u.name''',(cid,)).fetchall()]
                return self.send_json({'ok':True,'metrics':metrics,'campaigns':campaigns,'staff':staff})
        if path == '/api/client-admin/company':
            with connect(DB_PATH) as conn:
                s=self._require_auth(conn,'attendant')
                if not s:return
                if not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                c=conn.execute('SELECT id,name,code,logo_image,plan,loyalty_type,points_spend_cents,goal,reward_name,card_theme,min_stamp_interval_sec,max_stamps_per_hour,email_provider,smtp_host,smtp_port,smtp_user,smtp_from,smtp_from_name,smtp_security,brevo_sender_email,brevo_sender_name,brevo_reply_to,whatsapp_phone_number_id,whatsapp_waba_id,whatsapp_api_version,ecommerce_platform,ecommerce_store_url,ecommerce_webhook_secret,ecommerce_status FROM campaigns WHERE id=? AND company_id=?',(s['campaign_id'],s['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                company=rowdict(c); company['email_configured']=bool(email_configured(email_config_for_client(conn,s['campaign_id']))); company['whatsapp_configured']=bool(whatsapp_cloud_configured(whatsapp_config_for_client(conn,s['campaign_id'])))
                company['ecommerce_platform']=normalize_ecommerce_platform(company.get('ecommerce_platform')); company['ecommerce_status']=company.get('ecommerce_status') or ('awaiting_connection' if company['ecommerce_platform']!='none' else 'not_connected')
                public_base=(os.environ.get('PUBLIC_BASE_URL') or 'https://app.fidelizae.com.br').rstrip('/'); company['ecommerce_webhook_url']=(public_base+f"/api/integrations/ecommerce/{company['id']}/{company.get('ecommerce_webhook_secret')}") if company.get('ecommerce_webhook_secret') else ''; company.pop('ecommerce_webhook_secret',None)
                return self.send_json({'ok':True,'company':company,'features':PLAN_FEATURES[normalize_plan(c['plan'])]})
        if path == '/api/attendant/recent':
            with connect(DB_PATH) as conn:
                s=self._require_auth(conn,'attendant')
                if not s: return
                if not s['campaign_id']: return self.send_json({'ok':False,'error':'attendant_without_client'},403)
                tx=[rowdict(r) for r in conn.execute('''SELECT t.id,t.type,t.value,t.previous_progress,t.new_progress,t.rewards_delta,t.note,t.created_at,cu.name customer_name,c.name campaign_name,u.name user_name
                   FROM transactions t JOIN memberships m ON m.id=t.membership_id JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id LEFT JOIN users u ON u.id=t.user_id
                   WHERE c.id=? AND c.company_id=? ORDER BY t.id DESC LIMIT 50''',(s['campaign_id'],s['company_id'])).fetchall()]
                return self.send_json({'ok':True,'transactions':tx,'client':{'id':s['campaign_id'],'name':s['client_name']}})
        if path == '/api/admin/my-account':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                c=conn.execute("""SELECT id,name,code,reward_name,goal,loyalty_type,points_spend_cents,plan,created_at,subscription_id,
                                  email_provider,smtp_host,brevo_sender_email,whatsapp_phone_number_id,ecommerce_platform
                                  FROM campaigns WHERE id=? AND company_id=?""",(sess['campaign_id'],sess['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                signup=None
                if c['subscription_id']:
                    signup=conn.execute("SELECT responsible_name,email,phone,document,created_at FROM subscription_signups WHERE subscription_id=? ORDER BY id DESC LIMIT 1",(c['subscription_id'],)).fetchone()
                if not signup:
                    signup=conn.execute("SELECT responsible_name,email,phone,document,created_at FROM subscription_signups WHERE lower(email)=lower(?) AND lower(company_name)=lower(?) ORDER BY id DESC LIMIT 1",(sess['email'],c['name'])).fetchone()
                account=rowdict(c)
                account.update({
                    'company_name':c['name'],
                    'responsible_name':(signup['responsible_name'] if signup else sess['name']),
                    'email':(signup['email'] if signup else sess['email']),
                    'phone':(signup['phone'] if signup else None),
                    'document':(signup['document'] if signup else None),
                    'integrations':{
                        'email':bool(email_configured(email_config_for_client(conn,sess['campaign_id']))),
                        'whatsapp':bool(whatsapp_cloud_configured(whatsapp_config_for_client(conn,sess['campaign_id']))),
                        'ecommerce':({'nuvemshop':'Nuvemshop','shopify':'Shopify','woocommerce':'WooCommerce','tray':'Tray','vtex':'VTEX','custom':'API própria'}.get(normalize_ecommerce_platform(c['ecommerce_platform'])) if normalize_ecommerce_platform(c['ecommerce_platform'])!='none' else None)
                    }
                })
                # Histórico vinculado diretamente à empresa e, nos cadastros por assinatura,
                # também ao registro de adesão original. Nunca altera evidências antigas.
                rows=conn.execute("""SELECT la.id,la.email,la.terms_version,la.privacy_version,la.accepted_at,la.ip_address,u.name user_name
                                     FROM legal_acceptances la LEFT JOIN users u ON u.id=la.user_id
                                     WHERE la.campaign_id=? ORDER BY la.accepted_at DESC,la.id DESC""",(sess['campaign_id'],)).fetchall()
                legal=[rowdict(r) for r in rows]
                if signup:
                    signup_rows=conn.execute("""SELECT la.id,la.email,la.terms_version,la.privacy_version,la.accepted_at,la.ip_address,
                                               ? AS user_name FROM legal_acceptances la JOIN subscription_signups ss ON ss.id=la.signup_id
                                               WHERE lower(ss.email)=lower(?) AND lower(ss.company_name)=lower(?) ORDER BY la.accepted_at DESC,la.id DESC""",
                                             (signup['responsible_name'],signup['email'],c['name'])).fetchall()
                    seen={x['id'] for x in legal}
                    legal.extend(rowdict(r) for r in signup_rows if r['id'] not in seen)
                    legal.sort(key=lambda x:(int(x.get('accepted_at') or 0),int(x.get('id') or 0)),reverse=True)
                return self.send_json({'ok':True,'account':account,'legal_acceptances':legal})
        if path == '/api/admin/onboarding':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                cid=sess['campaign_id']
                c=conn.execute("SELECT name,logo_image,loyalty_type,reward_name FROM campaigns WHERE id=?",(cid,)).fetchone()
                staff=conn.execute("SELECT COUNT(*) n FROM users WHERE campaign_id=? AND role='attendant' AND is_client_admin=0 AND active=1",(cid,)).fetchone()['n']
                clients=conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=?",(cid,)).fetchone()['n']
                rewards=conn.execute("SELECT COUNT(*) n FROM reward_catalog WHERE campaign_id=? AND active=1",(cid,)).fetchone()['n']
                comm=bool(email_configured(email_config_for_client(conn,cid)) or whatsapp_cloud_configured(whatsapp_config_for_client(conn,cid)))
                wallet=conn.execute("SELECT COUNT(*) n FROM wallet_registrations wr JOIN memberships m ON m.id=wr.membership_id WHERE m.campaign_id=?",(cid,)).fetchone()['n']
                steps=[
                  {'key':'company','label':'Empresa e logo configuradas','done':bool(c and c['name'] and c['logo_image']),'target':'company'},
                  {'key':'program','label':'Programa de fidelidade configurado','done':bool(c and c['reward_name']),'target':'company'},
                  {'key':'reward','label':'Primeira recompensa cadastrada','done':bool(c and (c['loyalty_type']=='stamps' or rewards>0)),'target':'rewards'},
                  {'key':'staff','label':'Primeiro atendente cadastrado','done':staff>0,'target':'staff'},
                  {'key':'customer','label':'Primeiro cliente cadastrado','done':clients>0,'target':'customers'},
                  {'key':'communication','label':'Canal de comunicação configurado','done':comm,'target':'communication'},
                  {'key':'wallet','label':'Primeiro cartão salvo na Wallet','done':wallet>0,'target':'wallet'}
                ]
                return self.send_json({'ok':True,'steps':steps,'completed':sum(1 for x in steps if x['done']),'total':len(steps)})
        if path == '/api/admin/integrations':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,sess['campaign_id'],'advanced'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                keys=[rowdict(r) for r in conn.execute("SELECT id,name,token_prefix,active,last_used_at,created_at FROM api_keys WHERE campaign_id=? ORDER BY id DESC",(sess['campaign_id'],)).fetchall()]
                hooks=[]
                for r in conn.execute("SELECT id,url,events_json,active,created_at FROM webhook_subscriptions WHERE campaign_id=? ORDER BY id DESC",(sess['campaign_id'],)).fetchall():
                    d=rowdict(r)
                    try:d['events']=json.loads(d.pop('events_json') or '[]')
                    except Exception:d['events']=[]
                    hooks.append(d)
                deliveries=[rowdict(r) for r in conn.execute("SELECT event_type,status,http_status,attempts,last_error,created_at,delivered_at FROM webhook_deliveries WHERE campaign_id=? ORDER BY id DESC LIMIT 30",(sess['campaign_id'],)).fetchall()]
                return self.send_json({'ok':True,'api_keys':keys,'webhooks':hooks,'deliveries':deliveries})
        if path == '/api/v1/customers':
            with connect(DB_PATH) as conn:
                ctx=self._api_context(conn)
                if not ctx:return
                rows=conn.execute("""SELECT cu.name,cu.email,cu.phone,cu.phone_enc,cu.birth_date,cu.gender,cu.cpf,cu.cpf_enc,m.public_id,m.progress,m.points_balance,m.rewards_available,m.status,m.created_at
                                     FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? ORDER BY cu.name LIMIT 1000""",(ctx['campaign_id'],)).fetchall()
                out=[]
                for r in rows:
                    d=customer_rowdict(r); intel=customer_intelligence(conn,{**d,'id':conn.execute('SELECT id FROM memberships WHERE public_id=?',(d['public_id'],)).fetchone()['id'],'campaign_id':ctx['campaign_id'],'loyalty_type':ctx['loyalty_type'],'goal':ctx['goal']},ctx); d.update(intel); out.append(d)
                return self.send_json({'ok':True,'data':out})
        if path.startswith('/api/v1/customers/'):
            public_id=urllib.parse.unquote(path.rsplit('/',1)[1])
            with connect(DB_PATH) as conn:
                ctx=self._api_context(conn)
                if not ctx:return
                r=conn.execute("""SELECT cu.name,cu.email,cu.phone,cu.phone_enc,cu.birth_date,cu.gender,cu.cpf,cu.cpf_enc,m.* FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND m.public_id=?""",(ctx['campaign_id'],public_id)).fetchone()
                if not r:return self.send_json({'ok':False,'error':'customer_not_found'},404)
                d=customer_rowdict(r); d['intelligence']=customer_intelligence(conn,d,{**ctx,'goal':ctx['goal']})
                return self.send_json({'ok':True,'data':d})
        if path == '/api/attendant/customers':
            with connect(DB_PATH) as conn:
                s=self._require_auth(conn,'attendant')
                if not s: return
                if not s['campaign_id']: return self.send_json({'ok':False,'error':'attendant_without_client'},403)
                customers=[customer_rowdict(r) for r in conn.execute('''SELECT cu.id,cu.name,cu.email,cu.phone,cu.phone_enc,cu.birth_date,cu.gender,cu.cpf,cu.cpf_enc,cu.created_at,m.id membership_id,m.public_id,m.progress,m.points_balance,m.rewards_available,c.loyalty_type,c.goal,
                    COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at) last_activity,
                    (SELECT COUNT(*) FROM transactions t WHERE t.membership_id=m.id AND ((c.loyalty_type='points' AND t.type='adjustment' AND t.value>0) OR (c.loyalty_type='stamps' AND t.type='stamp' AND t.value>0))) visits,
                    (SELECT COUNT(*) FROM transactions t WHERE t.membership_id=m.id AND t.type='redeem') redeems
                    FROM customers cu JOIN memberships m ON m.customer_id=cu.id JOIN campaigns c ON c.id=m.campaign_id
                    WHERE m.campaign_id=? ORDER BY cu.name''',(s['campaign_id'],)).fetchall()]
                cheapest=conn.execute("SELECT MIN(points_cost) n FROM reward_catalog WHERE campaign_id=? AND active=1",(s['campaign_id'],)).fetchone()['n']
                vip_enabled=plan_allows(conn,s['campaign_id'],'vip_tiers')
                tiers=[rowdict(r) for r in conn.execute("SELECT name,min_points,benefit FROM loyalty_tiers WHERE campaign_id=? AND active=1 ORDER BY min_points",(s['campaign_id'],)).fetchall()] if vip_enabled else []
                for c in customers:
                    intel=customer_intelligence(conn,{**c,'id':c['membership_id'],'campaign_id':s['campaign_id']},c)
                    c.update(intel)
                    if c['segment']=='vip' and not vip_enabled:c['segment']='recurrent'
                    if tiers and c['loyalty_type']=='points':
                        eligible=[t for t in tiers if int(c.get('points_balance') or 0)>=int(t.get('min_points') or 0)]; c['level']=(eligible[-1]['name'] if eligible else 'Inicial')
                    else: c['level']='VIP' if vip_enabled and int(c.get('visits') or 0)>=12 else ('Frequente' if int(c.get('visits') or 0)>=5 else 'Inicial')
                    if c['loyalty_type']=='points': c['to_reward']=max(0,int(cheapest or 0)-int(c.get('points_balance') or 0)) if cheapest else None
                    else: c['to_reward']=max(0,int(c.get('goal') or 0)-int(c.get('progress') or 0))
                month=datetime.now(ZoneInfo('America/Sao_Paulo')).month
                birthdays=[c for c in customers if c.get('birth_date') and len(c['birth_date'])>=10 and int(c['birth_date'][5:7])==month]
                birthdays.sort(key=lambda c: (int(c['birth_date'][8:10]), c['name'].lower()))
                comm=plan_allows(conn,s['campaign_id'],'communications'); return self.send_json({'ok':True,'customers':customers,'birthdays':birthdays,'month':month,'whatsapp_cloud':comm and whatsapp_cloud_configured(whatsapp_config_for_client(conn,s['campaign_id'])),'whatsapp_configured':comm and whatsapp_cloud_configured(whatsapp_config_for_client(conn,s['campaign_id'])),'email_configured':comm and email_configured(email_config_for_client(conn,s['campaign_id']))})

        if path == '/api/card/rewards':
            public_id=(qs.get('id') or [''])[0].strip()
            with connect(DB_PATH) as conn:
                m=conn.execute("SELECT m.id,m.public_id,m.points_balance,c.id campaign_id,c.name campaign_name,c.loyalty_type FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=?",(public_id,)).fetchone()
                if not m:return self.send_json({'ok':False,'error':'card_not_found'},404)
                rewards=[rowdict(r) for r in conn.execute("SELECT id,name,description,points_cost,image_data,active,stock,starts_at,ends_at FROM reward_catalog WHERE campaign_id=? AND active=1 ORDER BY points_cost,name",(m['campaign_id'],)).fetchall()]
                return self.send_json({'ok':True,'loyalty_type':m['loyalty_type'],'points_balance':m['points_balance'],'campaign_name':m['campaign_name'],'rewards':rewards})
        if path == '/api/attendant/rewards':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                c=conn.execute("SELECT id,name,loyalty_type,points_spend_cents FROM campaigns WHERE id=? AND company_id=?",(sess['campaign_id'],sess['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                rewards=[rowdict(r) for r in conn.execute("SELECT id,name,description,points_cost,image_data,active,stock,starts_at,ends_at FROM reward_catalog WHERE campaign_id=? AND active=1 ORDER BY points_cost,name",(sess['campaign_id'],)).fetchall()]
                return self.send_json({'ok':True,'campaign':rowdict(c),'rewards':rewards})
        if path == '/api/admin/rewards':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                rewards=[rowdict(r) for r in conn.execute("SELECT id,name,description,points_cost,image_data,active,stock,starts_at,ends_at,created_at,updated_at FROM reward_catalog WHERE campaign_id=? ORDER BY active DESC,points_cost,name",(sess['campaign_id'],)).fetchall()]
                c=conn.execute("SELECT loyalty_type,points_spend_cents FROM campaigns WHERE id=?",(sess['campaign_id'],)).fetchone()
                return self.send_json({'ok':True,'campaign':rowdict(c) if c else {},'rewards':rewards})

        if path == '/api/attendant/messages':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not plan_allows(conn,sess['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                if not sess['campaign_id']:return self.send_json({'ok':False,'error':'attendant_without_client'},403)
                rows=[queue_rowdict(r) for r in conn.execute("SELECT id,kind,recipient,recipient_hash,status,attempts,last_error,created_at,sent_at,available_at FROM message_queue WHERE campaign_id=? ORDER BY id DESC LIMIT 30",(sess['campaign_id'],)).fetchall()]
                return self.send_json({'ok':True,'messages':rows})
        if path == '/api/admin/branches':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                rows=[rowdict(r) for r in conn.execute("""SELECT b.id,b.name,b.code,b.active,b.created_at,
                    (SELECT COUNT(*) FROM users u WHERE u.branch_id=b.id AND u.campaign_id=b.campaign_id AND u.role='attendant' AND u.active=1) staff_count,
                    (SELECT COUNT(*) FROM transactions t JOIN memberships m ON m.id=t.membership_id LEFT JOIN users ux ON ux.id=t.user_id WHERE m.campaign_id=b.campaign_id AND COALESCE(t.branch_id,ux.branch_id)=b.id) operations_count
                    FROM branches b WHERE b.campaign_id=? ORDER BY b.active DESC,b.name""",(sess['campaign_id'],)).fetchall()]
                me=conn.execute('SELECT branch_id FROM users WHERE id=?',(sess['user_id'],)).fetchone()
                return self.send_json({'ok':True,'branches':rows,'current_branch_id':me['branch_id'] if me else None})
        if path == '/api/admin/audit-center':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                action=(qs.get('action') or [''])[0].strip(); user_id=(qs.get('user_id') or [''])[0].strip(); branch_id=(qs.get('branch_id') or [''])[0].strip()
                params=[sess['campaign_id']]; where=['a.campaign_id=?']
                if action: where.append('a.action=?'); params.append(action)
                if user_id:
                    try: uid=int(user_id)
                    except (TypeError,ValueError): return self.send_json({'ok':False,'error':'invalid_user'},400)
                    where.append('a.user_id=?'); params.append(uid)
                if branch_id:
                    try: bid=int(branch_id)
                    except (TypeError,ValueError): return self.send_json({'ok':False,'error':'invalid_branch'},400)
                    owned=conn.execute('SELECT id FROM branches WHERE id=? AND campaign_id=?',(bid,sess['campaign_id'])).fetchone()
                    if not owned:return self.send_json({'ok':False,'error':'branch_not_found'},404)
                    where.append('COALESCE(a.branch_id,u.branch_id)=?'); params.append(bid)
                rows=[rowdict(r) for r in conn.execute('SELECT a.*,u.name user_name,COALESCE(b.name,\'Sem unidade\') branch_name,b.code branch_code FROM audit_log a LEFT JOIN users u ON u.id=a.user_id LEFT JOIN branches b ON b.id=COALESCE(a.branch_id,u.branch_id) WHERE '+ ' AND '.join(where)+' ORDER BY a.id DESC LIMIT 250',tuple(params)).fetchall()]
                users=[rowdict(r) for r in conn.execute("SELECT u.id,u.name,u.email,u.branch_id,b.name branch_name FROM users u LEFT JOIN branches b ON b.id=u.branch_id WHERE u.company_id=? AND u.active=1 AND (u.campaign_id=? OR u.id=?) ORDER BY u.name,u.email",(sess['company_id'],sess['campaign_id'],sess['user_id'])).fetchall()]
                actions=[r['action'] for r in conn.execute('SELECT DISTINCT action FROM audit_log WHERE campaign_id=? ORDER BY action',(sess['campaign_id'],)).fetchall()]
                branches=[rowdict(r) for r in conn.execute('SELECT id,name,code FROM branches WHERE campaign_id=? AND active=1 ORDER BY name',(sess['campaign_id'],)).fetchall()]
                return self.send_json({'ok':True,'audit':rows,'users':users,'actions':actions,'branches':branches})
        if path == '/api/admin/commercial-report':
            with connect(DB_PATH) as conn:
                sess=self._require_auth(conn,'attendant')
                if not sess:return
                if not sess['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,sess['campaign_id'],'advanced_reports'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                cid=sess['campaign_id']
                top_rewards=[rowdict(r) for r in conn.execute("SELECT COALESCE(rc.name,'Recompensa') name,COUNT(*) total FROM reward_redemptions rr LEFT JOIN reward_catalog rc ON rc.id=rr.reward_id JOIN memberships m ON m.id=rr.membership_id WHERE m.campaign_id=? GROUP BY rc.name ORDER BY total DESC LIMIT 10",(cid,)).fetchall()]
                by_staff=[rowdict(r) for r in conn.execute("""SELECT COALESCE(u.name,'Sistema') name,COALESCE(b.name,'Sem unidade') branch_name,COUNT(*) operations
                    FROM transactions t JOIN memberships m ON m.id=t.membership_id LEFT JOIN users u ON u.id=t.user_id LEFT JOIN branches b ON b.id=COALESCE(t.branch_id,u.branch_id)
                    WHERE m.campaign_id=? GROUP BY u.name,b.name ORDER BY operations DESC LIMIT 30""",(cid,)).fetchall()]
                by_branch=[rowdict(r) for r in conn.execute("""SELECT COALESCE(b.name,'Sem unidade') name,COALESCE(b.code,'—') code,COUNT(*) operations,COUNT(DISTINCT t.user_id) staff_count,COUNT(DISTINCT t.membership_id) customers
                    FROM transactions t JOIN memberships m ON m.id=t.membership_id LEFT JOIN users u ON u.id=t.user_id LEFT JOIN branches b ON b.id=COALESCE(t.branch_id,u.branch_id)
                    WHERE m.campaign_id=? GROUP BY b.id,b.name,b.code ORDER BY operations DESC""",(cid,)).fetchall()]
                return self.send_json({'ok':True,'top_rewards':top_rewards,'by_staff':by_staff,'by_branch':by_branch})
        if path == '/api/attendant/lookup':
            token,token_error=resolve_member_token((qs.get('token') or [''])[0])
            if token_error:return self.send_json({'ok':False,'error':token_error},410 if token_error=='qr_expired' else 400)
            with connect(DB_PATH) as conn:
                s=self._require_auth(conn,'attendant')
                if not s: return
                if not s['campaign_id']: return self.send_json({'ok':False,'error':'attendant_without_client'},403)
                m=conn.execute('''SELECT m.*,cu.name customer_name,c.name campaign_name,c.reward_name,c.goal,c.icon,c.logo_image,c.company_id,c.loyalty_type,c.points_spend_cents
                  FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id WHERE (m.public_id=? OR m.qr_token=?) AND c.company_id=? AND c.id=?''',(token,token,s['company_id'],s['campaign_id'])).fetchone()
                if not m: return self.send_json({'ok':False,'error':'membership_not_found'},404)
                return self.send_json({'ok':True,'membership':rowdict(m)})
        return self.send_text('Not found',404,'text/plain')

    def do_DELETE(self):
        p=urllib.parse.urlparse(self.path); path=p.path
        if path.startswith('/api/apple-wallet/v1/devices/') and '/registrations/' in path:
            parts=path.split('/'); device=parts[5] if len(parts)>5 else ''; public_id=urllib.parse.unquote(parts[-1]); auth=(self.headers.get('Authorization') or '').replace('ApplePass ','').strip()
            if not hmac.compare_digest(auth,apple_auth_token(public_id)):return self.send_json({'ok':False,'error':'unauthorized'},401)
            with connect(DB_PATH) as conn:
                m=conn.execute('SELECT id FROM memberships WHERE public_id=?',(public_id,)).fetchone()
                if m:conn.execute('DELETE FROM wallet_registrations WHERE membership_id=? AND device_library_id=?',(m['id'],device))
            return self.send_json({'ok':True})
        return self.send_json({'ok':False,'error':'not_found'},404)

    def do_POST(self):
        p=urllib.parse.urlparse(self.path); path=p.path
        print(f'[FORM] POST {path} ip={self._ip()} content_type={self.headers.get("Content-Type", "")}')
        try: payload, payload_kind=self._body_payload()
        except Exception:
            if path in ['/login','/join']:
                return self.send_redirect('/login?error=1' if path == '/login' else '/join?error=1')
            return self.send_json({'ok':False,'error':'invalid_json'},400)
        if path=='/api/public/signup':
            name=str(payload.get('responsible_name') or '').strip()[:100]; company=str(payload.get('company_name') or '').strip()[:120]; email=normalize_email(payload.get('email')); phone=str(payload.get('phone') or '').strip()[:40]; document=str(payload.get('document') or '').strip()[:30]; password=str(payload.get('password') or ''); plan=normalize_plan(payload.get('plan')); loyalty=str(payload.get('loyalty_type') or 'stamps').lower(); device_id=str(payload.get('mp_device_id') or '').strip()[:240]; billing_option=normalize_billing_option(plan,payload.get('billing_option'))
            if not payload.get('terms_accepted') or not payload.get('privacy_accepted'):return self.send_json({'ok':False,'error':'legal_acceptance_required'},400)
            if not name or not company or not email or not password_is_strong(password,12) or plan not in PLAN_PRICES:return self.send_json({'ok':False,'error':'invalid_signup'},400)
            try: logo_image=validate_logo_data(payload.get('logo_image'))
            except ValueError as exc:return self.send_json({'ok':False,'error':str(exc)},400)
            if not logo_image:return self.send_json({'ok':False,'error':'logo_required'},400)
            if plan=='beginner': loyalty='stamps'
            with connect(DB_PATH) as conn:
                if conn.execute('SELECT id FROM users WHERE lower(email)=lower(?)',(email,)).fetchone():return self.send_json({'ok':False,'error':'email_exists'},409)
                # Evita gerar várias assinaturas pendentes idênticas em sequência. Se houver
                # uma tentativa recente ainda válida no MP, reutilizamos o checkout existente.
                if plan!='beginner':
                    cutoff=now_ts()-15*60
                    recent=conn.execute("SELECT * FROM subscription_signups WHERE lower(email)=lower(?) AND plan=? AND billing_option=? AND status='pending' AND subscription_id IS NOT NULL AND created_at>=? ORDER BY id DESC LIMIT 1",(email,plan,billing_option,cutoff)).fetchone()
                    if recent:
                        try:
                            existing=mp_request('GET','/preapproval/'+urllib.parse.quote(str(recent['subscription_id']),safe=''))
                            if str(existing.get('status') or '').lower()=='pending' and existing.get('init_point'):
                                print('[BILLING] MP_REUSE_PENDING subscription_id=%s' % recent['subscription_id'],flush=True)
                                return self.send_json({'ok':True,'active':False,'checkout_url':existing.get('init_point'),'reused':True})
                        except Exception as exc:
                            print('[BILLING] MP_REUSE_CHECK_UNAVAILABLE type=%s' % type(exc).__name__,flush=True)
                token=secrets.token_urlsafe(24); _,bcfg=billing_config(plan,billing_option); sid=insert_id(conn,'INSERT INTO subscription_signups(token,company_name,responsible_name,email,phone,document,password_hash,plan,loyalty_type,status,created_at,logo_image,billing_option,billing_amount) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(token,company,name,email,phone,document,hash_password(password),plan,loyalty,'pending',now_ts(),logo_image,billing_option,bcfg['amount']))
                conn.execute('INSERT INTO legal_acceptances(signup_id,email,terms_version,privacy_version,accepted_at,ip_address) VALUES(?,?,?,?,?,?)',(sid,email,TERMS_VERSION,PRIVACY_VERSION,now_ts(),self._ip()))
                row=conn.execute('SELECT * FROM subscription_signups WHERE id=?',(sid,)).fetchone()
                if plan=='beginner':
                    provision_signup(conn,row); return self.send_json({'ok':True,'active':True,'redirect':'/login'})
                try: sub=create_mp_subscription(email,plan,'signup:'+token,device_id=device_id,billing_option=billing_option)
                except RuntimeError as exc:return self.send_json({'ok':False,'error':str(exc)},503)
                conn.execute('UPDATE subscription_signups SET subscription_id=? WHERE id=?',(sub.get('id'),sid))
                return self.send_json({'ok':True,'active':False,'checkout_url':sub.get('init_point')})
        if path=='/api/webhooks/mercadopago':
            # Segurança em duas camadas: valida x-signature quando a secret está configurada
            # e nunca confia no payload para ativar acesso; consulta o recurso na API do MP.
            query=urllib.parse.parse_qs(p.query,keep_blank_values=True)
            if not validate_mp_webhook_signature(self.headers,query):
                print('[BILLING] webhook signature rejected')
                return self.send_json({'ok':False,'error':'invalid_signature'},401)
            data=payload.get('data') or {}; sub_id=str(data.get('id') or payload.get('id') or '').strip(); typ=str(payload.get('type') or payload.get('topic') or '')
            if typ=='subscription_authorized_payment' and sub_id:
                try:
                    ap=mp_request('GET','/authorized_payments/'+urllib.parse.quote(sub_id,safe=''))
                    safe_ap={
                        'id':str(ap.get('id') or '') or None,
                        'preapproval_id':str(ap.get('preapproval_id') or '') or None,
                        'status':str(ap.get('status') or '') or None,
                        'rejection_code':str(ap.get('rejection_code') or '') or None,
                        'retry_attempt':ap.get('retry_attempt'),
                        'next_retry_date':ap.get('next_retry_date'),
                        'payment_method_id':str(ap.get('payment_method_id') or '') or None,
                    }
                    print('[BILLING] MP_AUTHORIZED_PAYMENT '+json.dumps(safe_ap,ensure_ascii=False,default=str),flush=True)
                    pre_id=str(ap.get('preapproval_id') or '').strip()
                    rejection=str(ap.get('rejection_code') or '').strip().lower()
                    if pre_id:
                        payment=ap.get('payment') if isinstance(ap.get('payment'),dict) else {}
                        approved=str(payment.get('status') or '').lower()=='approved'
                        with connect(DB_PATH) as conn:
                            signup=conn.execute('SELECT * FROM subscription_signups WHERE subscription_id=?',(pre_id,)).fetchone()
                            if rejection and signup and signup['status']!='active':
                                # Mantemos pending enquanto o MP estiver em recycling, mas
                                # registramos a recusa para diagnóstico sem expor dados sensíveis.
                                print('[BILLING] MP_SIGNUP_PAYMENT_REJECTED subscription_id=%s code=%s' % (pre_id,rejection),flush=True)
                            camp=conn.execute('SELECT * FROM campaigns WHERE subscription_id=?',(pre_id,)).fetchone()
                            if approved and camp and not camp['subscription_cancel_at_period_end']:
                                _refresh_annual_commitment_after_payment(conn,camp)
                                camp=conn.execute('SELECT * FROM campaigns WHERE id=?',(camp['id'],)).fetchone()
                            if approved and camp and camp['subscription_cancel_at_period_end'] and normalize_billing_option(camp['plan'],camp['billing_option'])=='annual_monthly':
                                paid=_approved_subscription_invoice_count(pre_id)
                                if paid is not None and paid>=12 and _best_effort_cancel_subscription(pre_id):
                                    conn.execute("UPDATE campaigns SET subscription_status='non_renewing',subscription_next_payment_at=NULL,subscription_status_updated_at=? WHERE id=?",(now_ts(),camp['id']))
                                    audit(conn,camp['company_id'],None,'subscription_annual_commitment_completed','campaign',camp['id'],details='12_paid_installments')
                except Exception as exc:
                    print('[BILLING] MP_AUTHORIZED_PAYMENT_UNAVAILABLE type=%s' % type(exc).__name__,flush=True)
                return self.send_json({'ok':True})
            if not sub_id or ('preapproval' not in typ and typ not in ('subscription_preapproval','')): return self.send_json({'ok':True})
            try: sub=mp_request('GET','/preapproval/'+urllib.parse.quote(sub_id,safe=''))
            except Exception:return self.send_json({'ok':True})
            status=str(sub.get('status') or '').lower(); mapped=_mp_status(status); next_ts=_mp_timestamp(sub.get('next_payment_date')); now=now_ts()
            with connect(DB_PATH) as conn:
                signup=conn.execute('SELECT * FROM subscription_signups WHERE subscription_id=?',(sub_id,)).fetchone()
                if signup and status=='authorized': provision_signup(conn,signup,sub)
                camp=conn.execute('SELECT * FROM campaigns WHERE subscription_id=? OR pending_subscription_id=? OR previous_subscription_id=?',(sub_id,sub_id,sub_id)).fetchone()
                if camp:
                    if camp['pending_subscription_id']==sub_id:
                        # Upgrade: libera o novo plano somente depois da nova assinatura ser autorizada.
                        if mapped=='active' and camp['pending_plan']:
                            old_id=camp['subscription_id']; new_plan=normalize_plan(camp['pending_plan'])
                            new_option,new_cfg=billing_config(new_plan,normalize_billing_option(new_plan,camp['billing_option']))
                            conn.execute("UPDATE campaigns SET plan=?,billing_option=?,billing_amount=?,subscription_provider='mercadopago',subscription_id=?,pending_subscription_id=NULL,pending_plan=NULL,previous_subscription_id=?,subscription_status='active',subscription_started_at=COALESCE(subscription_started_at,?),subscription_current_period_end=?,subscription_next_payment_at=?,subscription_status_updated_at=?,subscription_cancel_at_period_end=0 WHERE id=?",
                                (new_plan,new_option,new_cfg['amount'],sub_id,old_id,now,next_ts,next_ts,now,camp['id']))
                            if old_id and old_id!=sub_id and _best_effort_cancel_subscription(old_id):
                                conn.execute('UPDATE campaigns SET previous_subscription_id=NULL WHERE id=?',(camp['id'],))
                            audit(conn,camp['company_id'],None,'plan_upgrade_confirmed','campaign',camp['id'],details=new_plan)
                        elif mapped=='cancelled':
                            # Checkout de upgrade cancelado: mantém intactos plano e assinatura atuais.
                            conn.execute('UPDATE campaigns SET pending_subscription_id=NULL,pending_plan=NULL,subscription_change_requested_at=NULL WHERE id=?',(camp['id'],))
                            audit(conn,camp['company_id'],None,'plan_upgrade_cancelled','campaign',camp['id'])
                        elif mapped=='past_due':
                            # A falha pertence apenas à nova assinatura de upgrade; a assinatura atual segue valendo.
                            pass
                    elif camp['subscription_id']==sub_id:
                        # Um downgrade agendado nunca é antecipado por um simples webhook de atualização.
                        effective=int(camp['subscription_current_period_end'] or 0)
                        pending=normalize_plan(camp['pending_plan']) if camp['pending_plan'] else None
                        conn.execute('UPDATE campaigns SET subscription_status=?,subscription_next_payment_at=?,subscription_status_updated_at=? WHERE id=?',(mapped,next_ts,now,camp['id']))
                        if pending and effective and now>=effective:
                            if pending=='beginner':
                                reconcile_campaign_billing(conn,camp['id'],allow_remote=False)
                            elif mapped=='active' and next_ts and next_ts>effective:
                                # A próxima cobrança já ocorreu no novo valor: troca os recursos agora.
                                new_option,new_cfg=billing_config(pending,normalize_billing_option(pending,camp['billing_option']))
                                conn.execute('UPDATE campaigns SET plan=?,billing_option=?,billing_amount=?,pending_plan=NULL,subscription_cancel_at_period_end=0,subscription_current_period_end=?,subscription_next_payment_at=?,subscription_change_requested_at=NULL WHERE id=?',(pending,new_option,new_cfg['amount'],next_ts,next_ts,camp['id']))
                                audit(conn,camp['company_id'],None,'plan_downgrade_confirmed','campaign',camp['id'],details=pending)
                        elif not pending and next_ts:
                            conn.execute('UPDATE campaigns SET subscription_current_period_end=? WHERE id=?',(next_ts,camp['id']))
                    elif camp['previous_subscription_id']==sub_id and mapped=='cancelled':
                        conn.execute('UPDATE campaigns SET previous_subscription_id=NULL WHERE id=?',(camp['id'],))
            return self.send_json({'ok':True})
        if path=='/api/public/contact':
            # Formulário comercial público da landing page. O campo `website` é
            # um honeypot simples para bots; para visitantes reais fica vazio.
            if str(payload.get('website') or '').strip():
                return self.send_json({'ok':True})
            name=str(payload.get('name') or '').strip()[:120]
            company=str(payload.get('company') or '').strip()[:160]
            email=str(payload.get('email') or '').strip().lower()[:180]
            phone=str(payload.get('phone') or '').strip()[:40]
            segment=str(payload.get('segment') or '').strip()[:100]
            interest=str(payload.get('interest') or '').strip()[:80]
            message=str(payload.get('message') or '').strip()[:3000]
            consent=bool(payload.get('consent'))
            if not name or not company or not email or not phone or not segment or not interest or not consent:
                return self.send_json({'ok':False,'error':'required_fields'},400)
            if not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+',email):
                return self.send_json({'ok':False,'error':'invalid_email'},400)
            cfg=global_email_config()
            print(f'[CONTACT] EMAIL_CONFIG provider={cfg.get("provider") or "auto"} source={cfg.get("source") or "unknown"} configured={email_configured(cfg)} sender={bool(cfg.get("sender_email") or cfg.get("from_addr"))}')
            if not email_configured(cfg):
                print('[CONTACT] EMAIL_NOT_CONFIGURED')
                return self.send_json({'ok':False,'error':'email_not_configured'},503)
            msg=EmailMessage()
            msg['Subject']=f'Novo contato Fidelizaê! • {company}'
            msg['To']='gustavo@agenciataboo.com.br'
            msg['Reply-To']=email
            msg.set_content(
                'Novo contato recebido pelo site do Fidelizaê!\n\n'
                f'Nome: {name}\nEmpresa: {company}\nE-mail: {email}\nWhatsApp/Celular: {phone}\n'
                f'Segmento: {segment}\nInteresse: {interest}\n\nMensagem:\n{message or "Não informada."}\n'
            )
            result=send_email_message(msg,cfg)
            if not result.get('sent'):
                reason=result.get('reason') or 'send_failed'
                if reason=='brevo_ip_blocked':
                    blocked_ip=result.get('blocked_ip') or 'nao_informado'
                    print(f'[CONTACT] BREVO_IP_BLOCKED ip={blocked_ip} status={result.get("status")} action=aguardar_autorizacao_automatica_ou_revisar_brevo_security')
                    # Código seguro para a interface: informa a causa sem expor o IP público.
                    return self.send_json({'ok':False,'error':'email_provider_ip_blocked'},503)
                print(f'[CONTACT] SEND_FAILED email={email} reason={reason} status={result.get("status")} source={result.get("source")}')
                return self.send_json({'ok':False,'error':'send_failed'},502)
            print(f'[CONTACT] SENT company={company!r} email={email}')
            return self.send_json({'ok':True})
        mweb=re.fullmatch(r'/api/integrations/ecommerce/(\d+)/([^/]+)',path)
        if mweb:
            campaign_id=int(mweb.group(1)); supplied_secret=mweb.group(2)
            with connect(DB_PATH) as conn:
                c=rowdict(conn.execute('SELECT * FROM campaigns WHERE id=? AND active=1',(campaign_id,)).fetchone())
                if not c or normalize_ecommerce_platform(c.get('ecommerce_platform'))=='none':return self.send_json({'ok':False,'error':'ecommerce_not_configured'},404)
                expected=str(c.get('ecommerce_webhook_secret') or '')
                if not expected or not hmac.compare_digest(expected,supplied_secret):return self.send_json({'ok':False,'error':'invalid_webhook_secret'},403)
                platform=normalize_ecommerce_platform(c.get('ecommerce_platform')); adapted=platform_order(payload,platform); info=ecommerce_extract(adapted,platform)
                if not info['order_id']:return self.send_json({'ok':False,'error':'order_id_required'},400)
                paid_states={'paid','approved','completed','processing','authorized'}
                reversed_states={'refunded','cancelled','canceled','voided','reversed','chargeback'}
                if info['status'] not in paid_states|reversed_states:return self.send_json({'ok':True,'ignored':True,'reason':'status_not_final','status':info['status']})
                existing=rowdict(conn.execute('SELECT * FROM ecommerce_orders WHERE campaign_id=? AND platform=? AND order_id=?',(campaign_id,platform,info['order_id'])).fetchone())
                if info['status'] in paid_states:
                    if existing and existing.get('processed_at') and not existing.get('reversed_at'):
                        return self.send_json({'ok':True,'duplicate':True,'order_id':info['order_id']})
                    member=ecommerce_find_membership(conn,campaign_id,info)
                    customer_ref=info.get('cpf') or info.get('email') or info.get('phone') or ''
                    if not member:
                        if existing: conn.execute('UPDATE ecommerce_orders SET order_status=?,customer_ref=?,total_cents=?,updated_at=? WHERE id=?',(info['status'],customer_ref,info['total_cents'],now_ts(),existing['id']))
                        else: conn.execute('INSERT INTO ecommerce_orders(campaign_id,platform,order_id,order_status,customer_ref,total_cents,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?)',(campaign_id,platform,info['order_id'],info['status'],customer_ref,info['total_cents'],now_ts(),now_ts()))
                        audit(conn,c['company_id'],None,'ecommerce_customer_not_found','campaign',campaign_id,details=f'{platform};pedido={info["order_id"]};cliente={customer_ref}',ip_address=self._ip())
                        return self.send_json({'ok':False,'error':'customer_not_found','message':'O pedido foi recebido, mas o cliente ainda não possui cartão neste programa.'},409)
                    event_ts=now_ts()
                    if c['loyalty_type']=='points':
                        rate=max(1,int(c['points_spend_cents'] or 200)); base_reward=info['total_cents']//rate; factor=active_multiplier(conn,campaign_id,event_ts) if plan_allows(conn,campaign_id,'multipliers') else 1.0; reward=int(base_reward*factor)
                        if reward<1:return self.send_json({'ok':True,'ignored':True,'reason':'purchase_below_point_rule'})
                        prev=int(member['points_balance'] or 0); new=prev+reward
                        tx_id=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(member['id'],None,None,'adjustment',reward,prev,new,0,f'ecom:{campaign_id}:{platform}:{info["order_id"]}',self._ip(),f'{platform} • Pedido {info["order_id"]} • R$ {info["total_cents"]/100:.2f} • {factor:g}x',event_ts))
                        conn.execute('UPDATE memberships SET points_balance=? WHERE id=?',(new,member['id']))
                        add_point_lot(conn,member['id'],tx_id,reward,int(c.get('points_expiry_days') or 180),event_ts)
                    else:
                        reward=1; prev=int(member['progress'] or 0); goal=max(1,int(c['goal'] or 5)); new=prev+1; rewards_delta=0
                        if new>=goal:new=0;rewards_delta=1
                        tx_id=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(member['id'],None,None,'stamp',1,prev,new,rewards_delta,f'ecom:{campaign_id}:{platform}:{info["order_id"]}',self._ip(),f'{platform} • Pedido {info["order_id"]} • R$ {info["total_cents"]/100:.2f}',event_ts))
                        conn.execute('UPDATE memberships SET progress=?,rewards_available=rewards_available+? WHERE id=?',(new,rewards_delta,member['id']))
                    record_purchase(conn,member['id'],tx_id,info['total_cents'],platform,event_ts)
                    if existing: conn.execute('UPDATE ecommerce_orders SET order_status=?,customer_ref=?,total_cents=?,reward_value=?,transaction_id=?,processed_at=?,reversed_at=NULL,reversal_transaction_id=NULL,updated_at=? WHERE id=?',(info['status'],customer_ref,info['total_cents'],reward,tx_id,now_ts(),now_ts(),existing['id']))
                    else: conn.execute('INSERT INTO ecommerce_orders(campaign_id,platform,order_id,order_status,customer_ref,total_cents,reward_value,transaction_id,processed_at,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(campaign_id,platform,info['order_id'],info['status'],customer_ref,info['total_cents'],reward,tx_id,now_ts(),now_ts(),now_ts()))
                    conn.execute("UPDATE campaigns SET ecommerce_status='connected',ecommerce_connected_at=COALESCE(ecommerce_connected_at,?) WHERE id=?",(now_iso(),campaign_id))
                    audit(conn,c['company_id'],None,'ecommerce_reward','membership',member['public_id'],details=f'{platform};pedido={info["order_id"]};valor={info["total_cents"]};recompensa={reward}',ip_address=self._ip()); notify_wallet_updates(conn,member['public_id'])
                    return self.send_json({'ok':True,'order_id':info['order_id'],'customer_name':member['customer_name'],'reward':reward,'loyalty_type':c['loyalty_type']})
                if not existing or not existing.get('processed_at') or existing.get('reversed_at'): return self.send_json({'ok':True,'ignored':True,'reason':'nothing_to_reverse'})
                tx=conn.execute('SELECT * FROM transactions WHERE id=?',(existing['transaction_id'],)).fetchone()
                if not tx:return self.send_json({'ok':False,'error':'original_transaction_not_found'},409)
                member=fetchone_for_update(conn,'SELECT * FROM memberships WHERE id=?',(tx['membership_id'],)); reward=max(0,int(existing['reward_value'] or 0))
                if c['loyalty_type']=='points':
                    prev=int(member['points_balance'] or 0); reversed_points=min(prev,reward); new=max(0,prev-reversed_points)
                    consume_point_lots(conn,member['id'],reversed_points)
                    rtx=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(member['id'],None,None,'adjustment',-reversed_points,prev,new,0,f'ecom-refund:{campaign_id}:{platform}:{info["order_id"]}',self._ip(),f'Estorno {platform} • Pedido {info["order_id"]}',now_ts()))
                    conn.execute('UPDATE memberships SET points_balance=? WHERE id=?',(new,member['id']))
                else:
                    prev=int(member['progress'] or 0); goal=max(1,int(c['goal'] or 5)); rewards_available=int(member['rewards_available'] or 0); new=prev-1; rd=0
                    if new<0:
                        new=goal-1
                        if rewards_available>0: rd=-1
                    rtx=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(member['id'],None,None,'adjustment',-1,prev,new,rd,f'ecom-refund:{campaign_id}:{platform}:{info["order_id"]}',self._ip(),f'Estorno {platform} • Pedido {info["order_id"]}',now_ts()))
                    new_rewards=max(0,rewards_available+rd); conn.execute('UPDATE memberships SET progress=?,rewards_available=? WHERE id=?',(new,new_rewards,member['id']))
                conn.execute('UPDATE ecommerce_orders SET order_status=?,reversal_transaction_id=?,reversed_at=?,updated_at=? WHERE id=?',(info['status'],rtx,now_ts(),now_ts(),existing['id']))
                audit(conn,c['company_id'],None,'ecommerce_reward_reversed','membership',member['public_id'],details=f'{platform};pedido={info["order_id"]};recompensa=-{reward}',ip_address=self._ip()); notify_wallet_updates(conn,member['public_id'])
                return self.send_json({'ok':True,'reversed':True,'order_id':info['order_id'],'reward_reversed':reward,'loyalty_type':c['loyalty_type']})

        if path in ['/api/login/2fa','/login/2fa']:
            code=str(payload.get('code') or '').strip()
            challenge_cookie=self._cookies().get('clube_2fa_challenge')
            raw=challenge_cookie.value if challenge_cookie else ''
            if not raw or not self._rate_ok('login-2fa-ip',10,600,self._ip(),1800): return
            th=hashlib.sha256(raw.encode()).hexdigest()
            with connect(DB_PATH) as conn:
                row=fetchone_for_update(conn,"SELECT ac.token_hash,ac.user_id,ac.expires_at,ac.attempts,u.* FROM auth_challenges ac JOIN users u ON u.id=ac.user_id WHERE ac.token_hash=?",(th,))
                if not row or int(row['expires_at'] or 0)<now_ts() or int(row['attempts'] or 0)>=6 or not row['active']:
                    if row: conn.execute('DELETE FROM auth_challenges WHERE token_hash=?',(th,))
                    if path=='/login/2fa': return self.send_redirect('/login?error=2fa_expired',303,{'Set-Cookie':_clear_cookie('clube_2fa_challenge')})
                    return self.send_json({'ok':False,'error':'two_factor_challenge_invalid'},401,{'Set-Cookie':_clear_cookie('clube_2fa_challenge')})
                secret=decrypt_secret(row['totp_secret_enc']) if row['totp_secret_enc'] else ''
                if not row['totp_enabled'] or not verify_totp(secret,code):
                    conn.execute('UPDATE auth_challenges SET attempts=attempts+1 WHERE token_hash=?',(th,))
                    audit(conn,row['company_id'],row['id'],'login_2fa_failed',details='totp',ip_address=self._ip())
                    if path=='/login/2fa': return self.send_redirect('/login/2fa?error=1')
                    return self.send_json({'ok':False,'error':'two_factor_invalid'},401)
                conn.execute('DELETE FROM auth_challenges WHERE token_hash=?',(th,))
                token,csrf=create_session(conn,row['id'])
                persistent_rate_reset('login-account:'+hashlib.sha256(str(row['email']).lower().encode()).hexdigest()[:32])
                audit(conn,row['company_id'],row['id'],'login_success','user',row['id'],details='2fa',ip_address=self._ip())
                auth_cookies=[_session_cookie(token),_clear_cookie('clube_2fa_challenge')]
                if path=='/login/2fa': return self.send_redirect('/manager' if row['role']=='manager' else '/attendant',303,{'Set-Cookie':auth_cookies})
                return self.send_json({'ok':True,'role':row['role'],'csrf':csrf},200,{'Set-Cookie':auth_cookies})

        if path in ['/api/login','/login']:
            email=normalize_email(payload.get('email')) or str(payload.get('email','')).lower().strip()[:160]
            password=str(payload.get('password',''))
            if not self._rate_ok('login-ip',12,300,self._ip(),900): return
            if not self._rate_ok('login-account',6,900,email or 'invalid',1800): return
            with connect(DB_PATH) as conn:
                u=conn.execute('SELECT * FROM users WHERE email=? AND active=1',(email,)).fetchone() if email else None
                password_ok=verify_password(password,u['password_hash'] if u else DUMMY_PASSWORD_HASH)
                admin_email=os.environ.get('CLUBE_ADMIN_EMAIL','').strip().lower()
                admin_password=os.environ.get('CLUBE_ADMIN_PASSWORD','').strip()
                allow_repair=os.environ.get('CLUBE_ALLOW_ADMIN_REPAIR','0' if (os.environ.get('APP_ENV') or '').lower()=='production' else '1')=='1'
                admin_login=bool(allow_repair and u and admin_email and admin_password and email==admin_email and hmac.compare_digest(password,admin_password))
                if admin_login:
                    # CLUBE_ADMIN_PASSWORD é credencial de bootstrap/recuperação de perfil, não uma
                    # senha mestre permanente. Nunca regrave password_hash aqui: isso anulava uma
                    # troca de senha feita pelo próprio administrador no Painel Fidelizaê!.
                    needs_repair=(u['role']!='manager') or (u['campaign_id'] is not None) or (not u['active'])
                    if needs_repair:
                        conn.execute("UPDATE users SET role='manager',campaign_id=NULL,active=1 WHERE id=?",(u['id'],))
                        u=conn.execute('SELECT * FROM users WHERE id=?',(u['id'],)).fetchone()
                        print(f'[AUTH] ADMIN_PROFILE_REPAIRED user={_email_tag(email)}')
                    # Permite a credencial de recuperação apenas quando explicitamente habilitada,
                    # sem alterar a senha persistida. Em produção o padrão continua desabilitado.
                    password_ok=True
                if not u or not password_ok:
                    print(f'[AUTH] LOGIN_FAILED user={_email_tag(email)}')
                    audit(conn,u['company_id'] if u else None,u['id'] if u else None,'login_failed',details='credential_rejected',ip_address=self._ip())
                    if path=='/login': return self.send_redirect('/login?error=1')
                    return self.send_json({'ok':False,'error':'invalid_credentials'},401)
                if u['role']=='attendant' and u['campaign_id']:
                    reconcile_campaign_billing(conn,u['campaign_id'])
                    camp=conn.execute('SELECT active,subscription_status FROM campaigns WHERE id=?',(u['campaign_id'],)).fetchone()
                    if not camp or not camp['active']:
                        if path=='/login': return self.send_redirect('/login?error=subscription_expired')
                        return self.send_json({'ok':False,'error':'subscription_expired'},403)
                # 2FA é exigido quando foi ativado pelo administrador/cliente-admin.
                if u['totp_enabled']:
                    challenge=create_2fa_challenge(conn,u['id'])
                    audit(conn,u['company_id'],u['id'],'login_2fa_challenge','user',u['id'],ip_address=self._ip())
                    if path=='/login': return self.send_redirect('/login/2fa',303,{'Set-Cookie':_challenge_cookie(challenge)})
                    return self.send_json({'ok':True,'requires_2fa':True},202,{'Set-Cookie':_challenge_cookie(challenge)})
                persistent_rate_reset('login-account:'+hashlib.sha256(str(email).lower().encode()).hexdigest()[:32])
                token,csrf=create_session(conn,u['id']); audit(conn,u['company_id'],u['id'],'login_success','user',u['id'],details='password',ip_address=self._ip())
                print(f'[AUTH] LOGIN_SUCCESS user={_email_tag(email)} role={u["role"]}')
                cookie=_session_cookie(token)
                if path=='/login': return self.send_redirect('/manager' if u['role']=='manager' else '/attendant',303,{'Set-Cookie':cookie})
                return self.send_json({'ok':True,'role':u['role'],'csrf':csrf},200,{'Set-Cookie':cookie})
        if path == '/api/logout':
            with connect(DB_PATH) as conn:
                token=self._session_token(); s=self._session(conn)
                if token: conn.execute('DELETE FROM sessions WHERE token=?',(token,))
                if s: audit(conn,s['company_id'],s['user_id'],'logout',ip_address=self._ip())
            return self.send_json({'ok':True},200,{'Set-Cookie':[_clear_cookie(SESSION_COOKIE),_clear_cookie('clube_2fa_challenge')]})
        if path in ['/api/join','/join']:
            if not self._rate_ok('join',12,600,self._ip(),900): return
            code=str(payload.get('campaign_code','')).upper().strip()
            name=str(payload.get('name','')).strip()[:80]
            email=normalize_email(payload.get('email'))
            phone=normalize_phone(payload.get('phone'))
            birth_date=normalize_birth_date(payload.get('birth_date'))
            gender=str(payload.get('gender') or '').strip().lower(); gender=gender if gender in ('female','male','other','prefer_not') else ''
            cpf=normalize_cpf(payload.get('cpf')); privacy_ok=str(payload.get('privacy_consent','')).lower() in ('1','true','on','yes'); marketing_email=str(payload.get('marketing_email','')).lower() in ('1','true','on','yes'); marketing_whatsapp=str(payload.get('marketing_whatsapp','')).lower() in ('1','true','on','yes')
            if len(name)<2 or not email or not phone or not birth_date or not cpf or not privacy_ok:
                error='invalid_customer_data'
                if len(name)<2: error='invalid_name'
                elif not email: error='invalid_email'
                elif not phone: error='invalid_phone'
                elif not birth_date: error='invalid_birth_date'
                elif not cpf: error='invalid_cpf'
                elif not privacy_ok: error='privacy_consent_required'
                return self.send_redirect('/join?campaign='+urllib.parse.quote(code or 'CAFE5')+'&error='+urllib.parse.quote(error)) if path=='/join' else self.send_json({'ok':False,'error':error},400)
            with connect(DB_PATH) as conn:
                c=conn.execute('SELECT * FROM campaigns WHERE code=? AND active=1',(code,)).fetchone()
                if not c:
                    return self.send_redirect('/join?campaign='+urllib.parse.quote(code or 'CAFE5')+'&error=campaign_not_found') if path=='/join' else self.send_json({'ok':False,'error':'campaign_not_found'},404)
                existing_customer=conn.execute('''SELECT cu.id FROM customers cu JOIN memberships m ON m.customer_id=cu.id
                    WHERE m.campaign_id=? AND cu.cpf_hash=? LIMIT 1''',(c['id'],pii_lookup_hash(cpf,'cpf'))).fetchone()
                customer_id=existing_customer['id'] if existing_customer else None
                duplicate_contact=conn.execute('''SELECT cu.id FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND (lower(cu.email)=lower(?) OR cu.phone_hash=?) AND (cu.cpf_hash IS NULL OR cu.cpf_hash<>?) LIMIT 1''',(c['id'],email,pii_lookup_hash(phone,'phone'),pii_lookup_hash(cpf,'cpf'))).fetchone()
                if duplicate_contact:
                    return self.send_redirect('/join?campaign='+urllib.parse.quote(code)+'&error=duplicate_contact') if path=='/join' else self.send_json({'ok':False,'error':'duplicate_contact'},409)
                if customer_id is None:
                    pii=protected_customer_pii(phone,cpf)
                    customer_id=insert_id(conn,'INSERT INTO customers(name,contact,email,phone,phone_enc,phone_hash,birth_date,gender,cpf,cpf_enc,cpf_hash,privacy_accepted_at,marketing_email,marketing_whatsapp,marketing_accepted_at,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                        (name,email,email,None,pii['phone_enc'],pii['phone_hash'],birth_date,gender or None,None,pii['cpf_enc'],pii['cpf_hash'],now_ts(),1 if marketing_email else 0,1 if marketing_whatsapp else 0,now_ts() if (marketing_email or marketing_whatsapp) else None,now_ts()))
                else:
                    pii=protected_customer_pii(phone,cpf)
                    conn.execute('UPDATE customers SET name=?,contact=?,email=?,phone=NULL,phone_enc=?,phone_hash=?,birth_date=?,gender=?,cpf=NULL,cpf_enc=?,cpf_hash=?,privacy_accepted_at=COALESCE(privacy_accepted_at,?),marketing_email=?,marketing_whatsapp=?,marketing_accepted_at=? WHERE id=?',
                        (name,email,email,pii['phone_enc'],pii['phone_hash'],birth_date,gender or None,pii['cpf_enc'],pii['cpf_hash'],now_ts(),1 if marketing_email else 0,1 if marketing_whatsapp else 0,now_ts() if (marketing_email or marketing_whatsapp) else None,customer_id))
                existing=conn.execute('SELECT public_id FROM memberships WHERE customer_id=? AND campaign_id=?',(customer_id,c['id'])).fetchone()
                if existing:
                    return self.send_redirect('/card?id='+urllib.parse.quote(existing['public_id'])) if path=='/join' else self.send_json({'ok':True,'public_id':existing['public_id'],'existing':True})
                plan=normalize_plan(c['plan'] if 'plan' in c.keys() else 'pro'); limit=PLAN_FEATURES[plan]['client_limit']; current=conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND status='active'",(c['id'],)).fetchone()['n'];
                if limit and current>=limit: return self.send_redirect('/join?campaign='+urllib.parse.quote(code)+'&error=plan_client_limit') if path=='/join' else self.send_json({'ok':False,'error':'plan_client_limit','limit':limit},403)
                public_id='mem_'+random_token(10); qr_token=random_token(24)
                conn.execute('INSERT INTO memberships(customer_id,campaign_id,public_id,qr_token,created_at) VALUES(?,?,?,?,?)',(customer_id,c['id'],public_id,qr_token,now_ts()))
                print(f'[JOIN] CREATED public_id={public_id} campaign={code} name={name!r}')
                audit(conn,c['company_id'],None,'customer_join','membership',public_id,details=name,ip_address=self._ip())
                welcome_result={'queued':False,'skipped':True,'reason':'email_provider_not_configured'}
                if email_configured(email_config_for_client(conn,c['id'])):
                    qid=enqueue_message(conn,c['id'],'customer_welcome',email,{'name':name,'email':email,'public_id':public_id}); welcome_result={'queued':True,'queue_id':qid}
                    audit(conn,c['company_id'],None,'customer_welcome_queued','membership',public_id,details=email,ip_address=self._ip())
                return self.send_redirect('/card?id='+urllib.parse.quote(public_id)) if path=='/join' else self.send_json({'ok':True,'public_id':public_id,'existing':False,'welcome_email':welcome_result})
        if path == '/api/forgot-password':
            email=normalize_email(payload.get('email'))
            if not self._rate_ok('forgot-password-ip',5,900,self._ip(),1800): return
            if email and not self._rate_ok('forgot-password-account',3,1800,email,3600): return
            if not email:
                return self.send_json({'ok':False,'error':'invalid_email'},400)
            with connect(DB_PATH) as conn:
                u=conn.execute("SELECT id,company_id,email,role,active,campaign_id FROM users WHERE email=? AND active=1",(email,)).fetchone()
                # Não revelamos se o endereço existe. A redefinição usa token único de 30 minutos.
                if u and u['role'] in ('attendant','manager'):
                    raw=random_token(32); token_hash=hashlib.sha256(raw.encode()).hexdigest(); ts=now_ts()
                    smtp_cfg=global_email_config()
                    if email_configured(smtp_cfg):
                        conn.execute('DELETE FROM password_reset_tokens WHERE user_id=? OR expires_at<?',(u['id'],ts))
                        conn.execute('INSERT INTO password_reset_tokens(token_hash,user_id,expires_at,created_at) VALUES(?,?,?,?)',(token_hash,u['id'],ts+1800,ts))
                        qid=enqueue_message(conn,None,'password_recovery',email,{'token':raw,'user_id':u['id']})
                        audit(conn,u['company_id'],u['id'],'password_recovery_queued','user',u['id'],details=f'queue={qid}',ip_address=self._ip())
                    else:
                        print('[AUTH] PASSWORD_RECOVERY_PROVIDER_UNAVAILABLE user=%s' % _email_tag(email))
                return self.send_json({'ok':True,'message':'Se o e-mail estiver cadastrado, enviaremos um link de redefinição.'})

        if path == '/api/reset-password':
            if not self._rate_ok('reset-password',6,900): return
            token=str(payload.get('token') or ''); password=str(payload.get('password') or '').strip()
            if not password_is_strong(password,12):return self.send_json({'ok':False,'error':'invalid_new_password'},400)
            th=hashlib.sha256(token.encode()).hexdigest()
            with connect(DB_PATH) as conn:
                r=conn.execute('SELECT * FROM password_reset_tokens WHERE token_hash=? AND used_at IS NULL AND expires_at>=?',(th,now_ts())).fetchone()
                if not r:return self.send_json({'ok':False,'error':'reset_token_invalid'},400)
                conn.execute('UPDATE users SET password_hash=? WHERE id=?',(hash_password(password),r['user_id'])); conn.execute('UPDATE password_reset_tokens SET used_at=? WHERE token_hash=?',(now_ts(),th)); conn.execute('DELETE FROM sessions WHERE user_id=?',(r['user_id'],)); u=conn.execute('SELECT company_id FROM users WHERE id=?',(r['user_id'],)).fetchone(); audit(conn,u['company_id'] if u else None,r['user_id'],'password_change','user',r['user_id'],details='reset_token',ip_address=self._ip())
            return self.send_json({'ok':True})

        if path == '/api/privacy/delete':
            if not self._rate_ok('privacy-delete',5,900): return
            public_id=str(payload.get('id','')).strip(); cpf=normalize_cpf(payload.get('cpf'))
            if not cpf:return self.send_json({'ok':False,'error':'invalid_cpf'},400)
            with connect(DB_PATH) as conn:
                row=conn.execute('''SELECT m.id membership_id,m.customer_id,c.company_id FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=? AND cu.cpf_hash=?''',(public_id,pii_lookup_hash(cpf,'cpf'))).fetchone()
                if not row:return self.send_json({'ok':False,'error':'not_found'},404)
                audit(conn,row['company_id'],None,'lgpd_delete','customer',row['customer_id'],details=public_id,ip_address=self._ip())
                conn.execute('DELETE FROM memberships WHERE id=?',(row['membership_id'],))
                conn.execute('DELETE FROM customers WHERE id=? AND id NOT IN (SELECT customer_id FROM memberships)',(row['customer_id'],))
            return self.send_json({'ok':True})
        if path == '/api/privacy/preferences':
            if not self._rate_ok('privacy-preferences',20,900): return
            public_id=str(payload.get('id','')).strip(); cpf=normalize_cpf(payload.get('cpf'))
            if not cpf:return self.send_json({'ok':False,'error':'invalid_cpf'},400)
            marketing_email=1 if str(payload.get('marketing_email','')).lower() in ('1','true','on','yes') else 0
            marketing_whatsapp=1 if str(payload.get('marketing_whatsapp','')).lower() in ('1','true','on','yes') else 0
            with connect(DB_PATH) as conn:
                row=conn.execute('''SELECT m.customer_id,c.company_id FROM memberships m JOIN customers cu ON cu.id=m.customer_id JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=? AND cu.cpf_hash=?''',(public_id,pii_lookup_hash(cpf,'cpf'))).fetchone()
                if not row:return self.send_json({'ok':False,'error':'not_found'},404)
                conn.execute('UPDATE customers SET marketing_email=?,marketing_whatsapp=?,marketing_accepted_at=? WHERE id=?',(marketing_email,marketing_whatsapp,now_ts() if (marketing_email or marketing_whatsapp) else None,row['customer_id']))
                audit(conn,row['company_id'],None,'privacy_preferences','customer',row['customer_id'],details=f'email={marketing_email};whatsapp={marketing_whatsapp}',ip_address=self._ip())
            return self.send_json({'ok':True,'marketing_email':bool(marketing_email),'marketing_whatsapp':bool(marketing_whatsapp)})
        if path == '/api/card/nps':
            if not self._rate_ok('card-nps',8,900): return
            public_id=str(payload.get('id') or '').strip()
            try: score=int(payload.get('score'))
            except (TypeError,ValueError): score=-1
            comment=str(payload.get('comment') or '').strip()[:500]
            if score<0 or score>10:return self.send_json({'ok':False,'error':'invalid_score'},400)
            with connect(DB_PATH) as conn:
                m=conn.execute("SELECT m.id,m.campaign_id,c.plan FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=? AND m.status='active'",(public_id,)).fetchone()
                if not m:return self.send_json({'ok':False,'error':'card_not_found'},404)
                if not PLAN_FEATURES[normalize_plan(m['plan'])]['nps']:return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                # Só aceita NPS após existir alguma interação real no programa.
                if not conn.execute('SELECT 1 FROM transactions WHERE membership_id=? LIMIT 1',(m['id'],)).fetchone():
                    return self.send_json({'ok':False,'error':'nps_not_due'},409)
                recent=conn.execute('SELECT 1 FROM nps_responses WHERE membership_id=? AND created_at>=? LIMIT 1',(m['id'],now_ts()-90*86400)).fetchone()
                if recent:return self.send_json({'ok':False,'error':'nps_already_answered'},409)
                insert_id(conn,"INSERT INTO nps_responses(campaign_id,membership_id,score,comment,created_at) VALUES(?,?,?,?,?)",(m['campaign_id'],m['id'],score,comment,now_ts()))
            return self.send_json({'ok':True})

        if path == '/api/apple-wallet/v1/log':
            # Apple may post diagnostic messages from Wallet; do not require app session/CSRF.
            return self.send_json({'ok':True})
        if path.startswith('/api/apple-wallet/v1/devices/') and '/registrations/' in path:
            parts=path.split('/'); device=urllib.parse.unquote(parts[5]) if len(parts)>5 else ''; public_id=urllib.parse.unquote(parts[-1])
            auth=(self.headers.get('Authorization') or '').replace('ApplePass ','').strip()
            if not hmac.compare_digest(auth,apple_auth_token(public_id)):return self.send_json({'ok':False,'error':'unauthorized'},401)
            push=str(payload.get('pushToken','')).strip()
            if not device or not push:return self.send_json({'ok':False,'error':'invalid_registration'},400)
            with connect(DB_PATH) as conn:
                m=conn.execute('SELECT id FROM memberships WHERE public_id=?',(public_id,)).fetchone()
                if not m:return self.send_json({'ok':False,'error':'invalid_registration'},400)
                try: conn.execute('INSERT INTO wallet_registrations(membership_id,device_library_id,push_token,created_at) VALUES(?,?,?,?)',(m['id'],device,push,now_ts()))
                except integrity_errors(): conn.execute('UPDATE wallet_registrations SET push_token=?,created_at=? WHERE membership_id=? AND device_library_id=?',(push,now_ts(),m['id'],device))
            return self.send_json({'ok':True},201)

        if path == '/api/v1/customers':
            with connect(DB_PATH) as conn:
                ctx=self._api_context(conn)
                if not ctx:return
                name=str(payload.get('name') or '').strip()[:120]
                email=normalize_email(payload.get('email')) if payload.get('email') else ''
                phone=str(payload.get('phone') or '').strip()[:40]
                cpf=normalize_cpf(payload.get('cpf')) if payload.get('cpf') else ''
                birth=str(payload.get('birth_date') or '').strip()[:10]
                if len(name)<2:return self.send_json({'ok':False,'error':'invalid_name'},400)
                plan=campaign_plan(conn,ctx['campaign_id']); limit=int(PLAN_FEATURES[plan].get('client_limit') or 0)
                if limit and int(conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=?",(ctx['campaign_id'],)).fetchone()['n'] or 0)>=limit:
                    return self.send_json({'ok':False,'error':'client_limit_reached'},409)
                existing=None
                if cpf: existing=conn.execute("SELECT cu.id FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND cu.cpf_hash=?",(ctx['campaign_id'],pii_lookup_hash(cpf,'cpf'))).fetchone()
                if not existing and email: existing=conn.execute("SELECT cu.id FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND lower(cu.email)=lower(?)",(ctx['campaign_id'],email)).fetchone()
                if existing:return self.send_json({'ok':False,'error':'customer_exists'},409)
                pii=protected_customer_pii(phone,cpf)
                customer_id=insert_id(conn,"INSERT INTO customers(name,contact,email,phone,phone_enc,phone_hash,birth_date,cpf,cpf_enc,cpf_hash,privacy_accepted_at,marketing_email,marketing_whatsapp,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(name,email,email,None,pii['phone_enc'],pii['phone_hash'],birth or None,None,pii['cpf_enc'],pii['cpf_hash'],None,0,0,now_ts()))
                public_id='mem_'+secrets.token_urlsafe(10); qr_token=random_token(18)
                mid=insert_id(conn,"INSERT INTO memberships(customer_id,campaign_id,public_id,qr_token,created_at) VALUES(?,?,?,?,?)",(customer_id,ctx['campaign_id'],public_id,qr_token,now_ts()))
                audit(conn,ctx['company_id'],None,'api_customer_create','customer',customer_id,details=public_id,ip_address=self._ip(),campaign_id=ctx['campaign_id'])
                queue_webhook_event(conn,ctx['campaign_id'],'customer.created',{'public_id':public_id,'name':name,'email':email,'phone':phone})
                return self.send_json({'ok':True,'data':{'public_id':public_id,'customer_id':customer_id}},201)
        if path == '/api/v1/purchases':
            with connect(DB_PATH) as conn:
                ctx=self._api_context(conn)
                if not ctx:return
                public_id=str(payload.get('public_id') or '').strip()
                try: amount=max(0,int(payload.get('amount_cents') or 0))
                except Exception:return self.send_json({'ok':False,'error':'invalid_amount'},400)
                idem_key=str(payload.get('idempotency_key') or '').strip()[:160] or None
                m=conn.execute("SELECT m.*,c.goal,c.loyalty_type,c.points_spend_cents,c.points_expiry_days FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=? AND m.campaign_id=?",(public_id,ctx['campaign_id'])).fetchone()
                if not m:return self.send_json({'ok':False,'error':'customer_not_found'},404)
                if m['status']!='active':return self.send_json({'ok':False,'error':'customer_blocked'},409)
                if idem_key:
                    prev=conn.execute("SELECT id,value,new_progress FROM transactions WHERE idempotency_key=?",(idem_key,)).fetchone()
                    if prev:return self.send_json({'ok':True,'idempotent':True,'data':rowdict(prev)})
                if m['loyalty_type']=='points':
                    value=max(0,amount//max(1,int(m['points_spend_cents'] or 200)))
                    previous=int(m['points_balance'] or 0); new=previous+value
                    tid=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,device_id,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(m['id'],None,'adjustment',value,previous,new,0,idem_key,'api',self._ip(),'Compra via API',now_ts()))
                    conn.execute("UPDATE memberships SET points_balance=? WHERE id=?",(new,m['id']))
                    if value:add_point_lot(conn,m['id'],tid,value,int(m['points_expiry_days'] or 180),now_ts())
                else:
                    previous=int(m['progress'] or 0); goal=max(1,int(m['goal'] or 1)); new=previous+1; reward=0
                    if new>=goal:new=0; reward=1
                    tid=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,device_id,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(m['id'],None,'stamp',1,previous,new,reward,idem_key,'api',self._ip(),'Compra via API',now_ts()))
                    conn.execute("UPDATE memberships SET progress=?,rewards_available=rewards_available+? WHERE id=?",(new,reward,m['id']))
                    value=1
                record_purchase(conn,m['id'],tid,amount,'api')
                queue_webhook_event(conn,ctx['campaign_id'],'purchase.created',{'public_id':public_id,'amount_cents':amount,'earned':value,'loyalty_type':m['loyalty_type']})
                return self.send_json({'ok':True,'data':{'transaction_id':tid,'earned':value,'loyalty_type':m['loyalty_type']}})
        if path == '/api/v1/redemptions':
            with connect(DB_PATH) as conn:
                ctx=self._api_context(conn)
                if not ctx:return
                public_id=str(payload.get('public_id') or '').strip()
                m=conn.execute("SELECT m.*,c.goal,c.loyalty_type FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=? AND m.campaign_id=?",(public_id,ctx['campaign_id'])).fetchone()
                if not m:return self.send_json({'ok':False,'error':'customer_not_found'},404)
                if m['loyalty_type']=='points':
                    try: reward_id=int(payload.get('reward_id') or 0)
                    except Exception: reward_id=0
                    r=conn.execute("SELECT id,name,points_cost FROM reward_catalog WHERE id=? AND campaign_id=? AND active=1",(reward_id,ctx['campaign_id'])).fetchone()
                    if not r:return self.send_json({'ok':False,'error':'reward_not_found'},404)
                    cost=int(r['points_cost']); balance=int(m['points_balance'] or 0)
                    if balance<cost:return self.send_json({'ok':False,'error':'insufficient_points'},409)
                    new=balance-cost
                    tid=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,type,value,previous_progress,new_progress,rewards_delta,device_id,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(m['id'],None,'redeem',-cost,balance,new,0,'api',self._ip(),r['name'],now_ts()))
                    conn.execute("UPDATE memberships SET points_balance=? WHERE id=?",(new,m['id'])); consume_point_lots(conn,m['id'],cost)
                    try:conn.execute("INSERT INTO reward_redemptions(reward_id,membership_id,user_id,points_cost,created_at) VALUES(?,?,?,?,?)",(r['id'],m['id'],None,cost,now_ts()))
                    except Exception:pass
                    reward_name=r['name']
                else:
                    available=int(m['rewards_available'] or 0)
                    if available<1:return self.send_json({'ok':False,'error':'reward_not_available'},409)
                    tid=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,type,value,previous_progress,new_progress,rewards_delta,device_id,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(m['id'],None,'redeem',0,int(m['progress'] or 0),int(m['progress'] or 0),-1,'api',self._ip(),'Resgate via API',now_ts()))
                    conn.execute("UPDATE memberships SET rewards_available=rewards_available-1 WHERE id=?",(m['id'],)); reward_name='Recompensa'
                queue_webhook_event(conn,ctx['campaign_id'],'reward.redeemed',{'public_id':public_id,'reward':reward_name,'transaction_id':tid})
                return self.send_json({'ok':True,'data':{'transaction_id':tid,'reward':reward_name}})

        with connect(DB_PATH) as conn:
            s=self._require_auth(conn)
            if not s: return
            if not self._require_csrf(s,payload): return self.send_json({'ok':False,'error':'csrf_failed'},403)
            if path == '/api/manager/sentry-test':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self._rate_ok('manager-sentry-test',5,900,self._ip(),1800): return
                if not SENTRY_ENABLED or sentry_sdk is None:
                    return self.send_json({'ok':False,'error':'sentry_not_configured'},503)
                try:
                    raise RuntimeError(f'Teste controlado do Sentry - Fidelizaê! {VERSION}')
                except RuntimeError as exc:
                    try:
                        with sentry_sdk.new_scope() as scope:
                            scope.set_tag('fidelizae_test','manager_controlled')
                            scope.set_tag('release_version',VERSION)
                            scope.set_context('test',{'source':'manager_diagnostics','user_id':int(s['user_id'])})
                            event_id=sentry_sdk.capture_exception(exc)
                        sentry_sdk.flush(timeout=2.0)
                    except Exception as sentry_exc:
                        print(f'[SENTRY] falha no teste controlado: {sentry_exc}')
                        return self.send_json({'ok':False,'error':'sentry_test_failed'},502)
                audit(conn,s['company_id'],s['user_id'],'sentry_test','platform',None,details=f'event_id={event_id or ""}',ip_address=self._ip())
                return self.send_json({'ok':True,'event_id':str(event_id or ''),'version':VERSION})
            if path == '/api/manager/backup':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self._rate_ok('manager-backup',5,900,self._ip(),1800): return
                password=str(payload.get('password') or '')
                u=conn.execute("SELECT id,password_hash FROM users WHERE id=? AND role='manager' AND active=1",(s['user_id'],)).fetchone()
                if not u or not verify_password(password,u['password_hash']):
                    audit(conn,s['company_id'],s['user_id'],'platform_backup_denied','backup',None,details='reauth_failed',ip_address=self._ip())
                    return self.send_json({'ok':False,'error':'invalid_password'},403)
                backup=build_platform_backup(conn)
                ok,reason=verify_platform_backup(backup)
                if not ok: return self.send_json({'ok':False,'error':reason},500)
                audit(conn,s['company_id'],s['user_id'],'platform_backup_download','backup',None,details=f'tables={len(backup["tables"])};sha256={backup["sha256"][:12]}',ip_address=self._ip())
                data=json.dumps(backup,ensure_ascii=False,indent=2).encode('utf-8')
                return self.send_bytes(data,'application/json; charset=utf-8',headers={'Content-Disposition':f'attachment; filename="fidelizae-backup-{datetime.now().strftime("%Y%m%d-%H%M")}.json"','Cache-Control':'no-store'})
            if path == '/api/admin/customers/import/preview':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                try: rows=_parse_import_file(payload.get('filename'),payload.get('data_base64'))
                except (ValueError,binascii.Error) as exc:return self.send_json({'ok':False,'error':str(exc)},400)
                existing_cpf={str(r['cpf_hash']) for r in conn.execute("SELECT cu.cpf_hash FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND cu.cpf_hash IS NOT NULL",(s['campaign_id'],)).fetchall() if r['cpf_hash']}
                existing_email={str(r['email']).lower() for r in conn.execute("SELECT cu.email FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND cu.email IS NOT NULL",(s['campaign_id'],)).fetchall() if r['email']}
                seen_cpf=set();seen_email=set()
                for r in rows:
                    cpf_h=pii_lookup_hash(r['cpf'],'cpf') if r['cpf'] else None
                    duplicate=bool((r['cpf'] and (cpf_h in existing_cpf or cpf_h in seen_cpf)) or (r['email'] and (r['email'].lower() in existing_email or r['email'].lower() in seen_email)))
                    r['duplicate']=duplicate
                    if r['cpf']:seen_cpf.add(cpf_h)
                    if r['email']:seen_email.add(r['email'].lower())
                valid=sum(1 for r in rows if not r['errors'] and not r['duplicate'])
                return self.send_json({'ok':True,'rows':rows,'summary':{'total':len(rows),'valid':valid,'duplicates':sum(1 for r in rows if r['duplicate']),'invalid':sum(1 for r in rows if r['errors'])}})
            if path == '/api/admin/customers/import/commit':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                rows=payload.get('rows') if isinstance(payload.get('rows'),list) else []
                if len(rows)>10000:return self.send_json({'ok':False,'error':'too_many_rows'},400)
                campaign=conn.execute("SELECT loyalty_type,goal,points_expiry_days FROM campaigns WHERE id=?",(s['campaign_id'],)).fetchone()
                plan=campaign_plan(conn,s['campaign_id']); limit=int(PLAN_FEATURES[plan].get('client_limit') or 0)
                current_count=int(conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=?",(s['campaign_id'],)).fetchone()['n'] or 0)
                created=skipped=0
                for r in rows:
                    if limit and current_count+created>=limit:
                        skipped+=1; continue
                    if r.get('errors') or r.get('duplicate'):skipped+=1;continue
                    name=str(r.get('name') or '').strip()[:120]; email=normalize_email(r.get('email')) if r.get('email') else ''; cpf=normalize_cpf(r.get('cpf')) if r.get('cpf') else ''; phone=str(r.get('phone') or '').strip()[:40]; birth=str(r.get('birth_date') or '').strip()[:10] or None
                    if len(name)<2:skipped+=1;continue
                    if cpf and conn.execute("SELECT 1 FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND cu.cpf_hash=?",(s['campaign_id'],pii_lookup_hash(cpf,'cpf'))).fetchone():skipped+=1;continue
                    if email and conn.execute("SELECT 1 FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND lower(cu.email)=lower(?)",(s['campaign_id'],email)).fetchone():skipped+=1;continue
                    pii=protected_customer_pii(phone,cpf)
                    customer_id=insert_id(conn,"INSERT INTO customers(name,contact,email,phone,phone_enc,phone_hash,birth_date,cpf,cpf_enc,cpf_hash,privacy_accepted_at,marketing_email,marketing_whatsapp,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(name,email,email,None,pii['phone_enc'],pii['phone_hash'],birth,None,pii['cpf_enc'],pii['cpf_hash'],None,0,0,now_ts()))
                    public_id='mem_'+secrets.token_urlsafe(10); mid=insert_id(conn,"INSERT INTO memberships(customer_id,campaign_id,public_id,qr_token,created_at) VALUES(?,?,?,?,?)",(customer_id,s['campaign_id'],public_id,random_token(18),now_ts()))
                    initial=max(0,int(r.get('initial_balance') or 0))
                    if initial:
                        if campaign['loyalty_type']=='points':
                            tid=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,type,value,previous_progress,new_progress,note,created_at) VALUES(?,?,?,?,?,?,?,?)",(mid,s['user_id'],'adjustment',initial,0,initial,'Saldo inicial importado',now_ts()))
                            conn.execute("UPDATE memberships SET points_balance=? WHERE id=?",(initial,mid)); add_point_lot(conn,mid,tid,initial,int(campaign['points_expiry_days'] or 180),now_ts())
                        else:
                            goal=max(1,int(campaign['goal'] or 1)); progress=initial%goal; rewards=initial//goal
                            conn.execute("UPDATE memberships SET progress=?,rewards_available=? WHERE id=?",(progress,rewards,mid))
                            insert_id(conn,"INSERT INTO transactions(membership_id,user_id,type,value,previous_progress,new_progress,rewards_delta,note,created_at) VALUES(?,?,?,?,?,?,?,?,?)",(mid,s['user_id'],'adjustment',initial,0,progress,rewards,'Saldo inicial importado',now_ts()))
                    audit(conn,s['company_id'],s['user_id'],'customer_import','customer',customer_id,details=public_id,ip_address=self._ip(),campaign_id=s['campaign_id'])
                    created+=1
                return self.send_json({'ok':True,'created':created,'skipped':skipped})
            if path == '/api/admin/integrations/api-key/create':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'advanced'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                name=str(payload.get('name') or 'Integração').strip()[:80] or 'Integração'
                token='fiz_'+secrets.token_urlsafe(32); digest=hashlib.sha256(token.encode()).hexdigest()
                kid=insert_id(conn,"INSERT INTO api_keys(campaign_id,name,token_hash,token_prefix,created_at) VALUES(?,?,?,?,?)",(s['campaign_id'],name,digest,token[:12],now_ts()))
                audit(conn,s['company_id'],s['user_id'],'api_key_create','api_key',kid,details=name,ip_address=self._ip(),campaign_id=s['campaign_id'])
                return self.send_json({'ok':True,'id':kid,'token':token,'warning':'O token é exibido uma única vez.'})
            if path == '/api/admin/integrations/api-key/revoke':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'advanced'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                kid=int(payload.get('id') or 0);conn.execute("UPDATE api_keys SET active=0 WHERE id=? AND campaign_id=?",(kid,s['campaign_id']))
                audit(conn,s['company_id'],s['user_id'],'api_key_revoke','api_key',kid,ip_address=self._ip(),campaign_id=s['campaign_id']);return self.send_json({'ok':True})
            if path == '/api/admin/integrations/webhook/create':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'advanced'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                url=str(payload.get('url') or '').strip()[:500]
                try:u=urllib.parse.urlparse(url)
                except Exception:u=None
                if not u or u.scheme!='https' or not u.netloc:return self.send_json({'ok':False,'error':'https_webhook_required'},400)
                events=payload.get('events') if isinstance(payload.get('events'),list) else []
                allowed={'customer.created','purchase.created','points.earned','stamp.earned','reward.redeemed','*'};events=[x for x in events if x in allowed]
                if not events:return self.send_json({'ok':False,'error':'events_required'},400)
                secret='whsec_'+secrets.token_urlsafe(24)
                try: enc=encrypt_secret(secret)
                except RuntimeError:return self.send_json({'ok':False,'error':'security_key_not_configured'},503)
                wid=insert_id(conn,"INSERT INTO webhook_subscriptions(campaign_id,url,secret_enc,events_json,created_at) VALUES(?,?,?,?,?)",(s['campaign_id'],url,enc,json.dumps(events),now_ts()))
                audit(conn,s['company_id'],s['user_id'],'webhook_create','webhook',wid,details=url,ip_address=self._ip(),campaign_id=s['campaign_id'])
                return self.send_json({'ok':True,'id':wid,'secret':secret,'warning':'O segredo é exibido uma única vez.'})
            if path == '/api/admin/integrations/webhook/delete':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'advanced'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                wid=int(payload.get('id') or 0);conn.execute("DELETE FROM webhook_subscriptions WHERE id=? AND campaign_id=?",(wid,s['campaign_id']))
                audit(conn,s['company_id'],s['user_id'],'webhook_delete','webhook',wid,ip_address=self._ip(),campaign_id=s['campaign_id']);return self.send_json({'ok':True})
            if path == '/api/security/sessions/revoke-other':
                current_token=self._session_token()
                before=conn.execute('SELECT COUNT(*) n FROM sessions WHERE user_id=? AND token<>?',(s['user_id'],current_token or '')).fetchone()['n']
                conn.execute('DELETE FROM sessions WHERE user_id=? AND token<>?',(s['user_id'],current_token or ''))
                audit(conn,s['company_id'],s['user_id'],'sessions_revoked','user',s['user_id'],details=f'other_sessions={int(before or 0)}',ip_address=self._ip())
                return self.send_json({'ok':True,'revoked':int(before or 0)})
            if path == '/api/security/2fa/setup':
                if not (s['role']=='manager' or bool(s['is_client_admin'])):return self.send_json({'ok':False,'error':'forbidden'},403)
                password=str(payload.get('password') or '')
                u=conn.execute('SELECT id,email,password_hash,totp_enabled FROM users WHERE id=?',(s['user_id'],)).fetchone()
                if not u or not verify_password(password,u['password_hash']):return self.send_json({'ok':False,'error':'invalid_password'},403)
                if u['totp_enabled']:return self.send_json({'ok':False,'error':'two_factor_already_enabled'},409)
                secret=generate_totp_secret(); uri=_totp_uri(secret,u['email'])
                try: encrypted_secret=encrypt_secret(secret)
                except RuntimeError: return self.send_json({'ok':False,'error':'security_key_not_configured'},503)
                conn.execute('UPDATE users SET totp_secret_enc=?,totp_enabled=0,totp_confirmed_at=NULL WHERE id=?',(encrypted_secret,u['id']))
                audit(conn,s['company_id'],s['user_id'],'two_factor_setup_started','user',s['user_id'],ip_address=self._ip())
                return self.send_json({'ok':True,'secret':secret,'qr_data':_totp_qr_data(uri)})
            if path == '/api/security/2fa/confirm':
                if not (s['role']=='manager' or bool(s['is_client_admin'])):return self.send_json({'ok':False,'error':'forbidden'},403)
                code=str(payload.get('code') or '')
                u=conn.execute('SELECT id,totp_secret_enc,totp_enabled FROM users WHERE id=?',(s['user_id'],)).fetchone()
                secret=decrypt_secret(u['totp_secret_enc']) if u and u['totp_secret_enc'] else ''
                if not secret or not verify_totp(secret,code):return self.send_json({'ok':False,'error':'two_factor_invalid'},400)
                conn.execute('UPDATE users SET totp_enabled=1,totp_confirmed_at=? WHERE id=?',(now_ts(),s['user_id']))
                conn.execute('DELETE FROM auth_challenges WHERE user_id=?',(s['user_id'],))
                audit(conn,s['company_id'],s['user_id'],'two_factor_enabled','user',s['user_id'],ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/security/2fa/disable':
                if not (s['role']=='manager' or bool(s['is_client_admin'])):return self.send_json({'ok':False,'error':'forbidden'},403)
                password=str(payload.get('password') or ''); code=str(payload.get('code') or '')
                u=conn.execute('SELECT id,password_hash,totp_secret_enc,totp_enabled FROM users WHERE id=?',(s['user_id'],)).fetchone()
                secret=decrypt_secret(u['totp_secret_enc']) if u and u['totp_secret_enc'] else ''
                if not u or not verify_password(password,u['password_hash']):return self.send_json({'ok':False,'error':'invalid_password'},403)
                if not u['totp_enabled'] or not verify_totp(secret,code):return self.send_json({'ok':False,'error':'two_factor_invalid'},400)
                conn.execute('UPDATE users SET totp_secret_enc=NULL,totp_enabled=0,totp_confirmed_at=NULL WHERE id=?',(s['user_id'],))
                conn.execute('DELETE FROM auth_challenges WHERE user_id=?',(s['user_id'],))
                audit(conn,s['company_id'],s['user_id'],'two_factor_disabled','user',s['user_id'],ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/admin/loyalty360/settings':
                s=self._require_auth(conn,'attendant')
                if not s:return
                if not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'advanced'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                campaign=conn.execute("SELECT loyalty_type FROM campaigns WHERE id=? AND company_id=?",(s['campaign_id'],s['company_id'])).fetchone()
                if not campaign:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                # O tipo principal é definido exclusivamente no Painel Taboo.
                # O administrador da empresa configura apenas recursos avançados do mesmo programa.
                if campaign['loyalty_type']=='points':
                    expiry=int(payload.get('points_expiry_days') or 180)
                    if expiry not in tuple(i*30 for i in range(1,13)): return self.send_json({'ok':False,'error':'invalid_points_expiry'},400)
                    conn.execute("UPDATE campaigns SET points_expiry_days=? WHERE id=?",(expiry,s['campaign_id']))
                return self.send_json({'ok':True,'loyalty_type':campaign['loyalty_type']})
            if path == '/api/admin/loyalty360/tier':
                s=self._require_auth(conn,'attendant');
                if not s:return
                if not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'vip_tiers'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                campaign=conn.execute("SELECT loyalty_type FROM campaigns WHERE id=?",(s['campaign_id'],)).fetchone()
                if not campaign or campaign['loyalty_type']!='points':return self.send_json({'ok':False,'error':'points_program_required'},409)
                name=str(payload.get('name') or '').strip()[:60]; mp=max(0,int(payload.get('min_points') or 0)); benefit=str(payload.get('benefit') or '').strip()[:200]
                if not name:return self.send_json({'ok':False,'error':'invalid_tier'},400)
                insert_id(conn,"INSERT INTO loyalty_tiers(campaign_id,name,min_points,benefit,active,created_at) VALUES(?,?,?,?,1,?)",(s['campaign_id'],name,mp,benefit,now_ts())); return self.send_json({'ok':True})
            if path == '/api/admin/loyalty360/multiplier':
                s=self._require_auth(conn,'attendant');
                if not s:return
                if not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'multipliers'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                campaign=conn.execute("SELECT loyalty_type FROM campaigns WHERE id=?",(s['campaign_id'],)).fetchone()
                if not campaign or campaign['loyalty_type']!='points':return self.send_json({'ok':False,'error':'points_program_required'},409)
                name=str(payload.get('name') or '').strip()[:80]; factor=max(1,min(10,float(payload.get('factor') or 1))); weekday=str(payload.get('weekday') or 'all')[:20]
                insert_id(conn,"INSERT INTO point_multipliers(campaign_id,name,factor,weekday,start_hour,end_hour,active,created_at) VALUES(?,?,?,?,?,?,1,?)",(s['campaign_id'],name,factor,weekday,str(payload.get('start_hour') or ''),str(payload.get('end_hour') or ''),now_ts())); return self.send_json({'ok':True})
            if path == '/api/admin/gift-card':
                s=self._require_auth(conn,'attendant');
                if not s:return
                if not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'gift_cards'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                value=max(100,int(payload.get('value_cents') or 0)); code='VALE-'+secrets.token_hex(4).upper(); purchaser=str(payload.get('purchaser_name') or '').strip()[:100]; beneficiary=str(payload.get('beneficiary_name') or '').strip()[:100]; gid=insert_id(conn,"INSERT INTO gift_cards(campaign_id,code,value_cents,balance_cents,status,purchaser_name,beneficiary_name,created_at) VALUES(?,?,?,?,?,?,?,?)",(s['campaign_id'],code,value,value,'active',purchaser,beneficiary,now_ts())); conn.execute('INSERT INTO gift_card_events(gift_card_id,user_id,event_type,amount_cents,balance_after_cents,note,created_at) VALUES(?,?,?,?,?,?,?)',(gid,s['user_id'],'created',value,value,'Vale criado',now_ts())); return self.send_json({'ok':True,'code':code,'qr_url':'/api/qr?data='+urllib.parse.quote('GIFT:'+code)})

            if path == '/api/admin/gift-card/delete':
                s=self._require_auth(conn,'attendant');
                if not s:return
                if not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'gift_cards'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                code=str(payload.get('code') or '').strip().upper()
                if not code:return self.send_json({'ok':False,'error':'gift_code_required'},400)
                gift=conn.execute("SELECT id,code FROM gift_cards WHERE campaign_id=? AND upper(code)=upper(?)",(s['campaign_id'],code)).fetchone()
                if not gift:return self.send_json({'ok':False,'error':'gift_not_found'},404)
                conn.execute("DELETE FROM gift_cards WHERE id=? AND campaign_id=?",(gift['id'],s['campaign_id']))
                audit(conn,s['company_id'],s['user_id'],'gift_card_delete','gift_card',gift['id'],details=gift['code'],ip_address=self._ip())
                return self.send_json({'ok':True,'code':gift['code']})

            if path == '/api/attendant/gift-card/redeem':
                if s['role']=='attendant' and not self._need_permission(s,'use_gift'): return
                if s['role']!='attendant' or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'gift_cards'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                code=str(payload.get('code') or '').strip().upper()
                try: amount=int(payload.get('amount_cents') or 0)
                except: amount=0
                if not code or amount<1:return self.send_json({'ok':False,'error':'invalid_gift_redeem'},400)
                with connect(DB_PATH) as conn2:
                    gift=fetchone_for_update(conn2,"SELECT * FROM gift_cards WHERE campaign_id=? AND upper(code)=upper(?)",(s['campaign_id'],code))
                    if not gift:return self.send_json({'ok':False,'error':'gift_not_found'},404)
                    if gift['status']!='active':return self.send_json({'ok':False,'error':'gift_inactive'},409)
                    balance=int(gift['balance_cents'] or 0)
                    if amount>balance:return self.send_json({'ok':False,'error':'insufficient_gift_balance','balance_cents':balance},409)
                    new_balance=balance-amount; new_status='used' if new_balance==0 else 'active'
                    conn2.execute("UPDATE gift_cards SET balance_cents=?,status=? WHERE id=?",(new_balance,new_status,gift['id'])); conn2.execute('INSERT INTO gift_card_events(gift_card_id,user_id,event_type,amount_cents,balance_after_cents,note,created_at) VALUES(?,?,?,?,?,?,?)',(gift['id'],s['user_id'],'redeem',amount,new_balance,'Uso do vale',now_ts()))
                    audit(conn2,s['company_id'],s['user_id'],'gift_card_redeem','gift_card',gift['id'],details=f'{code};R${amount/100:.2f};saldo=R${new_balance/100:.2f}',ip_address=self._ip())
                return self.send_json({'ok':True,'code':code,'amount_cents':amount,'balance_cents':new_balance,'status':new_status})

            if path == '/api/admin/marketing-campaign/save':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                name=str(payload.get('name','')).strip()[:100]; segment=str(payload.get('segment','all')); channel=str(payload.get('channel','both')); message=str(payload.get('message','')).strip()[:4096]
                if len(name)<2 or not message or segment not in ('all','new','active','recurrent','vip','at_risk','inactive','inactive60','inactive90','almost_reward','reward_ready','birthdays') or channel not in ('email','whatsapp','both'):return self.send_json({'ok':False,'error':'invalid_campaign'},400)
                mid=insert_id(conn,'INSERT INTO marketing_campaigns(campaign_id,name,segment,channel,message,status,created_at) VALUES(?,?,?,?,?,?,?)',(s['campaign_id'],name,segment,channel,message,'draft',now_ts()))
                audit(conn,s['company_id'],s['user_id'],'marketing_campaign_create','marketing_campaign',mid,details=name,ip_address=self._ip())
                return self.send_json({'ok':True,'id':mid})
            if path == '/api/admin/marketing-campaign/send':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                mid=int(payload.get('id') or 0); mc=conn.execute('SELECT * FROM marketing_campaigns WHERE id=? AND campaign_id=?',(mid,s['campaign_id'])).fetchone()
                if not mc:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                if mc['status']=='sent':return self.send_json({'ok':False,'error':'already_sent'},409)
                rows=campaign_recipient_rows(conn,s['campaign_id'],mc['segment']); queued=0; sent_at=now_ts()
                for r in rows:
                    q=False
                    if mc['channel'] in ('email','both') and r['email'] and r['marketing_email'] and email_configured(email_config_for_client(conn,s['campaign_id'])):
                        enqueue_message(conn,s['campaign_id'],'campaign_email',r['email'],{'name':r['name'],'message':mc['message'],'subject':'Fidelizaê! • '+mc['name']}); q=True
                    if mc['channel'] in ('whatsapp','both') and r['phone'] and r['marketing_whatsapp'] and whatsapp_cloud_configured(whatsapp_config_for_client(conn,s['campaign_id'])):
                        enqueue_message(conn,s['campaign_id'],'whatsapp',r['phone'],{'message':mc['message']}); q=True
                    if q:
                        cur=conn.execute('INSERT INTO marketing_campaign_recipients(marketing_campaign_id,membership_id,sent_at) VALUES(?,?,?) ON CONFLICT(marketing_campaign_id,membership_id) DO NOTHING',(mid,r['membership_id'],sent_at))
                        # O contador representa destinatários efetivamente novos.
                        if getattr(cur,'rowcount',0)>0: queued+=1
                conn.execute("UPDATE marketing_campaigns SET status='sent',sent_at=? WHERE id=?",(sent_at,mid)); audit(conn,s['company_id'],s['user_id'],'marketing_campaign_send','marketing_campaign',mid,details=f'queued={queued}',ip_address=self._ip())
                return self.send_json({'ok':True,'queued':queued})
            if path == '/api/admin/coupon/save':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'coupons'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                name=str(payload.get('name','')).strip()[:100]; code=re.sub(r'[^A-Z0-9_-]','',str(payload.get('code','')).upper())[:30]; typ=str(payload.get('benefit_type','percent')); segment=str(payload.get('segment','all'))
                try: val=int(payload.get('benefit_value') or 0); limit=int(payload.get('usage_limit') or 0); starts=int(payload.get('starts_at') or 0) or None; ends=int(payload.get('ends_at') or 0) or None
                except: return self.send_json({'ok':False,'error':'invalid_coupon'},400)
                if len(name)<2 or len(code)<3 or typ not in ('percent','fixed','bonus_points','bonus_stamps') or val<=0:return self.send_json({'ok':False,'error':'invalid_coupon'},400)
                try: cid=insert_id(conn,'INSERT INTO coupons(campaign_id,name,code,benefit_type,benefit_value,segment,starts_at,ends_at,usage_limit,active,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(s['campaign_id'],name,code,typ,val,segment,starts,ends,limit,1,now_ts()))
                except integrity_errors():return self.send_json({'ok':False,'error':'coupon_code_exists'},409)
                audit(conn,s['company_id'],s['user_id'],'coupon_create','coupon',cid,details=code,ip_address=self._ip());return self.send_json({'ok':True,'id':cid})
            if path == '/api/admin/coupon/toggle':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'coupons'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                conn.execute('UPDATE coupons SET active=? WHERE id=? AND campaign_id=?',(1 if payload.get('active') else 0,int(payload.get('id') or 0),s['campaign_id']));return self.send_json({'ok':True})
            if path == '/api/attendant/coupon/apply':
                if s['role']!='attendant' or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'coupons'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                code=str(payload.get('code','')).strip().upper(); public_id=str(payload.get('public_id','')).strip(); now=now_ts(); purchase_cents=max(0,int(payload.get('purchase_cents') or 0))
                m=conn.execute('SELECT m.*,c.loyalty_type,c.goal FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE m.public_id=? AND m.campaign_id=?',(public_id,s['campaign_id'])).fetchone(); cup=conn.execute('SELECT * FROM coupons WHERE campaign_id=? AND upper(code)=upper(?) AND active=1',(s['campaign_id'],code)).fetchone()
                if not m or not cup:return self.send_json({'ok':False,'error':'coupon_not_found'},404)
                if (cup['starts_at'] and now<cup['starts_at']) or (cup['ends_at'] and now>cup['ends_at']):return self.send_json({'ok':False,'error':'coupon_not_available'},409)
                if conn.execute('SELECT 1 FROM coupon_redemptions WHERE coupon_id=? AND membership_id=?',(cup['id'],m['id'])).fetchone():return self.send_json({'ok':False,'error':'coupon_already_used'},409)
                if cup['usage_limit'] and conn.execute('SELECT COUNT(*) n FROM coupon_redemptions WHERE coupon_id=?',(cup['id'],)).fetchone()['n']>=cup['usage_limit']:return self.send_json({'ok':False,'error':'coupon_limit_reached'},409)
                benefit={'type':cup['benefit_type'],'value':cup['benefit_value']}
                if cup['benefit_type']=='bonus_points':
                    prev=int(m['points_balance'] or 0); new=prev+int(cup['benefit_value']); conn.execute('UPDATE memberships SET points_balance=? WHERE id=?',(new,m['id'])); tx_coupon=insert_id(conn,'INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(m['id'],s['user_id'],current_branch_id(conn,s['user_id']),'adjustment',cup['benefit_value'],prev,new,0,self._ip(),'Cupom '+cup['code'],now)); exp=conn.execute('SELECT points_expiry_days FROM campaigns WHERE id=?',(s['campaign_id'],)).fetchone(); add_point_lot(conn,m['id'],tx_coupon,int(cup['benefit_value']),int(exp['points_expiry_days'] or 180),now)
                elif cup['benefit_type']=='bonus_stamps':
                    prev=int(m['progress'] or 0); total=prev+int(cup['benefit_value']); rewards=total//int(m['goal']); new=total%int(m['goal']); conn.execute('UPDATE memberships SET progress=?,rewards_available=rewards_available+? WHERE id=?',(new,rewards,m['id'])); conn.execute('INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(m['id'],s['user_id'],current_branch_id(conn,s['user_id']),'stamp',cup['benefit_value'],prev,new,rewards,self._ip(),'Cupom '+cup['code'],now))
                discount_cents=0
                if cup['benefit_type']=='percent' and purchase_cents>0: discount_cents=min(purchase_cents,round(purchase_cents*int(cup['benefit_value'])/100))
                elif cup['benefit_type']=='fixed' and purchase_cents>0: discount_cents=min(purchase_cents,int(cup['benefit_value']))
                benefit['discount_cents']=discount_cents; benefit['purchase_cents']=purchase_cents
                conn.execute('INSERT INTO coupon_redemptions(coupon_id,membership_id,user_id,purchase_cents,discount_cents,created_at) VALUES(?,?,?,?,?,?)',(cup['id'],m['id'],s['user_id'],purchase_cents,discount_cents,now)); audit(conn,s['company_id'],s['user_id'],'coupon_redeem','coupon',cup['id'],details=f"{cup['code']};membership={m['public_id']}",ip_address=self._ip()); notify_wallet_updates(conn,m['public_id'])
                return self.send_json({'ok':True,'coupon':cup['name'],'benefit':benefit})

            if path == '/api/manager/notification/state':
                if s['role']!='manager':return self.send_json({'ok':False,'error':'forbidden'},403)
                fp=str(payload.get('fingerprint') or '')[:40]; status=str(payload.get('status') or 'seen'); priority=str(payload.get('priority') or 'medium')
                if status not in ('new','seen','resolved') or priority not in ('low','medium','high'):return self.send_json({'ok':False,'error':'invalid_state'},400)
                try: conn.execute('INSERT INTO alert_states(fingerprint,status,priority,updated_at) VALUES(?,?,?,?)',(fp,status,priority,now_ts()))
                except integrity_errors(): conn.execute('UPDATE alert_states SET status=?,priority=?,updated_at=? WHERE fingerprint=?',(status,priority,now_ts(),fp))
                return self.send_json({'ok':True})

            if path == '/api/attendant/customer/note':
                if s['role']!='attendant' or not s['campaign_id']: return self.send_json({'ok':False,'error':'forbidden'},403)
                note=str(payload.get('note') or '').strip()[:800]
                try: customer_id=int(payload.get('customer_id') or 0)
                except: customer_id=0
                if not note:return self.send_json({'ok':False,'error':'note_required'},400)
                m=conn.execute('SELECT m.id FROM memberships m WHERE m.customer_id=? AND m.campaign_id=?',(customer_id,s['campaign_id'])).fetchone()
                if not m:return self.send_json({'ok':False,'error':'customer_not_found'},404)
                insert_id(conn,'INSERT INTO customer_notes(membership_id,user_id,note,created_at) VALUES(?,?,?,?)',(m['id'],s['user_id'],note,now_ts())); audit(conn,s['company_id'],s['user_id'],'customer_note_create','membership',m['id'],details=note[:120],ip_address=self._ip()); return self.send_json({'ok':True})

            if path == '/api/attendant/password':
                if s['role']!='attendant': return self.send_json({'ok':False,'error':'forbidden'},403)
                current_password=str(payload.get('current_password',''))
                new_password=str(payload.get('new_password','')).strip()
                if not password_is_strong(new_password,12): return self.send_json({'ok':False,'error':'invalid_new_password'},400)
                u=conn.execute('SELECT id,password_hash FROM users WHERE id=? AND role=\'attendant\' AND active=1',(s['user_id'],)).fetchone()
                if not u or not verify_password(current_password,u['password_hash']): return self.send_json({'ok':False,'error':'invalid_current_password'},401)
                if verify_password(new_password,u['password_hash']): return self.send_json({'ok':False,'error':'same_password'},409)
                conn.execute('UPDATE users SET password_hash=? WHERE id=?',(hash_password(new_password),s['user_id']))
                conn.execute('DELETE FROM sessions WHERE user_id=?',(s['user_id'],))
                new_token,new_csrf=create_session(conn,s['user_id'])
                audit(conn,s['company_id'],s['user_id'],'password_change','user',s['user_id'],details='sessions_revoked',ip_address=self._ip())
                print(f'[AUTH] ATTENDANT_PASSWORD_CHANGED user_id={s["user_id"]} sessions_revoked=True')
                return self.send_json({'ok':True,'csrf':new_csrf},200,{'Set-Cookie':_session_cookie(new_token)})
            if path == '/api/attendant/customer/update':
                if s['role']!='attendant' or not s['is_client_admin']: return self.send_json({'ok':False,'error':'forbidden'},403)
                if not s['campaign_id']: return self.send_json({'ok':False,'error':'attendant_without_client'},403)
                try: customer_id=int(payload.get('customer_id',0))
                except (TypeError,ValueError): customer_id=0
                name=str(payload.get('name','')).strip()[:80]
                email=normalize_email(payload.get('email'))
                phone=normalize_phone(payload.get('phone'))
                birth_date=normalize_birth_date(payload.get('birth_date'))
                gender=str(payload.get('gender') or '').strip().lower(); gender=gender if gender in ('female','male','other','prefer_not') else ''
                cpf=normalize_cpf(payload.get('cpf'))
                if customer_id<1 or len(name)<2 or not email or not phone or not birth_date or not cpf:
                    return self.send_json({'ok':False,'error':'invalid_customer_data'},400)
                member=conn.execute("""SELECT m.id,m.public_id FROM memberships m JOIN campaigns c ON c.id=m.campaign_id
                    WHERE m.customer_id=? AND m.campaign_id=? AND c.company_id=?""",(customer_id,s['campaign_id'],s['company_id'])).fetchone()
                if not member: return self.send_json({'ok':False,'error':'customer_not_found'},404)
                duplicate=conn.execute("""SELECT cu.id FROM customers cu JOIN memberships m ON m.customer_id=cu.id
                    WHERE m.campaign_id=? AND cu.cpf_hash=? AND cu.id<>? LIMIT 1""",(s['campaign_id'],pii_lookup_hash(cpf,'cpf'),customer_id)).fetchone()
                if duplicate: return self.send_json({'ok':False,'error':'cpf_exists'},409)
                pii=protected_customer_pii(phone,cpf)
                conn.execute('UPDATE customers SET name=?,contact=?,email=?,phone=NULL,phone_enc=?,phone_hash=?,birth_date=?,gender=?,cpf=NULL,cpf_enc=?,cpf_hash=? WHERE id=?',
                    (name,email,email,pii['phone_enc'],pii['phone_hash'],birth_date,gender or None,pii['cpf_enc'],pii['cpf_hash'],customer_id))
                audit(conn,s['company_id'],s['user_id'],'customer_update','customer',customer_id,details=member['public_id'],ip_address=self._ip())
                return self.send_json({'ok':True,'customer_id':customer_id})
            if path == '/api/attendant/customer/delete':
                if s['role']!='attendant' or not s['is_client_admin']: return self.send_json({'ok':False,'error':'forbidden'},403)
                if not s['campaign_id']: return self.send_json({'ok':False,'error':'attendant_without_client'},403)
                try: customer_id=int(payload.get('customer_id',0))
                except (TypeError,ValueError): customer_id=0
                member=conn.execute("""SELECT m.id,m.public_id FROM memberships m JOIN campaigns c ON c.id=m.campaign_id
                    WHERE m.customer_id=? AND m.campaign_id=? AND c.company_id=?""",(customer_id,s['campaign_id'],s['company_id'])).fetchone()
                if not member: return self.send_json({'ok':False,'error':'customer_not_found'},404)
                audit(conn,s['company_id'],s['user_id'],'customer_delete','membership',member['public_id'],details=f'customer_id={customer_id}',ip_address=self._ip())
                conn.execute('DELETE FROM memberships WHERE id=?',(member['id'],))
                remaining=conn.execute('SELECT COUNT(*) n FROM memberships WHERE customer_id=?',(customer_id,)).fetchone()['n']
                if remaining==0: conn.execute('DELETE FROM customers WHERE id=?',(customer_id,))
                return self.send_json({'ok':True,'deleted_customer_id':customer_id,'deleted_membership':member['public_id']})
            if path == '/api/attendant/whatsapp':
                if s['role']=='attendant' and not self._need_permission(s,'send_messages'): return
                if s['role']!='attendant': return self.send_json({'ok':False,'error':'forbidden'},403)
                if not s['campaign_id']: return self.send_json({'ok':False,'error':'attendant_without_client'},403)
                if not plan_allows(conn,s['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                recipient=str(payload.get('recipient','')).strip()
                message=str(payload.get('message','')).strip()
                if not message or len(message)>4096: return self.send_json({'ok':False,'error':'invalid_message'},400)
                if recipient == 'all' or recipient.startswith('segment:'):
                    extra=''; args=[s['campaign_id'],'']
                    if recipient=='segment:birthdays': extra=" AND substr(cu.birth_date,6,2)=?"; args.append(datetime.now(ZoneInfo('America/Sao_Paulo')).strftime('%m'))
                    elif recipient=='segment:inactive30': extra=" AND COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at)<?"; args.append(now_ts()-30*86400)
                    elif recipient=='segment:reward_ready': extra=" AND (m.rewards_available>0 OR m.points_balance>=(SELECT COALESCE(MIN(points_cost),999999999) FROM reward_catalog WHERE campaign_id=m.campaign_id AND active=1))"
                    rows=conn.execute('SELECT DISTINCT cu.id,cu.name,cu.phone,cu.phone_enc FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND cu.marketing_whatsapp=1 AND COALESCE(cu.phone_enc,cu.phone) IS NOT NULL AND COALESCE(cu.phone_enc,cu.phone)<>?'+extra+' ORDER BY cu.name',tuple(args)).fetchall()
                    rows=[customer_rowdict(r) for r in rows]
                else:
                    try: customer_id=int(recipient)
                    except (TypeError,ValueError): return self.send_json({'ok':False,'error':'invalid_recipient'},400)
                    rows=conn.execute('''SELECT DISTINCT cu.id,cu.name,cu.phone,cu.phone_enc FROM customers cu JOIN memberships m ON m.customer_id=cu.id
                        WHERE m.campaign_id=? AND cu.id=? AND cu.marketing_whatsapp=1 AND COALESCE(cu.phone_enc,cu.phone) IS NOT NULL AND COALESCE(cu.phone_enc,cu.phone)<>?''',(s['campaign_id'],customer_id,'')).fetchall()
                rows=[customer_rowdict(r) for r in rows]
                if not rows: return self.send_json({'ok':False,'error':'no_recipients'},404)
                wa_cfg=whatsapp_config_for_client(conn,s['campaign_id'])
                cloud=whatsapp_cloud_configured(wa_cfg)
                if not cloud:return self.send_json({'ok':False,'error':'whatsapp_not_configured'},503)
                results=[]
                for r in rows:
                    qid=enqueue_message(conn,s['campaign_id'],'whatsapp',r['phone'],{'message':message})
                    results.append({'customer_id':r['id'],'name':r['name'],'phone':r['phone'],'queued':True,'queue_id':qid})
                    audit(conn,s['company_id'],s['user_id'],'whatsapp_queued','customer',r['id'],details=f'queue={qid}',ip_address=self._ip())
                return self.send_json({'ok':True,'queued_count':len(results),'results':results})
            if path == '/api/attendant/email':
                if s['role']=='attendant' and not self._need_permission(s,'send_messages'): return
                if s['role']!='attendant': return self.send_json({'ok':False,'error':'forbidden'},403)
                if not s['campaign_id']: return self.send_json({'ok':False,'error':'attendant_without_client'},403)
                if not plan_allows(conn,s['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                smtp_cfg=email_config_for_client(conn,s['campaign_id'])
                if not email_configured(smtp_cfg): return self.send_json({'ok':False,'error':'email_provider_not_configured'},503)
                recipient=str(payload.get('recipient','')).strip()
                message=str(payload.get('message','')).strip()
                image_data=payload.get('image_data')
                if not message and not image_data: return self.send_json({'ok':False,'error':'message_or_image_required'},400)
                if len(message)>10000: return self.send_json({'ok':False,'error':'invalid_message'},400)
                try:
                    if image_data: decode_image_data(image_data)
                except ValueError as exc:
                    return self.send_json({'ok':False,'error':str(exc)},400)
                if recipient == 'all' or recipient.startswith('segment:'):
                    extra=''; args=[s['campaign_id'],'']
                    if recipient=='segment:birthdays': extra=" AND substr(cu.birth_date,6,2)=?"; args.append(datetime.now(ZoneInfo('America/Sao_Paulo')).strftime('%m'))
                    elif recipient=='segment:inactive30': extra=" AND COALESCE((SELECT MAX(t.created_at) FROM transactions t WHERE t.membership_id=m.id),m.created_at)<?"; args.append(now_ts()-30*86400)
                    elif recipient=='segment:reward_ready': extra=" AND (m.rewards_available>0 OR m.points_balance>=(SELECT COALESCE(MIN(points_cost),999999999) FROM reward_catalog WHERE campaign_id=m.campaign_id AND active=1))"
                    rows=conn.execute('SELECT DISTINCT cu.id,cu.name,cu.email FROM customers cu JOIN memberships m ON m.customer_id=cu.id WHERE m.campaign_id=? AND cu.marketing_email=1 AND cu.email IS NOT NULL AND cu.email<>?'+extra+' ORDER BY cu.name',tuple(args)).fetchall()
                    rows=[customer_rowdict(r) for r in rows]
                else:
                    try: customer_id=int(recipient)
                    except (TypeError,ValueError): return self.send_json({'ok':False,'error':'invalid_recipient'},400)
                    rows=conn.execute('''SELECT DISTINCT cu.id,cu.name,cu.email FROM customers cu JOIN memberships m ON m.customer_id=cu.id
                        WHERE m.campaign_id=? AND cu.id=? AND cu.marketing_email=1 AND cu.email IS NOT NULL AND cu.email<>?''',(s['campaign_id'],customer_id,'')).fetchall()
                rows=[customer_rowdict(r) for r in rows]
                if not rows: return self.send_json({'ok':False,'error':'no_recipients'},404)
                results=[]
                for r in rows:
                    qid=enqueue_message(conn,s['campaign_id'],'campaign_email',r['email'],{'name':r['name'],'message':message,'image_data':image_data,'subject':f'Fidelizaê! • {s["client_name"] or "Mensagem"}'})
                    results.append({'customer_id':r['id'],'name':r['name'],'email':r['email'],'queued':True,'queue_id':qid})
                    audit(conn,s['company_id'],s['user_id'],'email_queued','customer',r['id'],details=f'queue={qid}',ip_address=self._ip())
                return self.send_json({'ok':True,'results':results,'queued_count':len(results)})
            if path == '/api/admin/template/save':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                name=str(payload.get('name','')).strip()[:80];channel=str(payload.get('channel','both'));subject=str(payload.get('subject','')).strip()[:150];body=str(payload.get('body','')).strip()[:4000]
                if not name or not body or channel not in ('email','whatsapp','both'):return self.send_json({'ok':False,'error':'invalid_template'},400)
                tid=insert_id(conn,'INSERT INTO message_templates(campaign_id,name,channel,subject,body,created_at) VALUES(?,?,?,?,?,?)',(s['campaign_id'],name,channel,subject,body,now_ts()))
                return self.send_json({'ok':True,'template_id':tid})
            if path == '/api/admin/template/test':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                try: template_id=int(payload.get('template_id') or 0)
                except (TypeError,ValueError): template_id=0
                try: customer_id=int(payload.get('customer_id') or 0)
                except (TypeError,ValueError): customer_id=0
                channel=str(payload.get('channel') or '').strip().lower()
                if not template_id or not customer_id:return self.send_json({'ok':False,'error':'invalid_test_recipient'},400)
                if channel not in ('email','whatsapp','both'):return self.send_json({'ok':False,'error':'invalid_test_channel'},400)
                customer=conn.execute('''SELECT DISTINCT cu.id,cu.name,cu.email,cu.phone,cu.phone_enc,cu.marketing_email,cu.marketing_whatsapp FROM customers cu JOIN memberships m ON m.customer_id=cu.id
                    WHERE cu.id=? AND m.campaign_id=?''',(customer_id,s['campaign_id'])).fetchone()
                if not customer:return self.send_json({'ok':False,'error':'customer_not_found'},404)
                customer=customer_rowdict(customer)
                tpl=conn.execute('SELECT id,name,channel,subject,body FROM message_templates WHERE id=? AND campaign_id=?',(template_id,s['campaign_id'])).fetchone()
                if not tpl:return self.send_json({'ok':False,'error':'template_not_found'},404)
                allowed={'email':('email',),'whatsapp':('whatsapp',),'both':('email','whatsapp','both')}
                if channel not in allowed.get(tpl['channel'],()):return self.send_json({'ok':False,'error':'template_channel_mismatch'},409)
                if channel in ('email','both') and not customer['marketing_email']:return self.send_json({'ok':False,'error':'email_consent_required' if channel=='email' else 'both_consent_required'},403)
                if channel in ('whatsapp','both') and not customer['marketing_whatsapp']:return self.send_json({'ok':False,'error':'whatsapp_consent_required' if channel=='whatsapp' else 'both_consent_required'},403)
                email=str(customer['email'] or '').strip()
                phone=re.sub(r'\D','',str(customer['phone'] or ''))[:20]
                if channel in ('email','both') and ('@' not in email or len(email)>254):return self.send_json({'ok':False,'error':'invalid_test_recipient'},400)
                if channel in ('whatsapp','both') and (len(phone)<8 or len(phone)>15):return self.send_json({'ok':False,'error':'invalid_test_recipient'},400)
                camp=conn.execute('SELECT name,reward_name,goal FROM campaigns WHERE id=? AND company_id=?',(s['campaign_id'],s['company_id'])).fetchone()
                if not camp:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                message=render_test_template(tpl['body'],rowdict(camp),customer['name'])
                sent=[]
                if channel in ('email','both'):
                    email_cfg=email_config_for_client(conn,s['campaign_id'])
                    if not email_configured(email_cfg):return self.send_json({'ok':False,'error':'email_not_configured'},503)
                    subject=(str(tpl['subject'] or '').strip() or f'Teste • {camp["name"]}')[:150]
                    result=send_campaign_email(email,customer['name'],message,None,subject,email_cfg)
                    if not result.get('sent'):
                        reason=str(result.get('reason') or 'email_test_failed')[:300]
                        print(f'[EMAIL_TEST] SEND_FAILED campaign_id={s["campaign_id"]} template_id={template_id} reason={reason}')
                        return self.send_json({'ok':False,'error':'email_test_failed','detail':reason},502)
                    audit(conn,s['company_id'],s['user_id'],'email_test_send','message_template',template_id,details=f'email=***{email[-8:]}',ip_address=self._ip())
                    sent.append('E-mail')
                if channel in ('whatsapp','both'):
                    cfg,wa_mode=whatsapp_test_delivery_config(conn,s['campaign_id'],phone)
                    if not cfg:
                        if whatsapp_cloud_configured(whatsapp_meta_test_config()):
                            return self.send_json({'ok':False,'error':'meta_test_recipient_not_allowed'},403)
                        return self.send_json({'ok':False,'error':'whatsapp_test_mode_not_configured'},503)
                    try:
                        response=send_whatsapp_cloud(phone,message,cfg)
                    except Exception as exc:
                        reason=str(exc)[:700]
                        print(f'[WHATSAPP_TEST] SEND_FAILED campaign_id={s["campaign_id"]} template_id={template_id} mode={wa_mode} reason={reason}')
                        return self.send_json({'ok':False,'error':'whatsapp_test_failed','detail':reason,'mode':wa_mode},502)
                    message_id=((response.get('messages') or [{}])[0]).get('id') if isinstance(response,dict) else None
                    audit(conn,s['company_id'],s['user_id'],'whatsapp_test_send','message_template',template_id,details=f'mode={wa_mode};phone=***{phone[-4:]};message_id={message_id or ""}',ip_address=self._ip())
                    sent.append('WhatsApp' + (' (Modo de teste Meta)' if wa_mode=='meta_test' else ''))
                return self.send_json({'ok':True,'message':'Teste enviado com sucesso por '+(' e '.join(sent))+'.','preview':message,'channels':sent})
            if path == '/api/admin/template/delete':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                conn.execute('DELETE FROM message_templates WHERE id=? AND campaign_id=?',(int(payload.get('template_id') or 0),s['campaign_id']));return self.send_json({'ok':True})

            if path == '/api/admin/reward/save':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                name=str(payload.get('name','')).strip()[:100]; description=str(payload.get('description','')).strip()[:300]
                try: points_cost=int(payload.get('points_cost') or 0)
                except: points_cost=0
                image_data=str(payload.get('image_data','') or '')
                try: stock=int(payload.get('stock',-1)); starts_at=int(payload.get('starts_at') or 0) or None; ends_at=int(payload.get('ends_at') or 0) or None
                except: stock=-1; starts_at=None; ends_at=None
                if stock < -1: stock=-1
                if not name or points_cost<1:return self.send_json({'ok':False,'error':'invalid_reward'},400)
                if image_data and (len(image_data)>900000 or not image_data.startswith(('data:image/png;base64,','data:image/jpeg;base64,','data:image/webp;base64,'))):return self.send_json({'ok':False,'error':'invalid_reward_image'},400)
                with connect(DB_PATH) as conn:
                    c=conn.execute("SELECT loyalty_type FROM campaigns WHERE id=? AND company_id=?",(s['campaign_id'],s['company_id'])).fetchone()
                    if not c or c['loyalty_type']!='points':return self.send_json({'ok':False,'error':'points_program_required'},409)
                    rid=insert_id(conn,"INSERT INTO reward_catalog(campaign_id,name,description,points_cost,image_data,active,stock,starts_at,ends_at,created_at,updated_at) VALUES(?,?,?,?,?,1,?,?,?,?,?)",(s['campaign_id'],name,description,points_cost,image_data or None,stock,starts_at,ends_at,now_ts(),now_ts()))
                    audit(conn,s['company_id'],s['user_id'],'reward_catalog_create','reward',rid,details=f'{name};{points_cost} pontos',ip_address=self._ip())
                return self.send_json({'ok':True,'reward_id':rid})
            if path == '/api/admin/reward/update':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                try: rid=int(payload.get('reward_id') or 0); points_cost=int(payload.get('points_cost') or 0)
                except: rid=0; points_cost=0
                name=str(payload.get('name','')).strip()[:100]; description=str(payload.get('description','')).strip()[:300]; image_data=payload.get('image_data',None)
                try: stock=int(payload.get('stock',-1)); starts_at=int(payload.get('starts_at') or 0) or None; ends_at=int(payload.get('ends_at') or 0) or None
                except: stock=-1; starts_at=None; ends_at=None
                if stock < -1: stock=-1
                if rid<1 or not name or points_cost<1:return self.send_json({'ok':False,'error':'invalid_reward'},400)
                with connect(DB_PATH) as conn:
                    r=conn.execute("SELECT id FROM reward_catalog WHERE id=? AND campaign_id=?",(rid,s['campaign_id'])).fetchone()
                    if not r:return self.send_json({'ok':False,'error':'reward_not_found'},404)
                    if image_data is None: conn.execute("UPDATE reward_catalog SET name=?,description=?,points_cost=?,stock=?,starts_at=?,ends_at=?,updated_at=? WHERE id=?",(name,description,points_cost,stock,starts_at,ends_at,now_ts(),rid))
                    else:
                        image_data=str(image_data or '')
                        if image_data and (len(image_data)>900000 or not image_data.startswith(('data:image/png;base64,','data:image/jpeg;base64,','data:image/webp;base64,'))):return self.send_json({'ok':False,'error':'invalid_reward_image'},400)
                        conn.execute("UPDATE reward_catalog SET name=?,description=?,points_cost=?,image_data=?,stock=?,starts_at=?,ends_at=?,updated_at=? WHERE id=?",(name,description,points_cost,image_data or None,stock,starts_at,ends_at,now_ts(),rid))
                    audit(conn,s['company_id'],s['user_id'],'reward_catalog_update','reward',rid,details=f'{name};{points_cost} pontos',ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/admin/reward/toggle':
                if s['role']!='attendant' or not s['is_client_admin'] or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                try: rid=int(payload.get('reward_id') or 0)
                except: rid=0
                active=1 if payload.get('active') else 0
                with connect(DB_PATH) as conn:
                    r=conn.execute("SELECT id FROM reward_catalog WHERE id=? AND campaign_id=?",(rid,s['campaign_id'])).fetchone()
                    if not r:return self.send_json({'ok':False,'error':'reward_not_found'},404)
                    conn.execute("UPDATE reward_catalog SET active=?,updated_at=? WHERE id=?",(active,now_ts(),rid)); audit(conn,s['company_id'],s['user_id'],'reward_catalog_toggle','reward',rid,details=f'active={active}',ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/attendant/points/earn':
                if s['role']=='attendant' and not self._need_permission(s,'add_balance'): return
                if s['role']!='attendant' or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                token,token_error=resolve_member_token(str(payload.get('token','')))
                if token_error:return self.send_json({'ok':False,'error':token_error},410 if token_error=='qr_expired' else 400)
                try: purchase_cents=int(payload.get('purchase_cents') or 0)
                except: purchase_cents=0
                idem=str(payload.get('idempotency_key','')).strip()
                if purchase_cents<1 or not idem:return self.send_json({'ok':False,'error':'invalid_purchase'},400)
                with connect(DB_PATH) as conn:
                    m=fetchone_for_update(conn,"SELECT m.*,c.company_id,c.name campaign_name,c.loyalty_type,c.points_spend_cents,cu.name customer_name FROM memberships m JOIN campaigns c ON c.id=m.campaign_id JOIN customers cu ON cu.id=m.customer_id WHERE (m.public_id=? OR m.qr_token=?) AND c.company_id=? AND c.id=?",(token,token,s['company_id'],s['campaign_id']))
                    if not m:return self.send_json({'ok':False,'error':'membership_not_found'},404)
                    if m['loyalty_type']!='points':return self.send_json({'ok':False,'error':'points_program_required'},409)
                    rate=max(1,int(m['points_spend_cents'] or 200)); base_earned=purchase_cents//rate; factor=active_multiplier(conn,s['campaign_id']) if plan_allows(conn,s['campaign_id'],'multipliers') else 1.0; earned=int(base_earned*factor)
                    if earned<1:return self.send_json({'ok':False,'error':'purchase_below_point_rule','message':'O valor da compra não gera nenhum ponto nesta regra.'},409)
                    prev=int(m['points_balance'] or 0); new=prev+earned; ts=now_ts()
                    tx_id=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,device_id,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",(m['id'],s['user_id'],current_branch_id(conn,s['user_id'],s['campaign_id']),'adjustment',earned,prev,new,0,idem,str(payload.get('device_id',''))[:120],self._ip(),f'Pontos por compra de R$ {purchase_cents/100:.2f} • {factor:g}x',ts))
                    conn.execute("UPDATE memberships SET points_balance=? WHERE id=?",(new,m['id'])); camp_exp=conn.execute('SELECT points_expiry_days FROM campaigns WHERE id=?',(s['campaign_id'],)).fetchone(); add_point_lot(conn,m['id'],tx_id,earned,int(camp_exp['points_expiry_days'] or 180),ts); record_purchase(conn,m['id'],tx_id,purchase_cents,'in_store',ts); audit(conn,s['company_id'],s['user_id'],'points_earn','membership',m['public_id'],details=f'R${purchase_cents/100:.2f};+{earned} pontos;{factor:g}x',ip_address=self._ip()); notify_wallet_updates(conn,m['public_id'])
                return self.send_json({'ok':True,'transaction_id':tx_id,'customer_name':m['customer_name'],'points_earned':earned,'points_balance':new})
            if path == '/api/attendant/points/redeem':
                if s['role']=='attendant' and not self._need_permission(s,'redeem_reward'): return
                if s['role']!='attendant' or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                token,token_error=resolve_member_token(str(payload.get('token','')))
                if token_error:return self.send_json({'ok':False,'error':token_error},410 if token_error=='qr_expired' else 400)
                try: reward_id=int(payload.get('reward_id') or 0)
                except: reward_id=0
                idem=str(payload.get('idempotency_key','')).strip()
                if reward_id<1 or not idem:return self.send_json({'ok':False,'error':'invalid_reward'},400)
                with connect(DB_PATH) as conn:
                    m=fetchone_for_update(conn,"SELECT m.*,c.company_id,c.loyalty_type,cu.name customer_name FROM memberships m JOIN campaigns c ON c.id=m.campaign_id JOIN customers cu ON cu.id=m.customer_id WHERE (m.public_id=? OR m.qr_token=?) AND c.company_id=? AND c.id=?",(token,token,s['company_id'],s['campaign_id']))
                    if not m:return self.send_json({'ok':False,'error':'membership_not_found'},404)
                    if m['loyalty_type']!='points':return self.send_json({'ok':False,'error':'points_program_required'},409)
                    r=conn.execute("SELECT * FROM reward_catalog WHERE id=? AND campaign_id=? AND active=1",(reward_id,s['campaign_id'])).fetchone()
                    if not r:return self.send_json({'ok':False,'error':'reward_not_found'},404)
                    prev=int(m['points_balance'] or 0); cost=int(r['points_cost'])
                    if prev<cost:return self.send_json({'ok':False,'error':'insufficient_points','message':'Saldo de pontos insuficiente para esta recompensa.'},409)
                    new=prev-cost
                    tx_id=insert_id(conn,"INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(m['id'],s['user_id'],current_branch_id(conn,s['user_id'],s['campaign_id']),'redeem',-cost,prev,new,0,idem,self._ip(),f'Resgate: {r["name"]} ({cost} pontos)',now_ts()))
                    consume_point_lots(conn,m['id'],cost); conn.execute("UPDATE memberships SET points_balance=? WHERE id=?",(new,m['id'])); conn.execute("INSERT INTO reward_redemptions(membership_id,reward_id,user_id,points_cost,created_at) VALUES(?,?,?,?,?)",(m['id'],r['id'],s['user_id'],cost,now_ts())); conn.execute("UPDATE reward_catalog SET stock=CASE WHEN stock>0 THEN stock-1 ELSE stock END,updated_at=? WHERE id=?",(now_ts(),r['id'])); audit(conn,s['company_id'],s['user_id'],'points_redeem','membership',m['public_id'],details=f'{r["name"]};-{cost} pontos',ip_address=self._ip()); notify_wallet_updates(conn,m['public_id'])
                return self.send_json({'ok':True,'transaction_id':tx_id,'customer_name':m['customer_name'],'reward_name':r['name'],'points_spent':cost,'points_balance':new})

            if path in ('/api/attendant/stamp','/api/attendant/stamp/remove','/api/attendant/redeem'):
                if s['role']!='attendant': return self.send_json({'ok':False,'error':'forbidden'},403)
                if not s['campaign_id']: return self.send_json({'ok':False,'error':'attendant_without_client'},403)
            if path == '/api/attendant/stamp':
                if not self._need_permission(s,'add_balance'): return
                token,token_error=resolve_member_token(payload.get('token'));
                if token_error:return self.send_json({'ok':False,'error':token_error},410 if token_error=='qr_expired' else 400)
                qty=int(payload.get('quantity',1)); idem=str(payload.get('idempotency_key','')).strip()[:100] or random_token(12); device=str(payload.get('device_id',''))[:100]
                try:
                    begin_write(conn)
                    dupe=conn.execute('SELECT id FROM transactions WHERE idempotency_key=?',(idem,)).fetchone()
                    if dupe: return self.send_json({'ok':True,'duplicate':True,'transaction_id':dupe['id']})
                    m=fetchone_for_update(conn,'''SELECT m.*,c.goal,c.min_stamp_interval_sec,c.max_stamps_per_hour,c.max_stamps_per_attendant_day,c.company_id,c.name campaign_name,c.loyalty_type,cu.name customer_name
                      FROM memberships m JOIN campaigns c ON c.id=m.campaign_id JOIN customers cu ON cu.id=m.customer_id WHERE (m.public_id=? OR m.qr_token=?) AND c.company_id=? AND c.id=?''',(token,token,s['company_id'],s['campaign_id']))
                    if not m: return self.send_json({'ok':False,'error':'membership_not_found'},404)
                    if m['loyalty_type']!='stamps': return self.send_json({'ok':False,'error':'stamps_program_required'},409)
                    validate_stamp(conn,m,m,s,qty)
                    prev=m['progress']; rewards=0; new=prev
                    for _ in range(qty):
                        new += 1
                        if new >= m['goal']:
                            rewards += 1; new = 0
                    tx_id=insert_id(conn,'''INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,device_id,ip_address,created_at)
                      VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',(m['id'],s['user_id'],current_branch_id(conn,s['user_id'],s['campaign_id']),'stamp',qty,prev,new,rewards,idem,device,self._ip(),now_ts()))
                    conn.execute('UPDATE memberships SET progress=?, rewards_available=rewards_available+? WHERE id=?',(new,rewards,m['id']))
                    purchase_cents=max(0,int(payload.get('purchase_cents') or 0))
                    if purchase_cents: record_purchase(conn,m['id'],tx_id,purchase_cents,'in_store',now_ts())
                    audit(conn,s['company_id'],s['user_id'],'stamp','membership',m['public_id'],details=f'qty={qty};reward+={rewards};valor={purchase_cents}',ip_address=self._ip()); notify_wallet_updates(conn,m['public_id'])
                    return self.send_json({'ok':True,'transaction_id':tx_id,'customer_name':m['customer_name'],'previous_progress':prev,'progress':new,'reward_added':rewards})
                except FraudError as e:
                    audit(conn,s['company_id'],s['user_id'],'stamp_blocked','membership',token,details=e.code,ip_address=self._ip())
                    return self.send_json({'ok':False,'error':e.code,'message':e.message,'requires_manager':e.requires_manager},409)
                except integrity_errors():
                    return self.send_json({'ok':False,'error':'duplicate_request'},409)
            if path == '/api/attendant/stamp/remove':
                if not self._need_permission(s,'remove_balance'): return
                reason=str(payload.get('reason','')).strip()[:300]
                if len(reason)<3:return self.send_json({'ok':False,'error':'removal_reason_required'},400)
                token,token_error=resolve_member_token(payload.get('token'));
                if token_error:return self.send_json({'ok':False,'error':token_error},410 if token_error=='qr_expired' else 400)
                idem=str(payload.get('idempotency_key','')).strip()[:100] or random_token(12)
                begin_write(conn)
                dupe=conn.execute('SELECT id FROM transactions WHERE idempotency_key=?',(idem,)).fetchone()
                if dupe: return self.send_json({'ok':True,'duplicate':True,'transaction_id':dupe['id']})
                m=fetchone_for_update(conn,'''SELECT m.*,c.goal,c.company_id,cu.name customer_name
                  FROM memberships m JOIN campaigns c ON c.id=m.campaign_id JOIN customers cu ON cu.id=m.customer_id
                  WHERE (m.public_id=? OR m.qr_token=?) AND c.company_id=? AND c.id=?''',(token,token,s['company_id'],s['campaign_id']))
                if not m: return self.send_json({'ok':False,'error':'membership_not_found'},404)
                if m['status']!='active': return self.send_json({'ok':False,'error':'membership_blocked'},409)
                prev=m['progress']; reward_delta=0
                if prev>0:
                    new=prev-1
                elif m['rewards_available']>0:
                    # Desfaz o selo que fechou o ciclo: restaura goal-1 e remove a recompensa gerada.
                    new=max(0,m['goal']-1); reward_delta=-1
                else:
                    return self.send_json({'ok':False,'error':'no_stamp_to_remove','message':'Este cartão não possui selo para remover.'},409)
                tx_id=insert_id(conn,'''INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,ip_address,note,created_at)
                  VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',(m['id'],s['user_id'],current_branch_id(conn,s['user_id'],s['campaign_id']),'adjustment',-1,prev,new,reward_delta,idem,self._ip(),reason,now_ts()))
                conn.execute('UPDATE memberships SET progress=?, rewards_available=rewards_available+? WHERE id=?',(new,reward_delta,m['id']))
                audit(conn,s['company_id'],s['user_id'],'stamp_remove','membership',m['public_id'],details=f'progress={prev}->{new};reward={reward_delta};reason={reason}',ip_address=self._ip()); notify_wallet_updates(conn,m['public_id'])
                return self.send_json({'ok':True,'transaction_id':tx_id,'customer_name':m['customer_name'],'previous_progress':prev,'progress':new,'reward_removed':1 if reward_delta<0 else 0})
            if path == '/api/attendant/redeem':
                if not self._need_permission(s,'redeem_reward'): return
                token,token_error=resolve_member_token(payload.get('token'));
                if token_error:return self.send_json({'ok':False,'error':token_error},410 if token_error=='qr_expired' else 400)
                idem=str(payload.get('idempotency_key','')).strip()[:100] or random_token(12)
                begin_write(conn)
                if conn.execute('SELECT id FROM transactions WHERE idempotency_key=?',(idem,)).fetchone(): return self.send_json({'ok':True,'duplicate':True})
                m=fetchone_for_update(conn,'''SELECT m.*,c.company_id,cu.name customer_name FROM memberships m JOIN campaigns c ON c.id=m.campaign_id JOIN customers cu ON cu.id=m.customer_id WHERE (m.public_id=? OR m.qr_token=?) AND c.company_id=? AND c.id=?''',(token,token,s['company_id'],s['campaign_id']))
                if not m: return self.send_json({'ok':False,'error':'membership_not_found'},404)
                if m['status']!='active': return self.send_json({'ok':False,'error':'membership_blocked'},409)
                if m['rewards_available']<1: return self.send_json({'ok':False,'error':'no_reward_available'},409)
                tx_id=insert_id(conn,'''INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,idempotency_key,ip_address,created_at)
                  VALUES(?,?,?,?,?,?,?,?,?,?,?)''',(m['id'],s['user_id'],current_branch_id(conn,s['user_id'],s['campaign_id']),'redeem',1,m['progress'],m['progress'],-1,idem,self._ip(),now_ts()))
                conn.execute('UPDATE memberships SET rewards_available=rewards_available-1 WHERE id=?',(m['id'],))
                audit(conn,s['company_id'],s['user_id'],'reward_redeem','membership',m['public_id'],details='Recompensa resgatada; novo ciclo permanece ativo.',ip_address=self._ip()); notify_wallet_updates(conn,m['public_id'])
                return self.send_json({'ok':True,'transaction_id':tx_id,'customer_name':m['customer_name']})
            if path == '/api/attendant/messages/retry':
                if s['role']!='attendant' or not s['campaign_id']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'communications'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                try: message_id=int(payload.get('message_id',0))
                except (TypeError,ValueError):message_id=0
                row=conn.execute("SELECT id,status FROM message_queue WHERE id=? AND campaign_id=?",(message_id,s['campaign_id'])).fetchone()
                if not row:return self.send_json({'ok':False,'error':'message_not_found'},404)
                if row['status'] not in ('failed','retry'):return self.send_json({'ok':False,'error':'message_not_retryable'},409)
                conn.execute("UPDATE message_queue SET status='pending',attempts=0,last_error=NULL,available_at=? WHERE id=?",(now_ts(),message_id))
                audit(conn,s['company_id'],s['user_id'],'message_retry','message_queue',message_id,ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/attendant/automations/update':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not plan_allows(conn,s['campaign_id'],'automations'):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                try: rule_id=int(payload.get('rule_id',0))
                except: rule_id=0
                channel=str(payload.get('channel','email')); enabled=1 if payload.get('enabled') else 0; message=str(payload.get('message','')).strip()[:1000]
                if channel not in ('email','whatsapp','both') or not message:return self.send_json({'ok':False,'error':'invalid_rule'},400)
                r=conn.execute('SELECT id,rule_type FROM automation_rules WHERE id=? AND campaign_id=?',(rule_id,s['campaign_id'])).fetchone()
                if not r:return self.send_json({'ok':False,'error':'rule_not_found'},404)
                if r['rule_type']=='one_to_reward':
                    campaign=conn.execute('SELECT loyalty_type FROM campaigns WHERE id=?',(s['campaign_id'],)).fetchone()
                    if not campaign or campaign['loyalty_type']!='stamps':return self.send_json({'ok':False,'error':'rule_not_available'},409)
                conn.execute('UPDATE automation_rules SET channel=?,enabled=?,message=? WHERE id=?',(channel,enabled,message,rule_id)); audit(conn,s['company_id'],s['user_id'],'automation_update','automation',rule_id,details=f'{channel}:{enabled}',ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/admin/branch/update':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                bid=int(payload.get('branch_id') or 0); name=str(payload.get('name') or '').strip()[:100]; code=re.sub(r'[^A-Z0-9_-]','',str(payload.get('code') or '').upper())[:30]
                if not bid or not name or not code:return self.send_json({'ok':False,'error':'invalid_branch'},400)
                conn.execute('UPDATE branches SET name=?,code=? WHERE id=? AND campaign_id=?',(name,code,bid,s['campaign_id'])); audit(conn,s['company_id'],s['user_id'],'branch_update','branch',bid,details=name,ip_address=self._ip()); return self.send_json({'ok':True})
            if path == '/api/admin/branch/toggle':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                bid=int(payload.get('branch_id') or 0); active=1 if payload.get('active') else 0
                conn.execute('UPDATE branches SET active=? WHERE id=? AND campaign_id=?',(active,bid,s['campaign_id'])); audit(conn,s['company_id'],s['user_id'],'branch_toggle','branch',bid,details='ativo' if active else 'inativo',ip_address=self._ip()); return self.send_json({'ok':True})

            if path == '/api/admin/branch/save':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                name=str(payload.get('name','')).strip()[:80]; code=str(payload.get('code','')).strip().upper()[:30]
                if not name or not code:return self.send_json({'ok':False,'error':'invalid_branch'},400)
                try: bid=insert_id(conn,'INSERT INTO branches(campaign_id,name,code,active,created_at) VALUES(?,?,?,?,?)',(s['campaign_id'],name,code,1,now_ts()))
                except integrity_errors():return self.send_json({'ok':False,'error':'branch_code_exists'},409)
                audit(conn,s['company_id'],s['user_id'],'branch_create','branch',bid,details=f'{name} ({code})',ip_address=self._ip());return self.send_json({'ok':True,'branch_id':bid})
            if path == '/api/admin/current-branch':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                branch_raw=payload.get('branch_id')
                try: branch_id=int(branch_raw) if branch_raw not in (None,'',0,'0') else None
                except (TypeError,ValueError):return self.send_json({'ok':False,'error':'invalid_branch'},400)
                if branch_id and not conn.execute('SELECT id FROM branches WHERE id=? AND campaign_id=? AND active=1',(branch_id,s['campaign_id'])).fetchone():return self.send_json({'ok':False,'error':'branch_not_found'},404)
                conn.execute('UPDATE users SET branch_id=? WHERE id=? AND campaign_id=? AND role=\'attendant\'',(branch_id,s['user_id'],s['campaign_id']))
                audit(conn,s['company_id'],s['user_id'],'current_branch_update','user',s['user_id'],details=f'branch_id={branch_id or "none"}',ip_address=self._ip(),branch_id=branch_id)
                return self.send_json({'ok':True,'branch_id':branch_id})
            if path == '/api/admin/staff/access':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                uid=int(payload.get('user_id') or 0); branch_id=payload.get('branch_id'); perms=payload.get('permissions') or {}
                allowed={k:bool(perms.get(k)) for k in ('add_balance','remove_balance','redeem_reward','use_gift','view_reports','send_messages')}
                conn.execute("UPDATE users SET permissions_json=?,branch_id=? WHERE id=? AND campaign_id=? AND role='attendant'",(json.dumps(allowed,ensure_ascii=False),int(branch_id) if branch_id else None,uid,s['campaign_id']))
                audit(conn,s['company_id'],s['user_id'],'staff_access_update','user',uid,details=json.dumps(allowed),ip_address=self._ip());return self.send_json({'ok':True})
            if path == '/api/client-admin/plan':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                reconcile_campaign_billing(conn,s['campaign_id'])
                plan=normalize_plan(payload.get('plan')); c=conn.execute('SELECT * FROM campaigns WHERE id=? AND company_id=?',(s['campaign_id'],s['company_id'])).fetchone(); current=normalize_plan(c['plan'])
                commitment_until=int(c['commitment_until'] or 0) if 'commitment_until' in c.keys() else 0
                if commitment_until and commitment_until>now_ts() and plan!=current:
                    return self.send_json({'ok':False,'error':'annual_commitment_active','commitment_until':commitment_until},409)
                order={'beginner':0,'intermediate':1,'pro':2}
                if c['pending_subscription_id']:
                    if c['pending_plan']==plan:
                        try: pending_sub=mp_request('GET','/preapproval/'+urllib.parse.quote(c['pending_subscription_id'],safe=''))
                        except RuntimeError as exc:return self.send_json({'ok':False,'error':str(exc)},503)
                        return self.send_json({'ok':True,'payment_required':True,'checkout_url':pending_sub.get('init_point'),'plan':current,'pending_plan':c['pending_plan']})
                    return self.send_json({'ok':False,'error':'plan_change_pending'},409)
                if plan==current:
                    if c['pending_plan']:
                        if c['subscription_id'] and current!='beginner':
                            try: mp_request('PUT','/preapproval/'+urllib.parse.quote(c['subscription_id'],safe=''),{'auto_recurring':{'transaction_amount':billing_config(current,c['billing_option'])[1]['amount'],'currency_id':'BRL'}})
                            except RuntimeError as exc:return self.send_json({'ok':False,'error':str(exc)},503)
                        conn.execute('UPDATE campaigns SET pending_plan=NULL,subscription_current_period_end=subscription_next_payment_at,subscription_cancel_at_period_end=0,subscription_change_requested_at=NULL WHERE id=?',(s['campaign_id'],))
                        audit(conn,s['company_id'],s['user_id'],'plan_change_cancelled','campaign',s['campaign_id'],details=current,ip_address=self._ip())
                        return self.send_json({'ok':True,'plan':current,'cancelled_pending':True})
                    return self.send_json({'ok':True,'plan':current,'unchanged':True})
                if plan=='beginner' and c['loyalty_type']!='stamps':return self.send_json({'ok':False,'error':'downgrade_requires_stamps'},409)
                if plan=='beginner' and conn.execute("SELECT COUNT(*) n FROM memberships WHERE campaign_id=? AND status='active'",(s['campaign_id'],)).fetchone()['n']>50:return self.send_json({'ok':False,'error':'downgrade_client_limit'},409)
                if order[plan]>order[current]:
                    # Todo upgrade pago usa uma nova assinatura; o plano atual fica intacto até o webhook.
                    token='upgrade:'+str(s['campaign_id'])+':'+plan+':'+secrets.token_urlsafe(8)
                    try: sub=create_mp_subscription(s['email'],plan,token,billing_option=normalize_billing_option(plan,c['billing_option']))
                    except RuntimeError as exc:return self.send_json({'ok':False,'error':str(exc)},503)
                    conn.execute('UPDATE campaigns SET pending_plan=?,pending_subscription_id=?,subscription_change_requested_at=? WHERE id=?',(plan,sub.get('id'),now_ts(),s['campaign_id']))
                    audit(conn,s['company_id'],s['user_id'],'plan_upgrade_requested','campaign',s['campaign_id'],details=plan,ip_address=self._ip())
                    return self.send_json({'ok':True,'payment_required':True,'checkout_url':sub.get('init_point'),'plan':current,'pending_plan':plan})
                # Downgrade: mantém o acesso atual até o fim do ciclo já pago.
                period_end=c['subscription_next_payment_at'] or c['subscription_current_period_end'] or (now_ts()+30*86400)
                if c['subscription_id'] and plan!='beginner':
                    # O novo preço vale na próxima recorrência; os recursos só mudam após confirmação do novo ciclo.
                    try: mp_request('PUT','/preapproval/'+urllib.parse.quote(c['subscription_id'],safe=''),{'auto_recurring':{'transaction_amount':billing_config(plan,normalize_billing_option(plan,c['billing_option']))[1]['amount'],'currency_id':'BRL'}})
                    except RuntimeError as exc:return self.send_json({'ok':False,'error':str(exc)},503)
                if c['subscription_id'] and plan=='beginner':
                    # Cancela cobranças futuras agora, preservando acesso local até period_end.
                    try: mp_request('PUT','/preapproval/'+urllib.parse.quote(c['subscription_id'],safe=''),{'status':'cancelled'})
                    except RuntimeError as exc:return self.send_json({'ok':False,'error':str(exc)},503)
                conn.execute('UPDATE campaigns SET pending_plan=?,subscription_current_period_end=?,subscription_cancel_at_period_end=?,subscription_change_requested_at=? WHERE id=?',(plan,period_end,1 if plan=='beginner' else 0,now_ts(),s['campaign_id']))
                audit(conn,s['company_id'],s['user_id'],'plan_downgrade_scheduled','campaign',s['campaign_id'],details=plan,ip_address=self._ip());return self.send_json({'ok':True,'plan':current,'pending_plan':plan,'effective_at':period_end})
            if path == '/api/client-admin/company/update':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                c=conn.execute('SELECT * FROM campaigns WHERE id=? AND company_id=?',(s['campaign_id'],s['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                plan=normalize_plan(c['plan']); name=str(payload.get('name',c['name'])).strip()[:80]; code=re.sub(r'[^A-Z0-9_-]','',str(payload.get('code',c['code'] or '')).upper())[:24]; loyalty=str(payload.get('loyalty_type',c['loyalty_type'])).lower();
                if plan=='beginner':loyalty='stamps'
                if not name or not code:return self.send_json({'ok':False,'error':'invalid_campaign'},400)
                if loyalty not in ('stamps','points') or (loyalty=='points' and not PLAN_FEATURES[plan]['points']):return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                try: points=int(payload.get('points_spend_cents') or c['points_spend_cents'] or 200); goal=int(payload.get('goal') or c['goal'] or 5)
                except:return self.send_json({'ok':False,'error':'invalid_campaign'},400)
                if points not in (200,300,500,1000):points=200
                if goal not in (3,5,8,10,15):goal=5
                reward=str(payload.get('reward_name',c['reward_name'] or '')).strip()[:100] or 'Recompensa do programa'; theme=str(payload.get('card_theme',c['card_theme'] or 'orange')).lower()
                if theme not in CARD_THEMES:theme=c['card_theme'] or 'orange'
                logo=None
                if payload.get('logo_image'):
                    try:logo=validate_logo_data(payload.get('logo_image'))
                    except ValueError as exc:return self.send_json({'ok':False,'error':str(exc)},400)
                
                try:
                    if plan=='pro':
                        email_provider=str(payload.get('email_provider',c['email_provider'] or 'smtp')).strip().lower()
                        if email_provider not in ('smtp','brevo'):email_provider='smtp'
                        smtp_host=str(payload.get('smtp_host',c['smtp_host'] or '')).strip()[:200];smtp_port=str(payload.get('smtp_port',c['smtp_port'] or '587')).strip()[:8] or '587';smtp_user=str(payload.get('smtp_user',c['smtp_user'] or '')).strip()[:200];smtp_from=str(payload.get('smtp_from',c['smtp_from'] or '')).strip()[:200];smtp_from_name=str(payload.get('smtp_from_name',c['smtp_from_name'] or '')).strip()[:100];smtp_security=str(payload.get('smtp_security',c['smtp_security'] or 'starttls')).strip().lower()
                        if smtp_security not in ('starttls','ssl','none'):smtp_security='starttls'
                        brevo_sender_email=str(payload.get('brevo_sender_email',c['brevo_sender_email'] or '')).strip()[:200];brevo_sender_name=str(payload.get('brevo_sender_name',c['brevo_sender_name'] or '')).strip()[:100];brevo_reply_to=str(payload.get('brevo_reply_to',c['brevo_reply_to'] or '')).strip()[:200]
                        wa_phone_id=str(payload.get('whatsapp_phone_number_id',c['whatsapp_phone_number_id'] or '')).strip()[:100];wa_waba_id=str(payload.get('whatsapp_waba_id',c['whatsapp_waba_id'] or '')).strip()[:100];wa_version=str(payload.get('whatsapp_api_version',c['whatsapp_api_version'] or 'v24.0')).strip()[:20]
                        ecommerce_platform=normalize_ecommerce_platform(payload.get('ecommerce_platform',c['ecommerce_platform'] or 'none')); ecommerce_store_url=str(payload.get('ecommerce_store_url',c['ecommerce_store_url'] or '')).strip()[:300]; old_ecommerce_platform=normalize_ecommerce_platform(c['ecommerce_platform']); ecommerce_secret=c['ecommerce_webhook_secret']
                        if ecommerce_platform!='none' and not ecommerce_secret:ecommerce_secret=secrets.token_urlsafe(24)
                        if ecommerce_platform=='none':ecommerce_secret=None
                        ecommerce_status=('awaiting_connection' if ecommerce_platform!='none' else 'not_connected') if ecommerce_platform!=old_ecommerce_platform else (c['ecommerce_status'] or ('awaiting_connection' if ecommerce_platform!='none' else 'not_connected'))
                        smtp_password_enc=c['smtp_password_enc'];brevo_api_key_enc=c['brevo_api_key_enc'];wa_token_enc=c['whatsapp_access_token_enc']
                        if payload.get('smtp_password'):smtp_password_enc=encrypt_secret(payload.get('smtp_password'))
                        if payload.get('brevo_api_key'):brevo_api_key_enc=encrypt_secret(payload.get('brevo_api_key'))
                        if payload.get('whatsapp_access_token'):wa_token_enc=encrypt_secret(payload.get('whatsapp_access_token'))
                        conn.execute('''UPDATE campaigns SET code=?,name=?,loyalty_type=?,points_spend_cents=?,goal=?,reward_name=?,card_theme=?,logo_image=COALESCE(?,logo_image),email_provider=?,smtp_host=?,smtp_port=?,smtp_user=?,smtp_password_enc=?,smtp_from=?,smtp_from_name=?,smtp_security=?,brevo_api_key_enc=?,brevo_sender_email=?,brevo_sender_name=?,brevo_reply_to=?,whatsapp_phone_number_id=?,whatsapp_waba_id=?,whatsapp_access_token_enc=?,whatsapp_api_version=?,whatsapp_integration_mode='manual',whatsapp_signup_status=?,ecommerce_platform=?,ecommerce_store_url=?,ecommerce_webhook_secret=?,ecommerce_status=? WHERE id=? AND company_id=?''',(code,name,loyalty,points,goal,reward,theme,logo,email_provider,smtp_host,smtp_port,smtp_user,smtp_password_enc,smtp_from,smtp_from_name,smtp_security,brevo_api_key_enc,brevo_sender_email,brevo_sender_name,brevo_reply_to,wa_phone_id,wa_waba_id,wa_token_enc,wa_version,'connected' if (wa_phone_id and wa_token_enc) else 'not_connected',ecommerce_platform,ecommerce_store_url,ecommerce_secret,ecommerce_status,s['campaign_id'],s['company_id']))
                    else:
                        conn.execute('UPDATE campaigns SET code=?,name=?,loyalty_type=?,points_spend_cents=?,goal=?,reward_name=?,card_theme=?,logo_image=COALESCE(?,logo_image) WHERE id=? AND company_id=?',(code,name,loyalty,points,goal,reward,theme,logo,s['campaign_id'],s['company_id']))
                except RuntimeError as exc:return self.send_json({'ok':False,'error':str(exc)},503)
                except integrity_errors():return self.send_json({'ok':False,'error':'campaign_code_exists'},409)
                audit(conn,s['company_id'],s['user_id'],'client_admin_company_update','campaign',s['campaign_id'],details=f'plan={plan}',ip_address=self._ip());return self.send_json({'ok':True})
            if path == '/api/client-admin/integration/ecommerce/rotate-secret':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                c=conn.execute('SELECT id,plan,ecommerce_platform FROM campaigns WHERE id=? AND company_id=?',(s['campaign_id'],s['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                if normalize_plan(c['plan'])!='pro':return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                if normalize_ecommerce_platform(c['ecommerce_platform'])=='none':return self.send_json({'ok':False,'error':'ecommerce_not_configured'},409)
                secret=secrets.token_urlsafe(24); conn.execute("UPDATE campaigns SET ecommerce_webhook_secret=?,ecommerce_status='awaiting_connection' WHERE id=? AND company_id=?",(secret,s['campaign_id'],s['company_id']))
                audit(conn,s['company_id'],s['user_id'],'ecommerce_secret_rotated','campaign',s['campaign_id'],ip_address=self._ip())
                public_base=(os.environ.get('PUBLIC_BASE_URL') or 'https://app.fidelizae.com.br').rstrip('/')
                return self.send_json({'ok':True,'webhook_url':public_base+f"/api/integrations/ecommerce/{s['campaign_id']}/{secret}"})

            if path == '/api/client-admin/subscription/cancel-renewal':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                password=str(payload.get('password') or '')
                u=conn.execute('SELECT password_hash FROM users WHERE id=?',(s['user_id'],)).fetchone()
                if not u or not verify_password(password,u['password_hash']):return self.send_json({'ok':False,'error':'invalid_password'},403)
                c=conn.execute('SELECT * FROM campaigns WHERE id=? AND company_id=?',(s['campaign_id'],s['company_id'])).fetchone()
                if not c or normalize_plan(c['plan'])=='beginner' or not c['subscription_id']:return self.send_json({'ok':False,'error':'no_paid_subscription'},409)
                if c['subscription_cancel_at_period_end']:
                    option=normalize_billing_option(c['plan'],c['billing_option'])
                    access_until=int((c['commitment_until'] if option in ('annual_monthly','annual_upfront') else c['subscription_current_period_end']) or 0)
                    return self.send_json({'ok':True,'already_cancelled':True,'billing_option':option,'access_until':access_until,'subscription_status':c['subscription_status']})
                option=normalize_billing_option(c['plan'],c['billing_option'])
                now=now_ts(); next_ts=int(c['subscription_next_payment_at'] or 0)
                try:
                    sub=mp_request('GET','/preapproval/'+urllib.parse.quote(str(c['subscription_id']),safe=''))
                    next_ts=_mp_timestamp(sub.get('next_payment_date')) or next_ts
                except Exception as exc:
                    print('[BILLING] cancel renewal preapproval read failed type=%s' % type(exc).__name__,flush=True)
                    return self.send_json({'ok':False,'error':'billing_provider_unavailable'},503)
                if option=='monthly':
                    if not next_ts:return self.send_json({'ok':False,'error':'billing_period_unknown'},409)
                    if not _best_effort_cancel_subscription(c['subscription_id']):return self.send_json({'ok':False,'error':'billing_cancel_failed'},503)
                    access_until=next_ts
                    conn.execute("UPDATE campaigns SET subscription_cancel_at_period_end=1,renewal_cancelled_at=?,subscription_status='non_renewing',subscription_current_period_end=?,subscription_next_payment_at=NULL,subscription_status_updated_at=? WHERE id=?",(now,access_until,now,c['id']))
                elif option=='annual_upfront':
                    access_until=int(c['commitment_until'] or 0)
                    if not access_until:return self.send_json({'ok':False,'error':'billing_period_unknown'},409)
                    if not _best_effort_cancel_subscription(c['subscription_id']):return self.send_json({'ok':False,'error':'billing_cancel_failed'},503)
                    conn.execute("UPDATE campaigns SET subscription_cancel_at_period_end=1,renewal_cancelled_at=?,subscription_status='non_renewing',subscription_next_payment_at=NULL,subscription_status_updated_at=? WHERE id=?",(now,now,c['id']))
                else:
                    access_until=int(c['commitment_until'] or 0)
                    if not access_until:return self.send_json({'ok':False,'error':'billing_period_unknown'},409)
                    paid=_approved_subscription_invoice_count(c['subscription_id'])
                    if paid is not None and paid>=12:
                        if not _best_effort_cancel_subscription(c['subscription_id']):return self.send_json({'ok':False,'error':'billing_cancel_failed'},503)
                    conn.execute('UPDATE campaigns SET subscription_cancel_at_period_end=1,renewal_cancelled_at=?,subscription_status_updated_at=? WHERE id=?',(now,now,c['id']))
                audit(conn,s['company_id'],s['user_id'],'subscription_renewal_cancelled','campaign',c['id'],details=option,ip_address=self._ip())
                return self.send_json({'ok':True,'billing_option':option,'access_until':access_until,'subscription_status':'non_renewing' if option!='annual_monthly' else 'active'})

            if path == '/api/client-admin/account/delete':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                password=str(payload.get('password') or '')
                u=conn.execute('SELECT password_hash FROM users WHERE id=?',(s['user_id'],)).fetchone()
                if not u or not verify_password(password,u['password_hash']):return self.send_json({'ok':False,'error':'invalid_password'},403)
                c=conn.execute('SELECT * FROM campaigns WHERE id=? AND company_id=?',(s['campaign_id'],s['company_id'])).fetchone()
                now=now_ts(); option=normalize_billing_option(c['plan'],c['billing_option']) if c else 'free'; commitment=int(c['commitment_until'] or 0) if c else 0
                # No anual parcelado, excluir a conta no meio do compromisso retiraria o acesso
                # enquanto as parcelas continuariam vencendo. Por isso a exclusão fica disponível
                # após o término do compromisso; o cliente pode cancelar a renovação imediatamente.
                if c and option=='annual_monthly' and commitment and commitment>now:
                    return self.send_json({'ok':False,'error':'annual_commitment_delete_blocked','commitment_until':commitment,'renewal_cancelled':bool(c['subscription_cancel_at_period_end'])},409)
                # Antes de apagar acessos, garantimos que não ficará nenhuma assinatura/checkout
                # remoto cobrando a empresa. Se o provedor falhar, não concluímos a exclusão.
                ids=[]
                if c:
                    for key in ('subscription_id','pending_subscription_id','previous_subscription_id'):
                        sid=c[key]
                        if sid and sid not in ids:ids.append(sid)
                for sid in ids:
                    if not _best_effort_cancel_subscription(sid):return self.send_json({'ok':False,'error':'billing_cancel_failed'},503)
                conn.execute('UPDATE campaigns SET active=0,subscription_status=?,subscription_cancel_at_period_end=1,renewal_cancelled_at=COALESCE(renewal_cancelled_at,?),subscription_next_payment_at=NULL WHERE id=?',('cancelled',now,s['campaign_id']))
                # Encerra os acessos e libera os e-mails para um cadastro futuro sem apagar
                # os IDs históricos usados por auditoria/transações. O endereço original não
                # permanece na tabela users e, portanto, não bloqueia a restrição UNIQUE(email).
                deleted_at=now
                campaign_users=conn.execute('SELECT id FROM users WHERE campaign_id=?',(s['campaign_id'],)).fetchall()
                for deleted_user in campaign_users:
                    uid=int(deleted_user['id'])
                    tombstone=f'deleted-{uid}-{deleted_at}-{secrets.token_hex(4)}@deleted.invalid'
                    conn.execute('UPDATE users SET active=0,email=?,password_hash=? WHERE id=?',(tombstone,hash_password(secrets.token_urlsafe(32)),uid))
                    conn.execute('DELETE FROM sessions WHERE user_id=?',(uid,))
                audit(conn,s['company_id'],s['user_id'],'client_admin_account_delete','campaign',s['campaign_id'],details='billing_cancelled;emails_released;billing_option='+option,ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/client-admin/staff/create':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                name=str(payload.get('name','')).strip()[:80]; email=normalize_email(payload.get('email')); password=str(payload.get('password','')).strip(); branch_raw=payload.get('branch_id')
                try: branch_id=int(branch_raw) if branch_raw not in (None,'',0,'0') else None
                except (TypeError,ValueError):return self.send_json({'ok':False,'error':'invalid_branch'},400)
                if branch_id and not conn.execute('SELECT id FROM branches WHERE id=? AND campaign_id=? AND active=1',(branch_id,s['campaign_id'])).fetchone():return self.send_json({'ok':False,'error':'branch_not_found'},404)
                if conn.execute('SELECT COUNT(*) n FROM branches WHERE campaign_id=? AND active=1',(s['campaign_id'],)).fetchone()['n'] and not branch_id:return self.send_json({'ok':False,'error':'branch_required'},400)
                if len(name)<2 or not email or not password_is_strong(password,12):return self.send_json({'ok':False,'error':'invalid_staff'},400)
                plan=campaign_plan(conn,s['campaign_id']); limit=PLAN_FEATURES[plan]['staff_limit']; current=conn.execute("SELECT COUNT(*) n FROM users WHERE campaign_id=? AND role='attendant' AND active=1 AND is_client_admin=0",(s['campaign_id'],)).fetchone()['n'];
                if limit and current>=limit:return self.send_json({'ok':False,'error':'plan_staff_limit','limit':limit},403)
                try:new_id=insert_id(conn,'INSERT INTO users(company_id,name,email,password_hash,role,campaign_id,is_client_admin,branch_id,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(s['company_id'],name,email,hash_password(password),'attendant',s['campaign_id'],0,branch_id,now_ts()))
                except integrity_errors():return self.send_json({'ok':False,'error':'email_exists'},409)
                c=conn.execute('SELECT name FROM campaigns WHERE id=?',(s['campaign_id'],)).fetchone(); q=None
                if email_configured(global_email_config()):q=enqueue_message(conn,None,'attendant_welcome',email,{'name':name,'client_name':c['name']})
                audit(conn,s['company_id'],s['user_id'],'client_admin_staff_create','user',new_id,details=f'{email};branch_id={branch_id or "none"}',ip_address=self._ip()); return self.send_json({'ok':True,'user_id':new_id,'queue_id':q})
            if path == '/api/client-admin/staff/update':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                if not self.csrf_ok():return self.send_json({'ok':False,'error':'csrf_failed'},403)
                uid=int(payload.get('user_id') or 0); name=str(payload.get('name','')).strip()[:80]; email=normalize_email(payload.get('email')); password=str(payload.get('password','')); branch_raw=payload.get('branch_id')
                try: branch_id=int(branch_raw) if branch_raw not in (None,'',0,'0') else None
                except (TypeError,ValueError):return self.send_json({'ok':False,'error':'invalid_branch'},400)
                if uid==int(s['user_id']):return self.send_json({'ok':False,'error':'cannot_edit_self'},400)
                if not name or not email or (password and not password_is_strong(password,12)):return self.send_json({'ok':False,'error':'invalid_staff'},400)
                if branch_id and not conn.execute('SELECT id FROM branches WHERE id=? AND campaign_id=? AND active=1',(branch_id,s['campaign_id'])).fetchone():return self.send_json({'ok':False,'error':'branch_not_found'},404)
                if conn.execute('SELECT COUNT(*) n FROM branches WHERE campaign_id=? AND active=1',(s['campaign_id'],)).fetchone()['n'] and not branch_id:return self.send_json({'ok':False,'error':'branch_required'},400)
                u=conn.execute("SELECT id FROM users WHERE id=? AND campaign_id=? AND role='attendant'",(uid,s['campaign_id'])).fetchone()
                if not u:return self.send_json({'ok':False,'error':'staff_not_found'},404)
                duplicate=conn.execute("SELECT id FROM users WHERE lower(email)=lower(?) AND id<>?",(email,uid)).fetchone()
                if duplicate:return self.send_json({'ok':False,'error':'email_exists'},409)
                if password: conn.execute("UPDATE users SET name=?,email=?,password_hash=?,branch_id=? WHERE id=?",(name,email,hash_password(password),branch_id,uid))
                else: conn.execute("UPDATE users SET name=?,email=?,branch_id=? WHERE id=?",(name,email,branch_id,uid))
                audit(conn,s['company_id'],s['user_id'],'client_admin_staff_update','user',uid,details=f'{email};branch_id={branch_id or "none"}',ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/client-admin/staff/delete':
                if s['role']!='attendant' or not s['is_client_admin']:return self.send_json({'ok':False,'error':'forbidden'},403)
                try: uid=int(payload.get('user_id',0))
                except:uid=0
                if uid==s['user_id']:return self.send_json({'ok':False,'error':'cannot_delete_self'},409)
                u=conn.execute("SELECT id FROM users WHERE id=? AND campaign_id=? AND role='attendant' AND is_client_admin=0",(uid,s['campaign_id'])).fetchone()
                if not u:return self.send_json({'ok':False,'error':'user_not_found'},404)
                conn.execute('DELETE FROM users WHERE id=?',(uid,)); audit(conn,s['company_id'],s['user_id'],'client_admin_staff_delete','user',uid,ip_address=self._ip());return self.send_json({'ok':True})
            if path == '/api/manager/campaign':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                name=str(payload.get('name','')).strip()[:80]; reward=str(payload.get('reward_name','')).strip()[:100] or 'Catálogo de recompensas'; code=re.sub(r'[^A-Z0-9_-]','',str(payload.get('code','')).upper())[:24]
                if not payload.get('terms_accepted') or not payload.get('privacy_accepted'): return self.send_json({'ok':False,'error':'legal_acceptance_required'},400)
                plan=normalize_plan(payload.get('plan')); loyalty_type=str(payload.get('loyalty_type','stamps')).strip().lower()
                try: points_spend_cents=int(payload.get('points_spend_cents') or 200)
                except: points_spend_cents=200
                icon=str(payload.get('icon','☕'))[:8]; goal=int(payload.get('goal',5))
                card_theme=str(payload.get('card_theme','green')).strip().lower()
                if card_theme not in CARD_THEMES: card_theme='green'
                if plan=='beginner' and loyalty_type!='stamps': return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                if not name or not code or loyalty_type not in ('stamps','points') or (loyalty_type=='stamps' and (not reward or goal not in (3,5,8,10,15))) or (loyalty_type=='points' and points_spend_cents not in (200,300,500,1000)): return self.send_json({'ok':False,'error':'invalid_campaign'},400)
                try:
                    logo_image=validate_logo_data(payload.get('logo_image'))
                    if not logo_image:
                        return self.send_json({'ok':False,'error':'logo_required'},400)
                except ValueError as exc:
                    return self.send_json({'ok':False,'error':str(exc)},400)
                # Integrações opcionais do próprio cliente podem ser cadastradas agora ou depois.
                email_provider=str(payload.get('email_provider','smtp')).strip().lower()
                if email_provider not in ('smtp','brevo'): email_provider='smtp'
                smtp_host=str(payload.get('smtp_host','')).strip()[:200]
                smtp_port=str(payload.get('smtp_port','587')).strip()[:8] or '587'
                smtp_user=str(payload.get('smtp_user','')).strip()[:200]
                smtp_from=str(payload.get('smtp_from','')).strip()[:200]
                smtp_from_name=str(payload.get('smtp_from_name','')).strip()[:100]
                smtp_security=str(payload.get('smtp_security','starttls')).strip().lower()
                if smtp_security not in ('starttls','ssl','none'): smtp_security='starttls'
                brevo_sender_email=str(payload.get('brevo_sender_email','')).strip()[:200]
                brevo_sender_name=str(payload.get('brevo_sender_name','')).strip()[:100]
                brevo_reply_to=str(payload.get('brevo_reply_to','')).strip()[:200]
                wa_phone_id=str(payload.get('whatsapp_phone_number_id','')).strip()[:100]
                wa_waba_id=str(payload.get('whatsapp_waba_id','')).strip()[:100]
                wa_version=str(payload.get('whatsapp_api_version','v24.0')).strip()[:20] or 'v24.0'
                wa_mode=str(payload.get('whatsapp_integration_mode','none')).strip().lower()
                if wa_mode not in ('embedded','manual','none'): wa_mode='none'
                ecommerce_platform=normalize_ecommerce_platform(payload.get('ecommerce_platform'))
                ecommerce_store_url=str(payload.get('ecommerce_store_url','')).strip()[:300]
                ecommerce_secret=secrets.token_urlsafe(24) if ecommerce_platform!='none' else None
                ecommerce_status='awaiting_connection' if ecommerce_platform!='none' else 'not_connected'
                try:
                    smtp_password_enc=encrypt_secret(payload.get('smtp_password')) if payload.get('smtp_password') else None
                    brevo_api_key_enc=encrypt_secret(payload.get('brevo_api_key')) if payload.get('brevo_api_key') else None
                    wa_token_enc=encrypt_secret(payload.get('whatsapp_access_token')) if payload.get('whatsapp_access_token') else None
                except RuntimeError as exc:
                    return self.send_json({'ok':False,'error':str(exc)},503)
                wa_status='connected' if (wa_phone_id and wa_token_enc) else ('awaiting_connection' if wa_mode=='embedded' else 'not_connected')
                try:
                    new_id=insert_id(conn,'''INSERT INTO campaigns(company_id,code,name,reward_name,goal,icon,logo_image,card_theme,plan,loyalty_type,points_spend_cents,min_stamp_interval_sec,max_stamps_per_hour,max_stamps_per_attendant_day,
                        smtp_host,smtp_port,smtp_user,smtp_password_enc,smtp_from,smtp_from_name,smtp_security,email_provider,brevo_api_key_enc,brevo_sender_email,brevo_sender_name,brevo_reply_to,
                        whatsapp_phone_number_id,whatsapp_waba_id,whatsapp_access_token_enc,whatsapp_api_version,whatsapp_integration_mode,whatsapp_signup_status,
                        ecommerce_platform,ecommerce_store_url,ecommerce_webhook_secret,ecommerce_status,created_at)
                     VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(
                        s['company_id'],code,name,reward,goal,icon,logo_image,card_theme,plan,loyalty_type,points_spend_cents,int(payload.get('min_interval',0)),int(payload.get('max_hour',0)),int(payload.get('max_day',500)),
                        smtp_host,smtp_port,smtp_user,smtp_password_enc,smtp_from,smtp_from_name,smtp_security,email_provider,brevo_api_key_enc,brevo_sender_email,brevo_sender_name,brevo_reply_to,
                        wa_phone_id,wa_waba_id,wa_token_enc,wa_version,wa_mode,wa_status,
                        ecommerce_platform,ecommerce_store_url,ecommerce_secret,ecommerce_status,now_ts()))
                except integrity_errors(): return self.send_json({'ok':False,'error':'campaign_code_exists'},409)
                conn.execute('INSERT INTO legal_acceptances(campaign_id,user_id,terms_version,privacy_version,accepted_at,ip_address) VALUES(?,?,?,?,?,?)',(new_id,s['user_id'],TERMS_VERSION,PRIVACY_VERSION,now_ts(),self._ip()))
                audit(conn,s['company_id'],s['user_id'],'campaign_create','campaign',new_id,details=code,ip_address=self._ip())
                return self.send_json({'ok':True,'campaign_id':new_id})
            if path == '/api/manager/campaign/update':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: campaign_id=int(payload.get('campaign_id',0))
                except (TypeError,ValueError): campaign_id=0
                name=str(payload.get('name','')).strip()[:80]; reward=str(payload.get('reward_name','')).strip()[:100] or 'Catálogo de recompensas'; code=re.sub(r'[^A-Z0-9_-]','',str(payload.get('code','')).upper())[:24]
                plan=normalize_plan(payload.get('plan')); loyalty_type=str(payload.get('loyalty_type','stamps')).strip().lower()
                try: points_spend_cents=int(payload.get('points_spend_cents') or 200)
                except: points_spend_cents=200
                icon=str(payload.get('icon','☕'))[:8]
                try: goal=int(payload.get('goal',5)); min_interval=int(payload.get('min_interval',0)); max_hour=int(payload.get('max_hour',0)); max_day=int(payload.get('max_day',500))
                except (TypeError,ValueError): return self.send_json({'ok':False,'error':'invalid_campaign'},400)
                if plan=='beginner' and loyalty_type!='stamps': return self.send_json({'ok':False,'error':'plan_feature_not_available'},403)
                if campaign_id<1 or not name or not code or loyalty_type not in ('stamps','points') or (loyalty_type=='stamps' and (not reward or goal not in (3,5,8,10,15))) or (loyalty_type=='points' and points_spend_cents not in (200,300,500,1000)) or min_interval<0 or max_hour<0 or max_day<1:
                    return self.send_json({'ok':False,'error':'invalid_campaign'},400)
                c=rowdict(conn.execute('SELECT * FROM campaigns WHERE id=? AND company_id=?',(campaign_id,s['company_id'])).fetchone())
                card_theme=str(payload.get('card_theme',c.get('card_theme') or 'green')).strip().lower()
                if card_theme not in CARD_THEMES: card_theme='green'
                if not c: return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                logo_image=c['logo_image']
                if payload.get('logo_image'):
                    try: logo_image=validate_logo_data(payload.get('logo_image'))
                    except ValueError as exc: return self.send_json({'ok':False,'error':str(exc)},400)
                # Integrações por cliente. Segredos vazios preservam os valores já salvos.
                smtp_host=str(payload.get('smtp_host','')).strip()[:200]
                smtp_port=str(payload.get('smtp_port','587')).strip()[:8] or '587'
                smtp_user=str(payload.get('smtp_user','')).strip()[:200]
                smtp_from=str(payload.get('smtp_from','')).strip()[:200]
                smtp_from_name=str(payload.get('smtp_from_name','')).strip()[:100]
                smtp_security=str(payload.get('smtp_security','starttls')).strip().lower()
                email_provider=str(payload.get('email_provider',c.get('email_provider') or 'smtp')).strip().lower()
                if email_provider not in ('smtp','brevo'): email_provider='smtp'
                brevo_sender_email=str(payload.get('brevo_sender_email','')).strip()[:200]
                brevo_sender_name=str(payload.get('brevo_sender_name','')).strip()[:100]
                brevo_reply_to=str(payload.get('brevo_reply_to','')).strip()[:200]
                wa_phone_id=str(payload.get('whatsapp_phone_number_id','')).strip()[:100]
                wa_waba_id=str(payload.get('whatsapp_waba_id','')).strip()[:100]
                wa_version=str(payload.get('whatsapp_api_version','v24.0')).strip()[:20]
                wa_mode=str(payload.get('whatsapp_integration_mode',c.get('whatsapp_integration_mode') or 'manual')).strip().lower()
                if wa_mode not in ('embedded','manual','none'): wa_mode='manual'
                ecommerce_platform=normalize_ecommerce_platform(payload.get('ecommerce_platform',c.get('ecommerce_platform') or 'none'))
                ecommerce_store_url=str(payload.get('ecommerce_store_url',c.get('ecommerce_store_url') or '')).strip()[:300]
                old_platform=normalize_ecommerce_platform(c.get('ecommerce_platform'))
                ecommerce_secret=c.get('ecommerce_webhook_secret')
                if ecommerce_platform!='none' and not ecommerce_secret: ecommerce_secret=secrets.token_urlsafe(24)
                if ecommerce_platform=='none': ecommerce_secret=None
                ecommerce_status=('awaiting_connection' if ecommerce_platform!='none' else 'not_connected') if ecommerce_platform!=old_platform else (c.get('ecommerce_status') or ('awaiting_connection' if ecommerce_platform!='none' else 'not_connected'))
                smtp_password_enc=c['smtp_password_enc']
                brevo_api_key_enc=c.get('brevo_api_key_enc')
                wa_token_enc=c['whatsapp_access_token_enc']
                try:
                    if payload.get('smtp_password'): smtp_password_enc=encrypt_secret(payload.get('smtp_password'))
                    if payload.get('brevo_api_key'): brevo_api_key_enc=encrypt_secret(payload.get('brevo_api_key'))
                    if payload.get('whatsapp_access_token'): wa_token_enc=encrypt_secret(payload.get('whatsapp_access_token'))
                except RuntimeError as exc:
                    return self.send_json({'ok':False,'error':str(exc)},503)
                try:
                    conn.execute('''UPDATE campaigns SET code=?,name=?,reward_name=?,goal=?,icon=?,logo_image=?,card_theme=?,plan=?,loyalty_type=?,points_spend_cents=?,min_stamp_interval_sec=?,max_stamps_per_hour=?,max_stamps_per_attendant_day=?,
                        smtp_host=?,smtp_port=?,smtp_user=?,smtp_password_enc=?,smtp_from=?,smtp_from_name=?,smtp_security=?,email_provider=?,brevo_api_key_enc=?,brevo_sender_email=?,brevo_sender_name=?,brevo_reply_to=?,
                        whatsapp_phone_number_id=?,whatsapp_waba_id=?,whatsapp_access_token_enc=?,whatsapp_api_version=?,whatsapp_integration_mode=?,whatsapp_signup_status=?,
                        ecommerce_platform=?,ecommerce_store_url=?,ecommerce_webhook_secret=?,ecommerce_status=?
                        WHERE id=? AND company_id=?''',(code,name,reward,goal,icon,logo_image,card_theme,plan,loyalty_type,points_spend_cents,min_interval,max_hour,max_day,
                        smtp_host,smtp_port,smtp_user,smtp_password_enc,smtp_from,smtp_from_name,smtp_security,email_provider,brevo_api_key_enc,brevo_sender_email,brevo_sender_name,brevo_reply_to,
                        wa_phone_id,wa_waba_id,wa_token_enc,wa_version,wa_mode,'connected' if (wa_phone_id and wa_token_enc) else ('awaiting_connection' if wa_mode=='embedded' else 'not_connected'),
                        ecommerce_platform,ecommerce_store_url,ecommerce_secret,ecommerce_status,campaign_id,s['company_id']))
                except integrity_errors(): return self.send_json({'ok':False,'error':'campaign_code_exists'},409)
                audit(conn,s['company_id'],s['user_id'],'campaign_update','campaign',campaign_id,details=code,ip_address=self._ip())
                return self.send_json({'ok':True,'campaign_id':campaign_id})
            if path == '/api/manager/integration/ecommerce/rotate-secret':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: campaign_id=int(payload.get('campaign_id',0))
                except: campaign_id=0
                c=rowdict(conn.execute('SELECT ecommerce_platform FROM campaigns WHERE id=? AND company_id=?',(campaign_id,s['company_id'])).fetchone())
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                platform=normalize_ecommerce_platform(c.get('ecommerce_platform'))
                if platform=='none':return self.send_json({'ok':False,'error':'ecommerce_not_configured'},409)
                secret=secrets.token_urlsafe(24)
                conn.execute("UPDATE campaigns SET ecommerce_webhook_secret=?,ecommerce_status='awaiting_connection' WHERE id=?",(secret,campaign_id))
                audit(conn,s['company_id'],s['user_id'],'ecommerce_secret_rotated','campaign',campaign_id,ip_address=self._ip())
                public_base=(os.environ.get('PUBLIC_BASE_URL') or 'https://app.fidelizae.com.br').rstrip('/')
                return self.send_json({'ok':True,'webhook_url':public_base+f'/api/integrations/ecommerce/{campaign_id}/{secret}', 'secret':secret})

            if path == '/api/manager/integration/test-email':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: campaign_id=int(payload.get('campaign_id',0))
                except (TypeError,ValueError): campaign_id=0
                cfg=email_config_for_client(conn,campaign_id)
                target=normalize_email(payload.get('email') or s['email'])
                if not target or not email_configured(cfg): return self.send_json({'ok':False,'error':'email_provider_not_configured'},503)
                msg=EmailMessage(); msg['Subject']='Teste de e-mail • Fidelizaê!'; msg['To']=target; msg.set_content('Configuração SMTP testada com sucesso.')
                result=send_email_message(msg,cfg)
                return self.send_json({'ok':bool(result.get('sent')),'result':result},200 if result.get('sent') else 502)
            if path == '/api/manager/integration/whatsapp/embedded-complete':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: campaign_id=int(payload.get('campaign_id',0))
                except (TypeError,ValueError): campaign_id=0
                code=str(payload.get('code','')).strip(); phone_id=str(payload.get('phone_number_id','')).strip(); waba_id=str(payload.get('waba_id','')).strip()
                if not (campaign_id and code and phone_id and waba_id): return self.send_json({'ok':False,'error':'embedded_signup_incomplete'},400)
                c=conn.execute('SELECT id FROM campaigns WHERE id=? AND company_id=?',(campaign_id,s['company_id'])).fetchone()
                if not c: return self.send_json({'ok':False,'error':'client_not_found'},404)
                try:
                    exchanged=meta_exchange_code(code); token=str(exchanged.get('access_token','')).strip()
                    if not token: raise RuntimeError('meta_token_missing')
                    details=meta_phone_details(phone_id,token); token_enc=encrypt_secret(token)
                    conn.execute("""UPDATE campaigns SET whatsapp_integration_mode='embedded',whatsapp_signup_status='connected',whatsapp_phone_number_id=?,whatsapp_waba_id=?,whatsapp_access_token_enc=?,whatsapp_api_version=?,whatsapp_connected_at=? WHERE id=? AND company_id=?""",(phone_id,waba_id,token_enc,os.environ.get('META_GRAPH_VERSION','v24.0'),now_iso(),campaign_id,s['company_id']))
                    conn.commit()
                    return self.send_json({'ok':True,'status':'connected','display_phone_number':details.get('display_phone_number'),'verified_name':details.get('verified_name')})
                except Exception as exc: return self.send_json({'ok':False,'error':'embedded_signup_failed','detail':str(exc)[:700]},502)
            if path == '/api/manager/integration/test-whatsapp':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: campaign_id=int(payload.get('campaign_id',0))
                except (TypeError,ValueError): campaign_id=0
                phone=normalize_phone(payload.get('phone'))
                cfg=whatsapp_config_for_client(conn,campaign_id)
                if not phone or not whatsapp_cloud_configured(cfg): return self.send_json({'ok':False,'error':'whatsapp_not_configured'},503)
                try: response=send_whatsapp_cloud(phone,'Teste de integração • Fidelizaê!',cfg)
                except Exception as exc: return self.send_json({'ok':False,'error':'whatsapp_test_failed','detail':str(exc)[:500]},502)
                return self.send_json({'ok':True,'message_id':((response.get('messages') or [{}])[0]).get('id')})
            if path == '/api/manager/password':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                current_password=str(payload.get('current_password',''))
                new_password=str(payload.get('new_password','')).strip()
                if not password_is_strong(new_password,12): return self.send_json({'ok':False,'error':'invalid_new_password'},400)
                u=conn.execute("SELECT id,password_hash FROM users WHERE id=? AND role='manager' AND active=1",(s['user_id'],)).fetchone()
                if not u or not verify_password(current_password,u['password_hash']): return self.send_json({'ok':False,'error':'invalid_current_password'},401)
                if verify_password(new_password,u['password_hash']): return self.send_json({'ok':False,'error':'same_password'},409)
                conn.execute('UPDATE users SET password_hash=? WHERE id=?',(hash_password(new_password),s['user_id']))
                conn.execute('DELETE FROM sessions WHERE user_id=?',(s['user_id'],))
                new_token,new_csrf=create_session(conn,s['user_id'])
                audit(conn,s['company_id'],s['user_id'],'password_change','user',s['user_id'],details='manager_sessions_revoked',ip_address=self._ip())
                print(f'[AUTH] MANAGER_PASSWORD_CHANGED user_id={s["user_id"]} sessions_revoked=True')
                return self.send_json({'ok':True,'csrf':new_csrf},200,{'Set-Cookie':_session_cookie(new_token)})

            if path == '/api/manager/admin':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                name=str(payload.get('name','')).strip()[:80]
                email=normalize_email(payload.get('email'))
                password=str(payload.get('password','')).strip()
                if len(name)<2 or not email or '@' not in email or not password_is_strong(password,12):
                    return self.send_json({'ok':False,'error':'invalid_manager_admin','message':'Preencha nome, e-mail e uma senha com pelo menos 12 caracteres, com maiúscula, minúscula e número.'},400)
                existing=conn.execute('SELECT id FROM users WHERE lower(email)=lower(?) LIMIT 1',(email,)).fetchone()
                if existing:
                    return self.send_json({'ok':False,'error':'email_exists','message':'Este e-mail já está cadastrado no sistema.'},409)
                try:
                    new_id=insert_id(conn,'INSERT INTO users(company_id,name,email,password_hash,role,active,is_client_admin,campaign_id,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(s['company_id'],name,email,hash_password(password),'manager',1,0,None,now_ts()))
                except integrity_errors():
                    return self.send_json({'ok':False,'error':'email_exists','message':'Este e-mail já está cadastrado no sistema.'},409)
                audit(conn,s['company_id'],s['user_id'],'manager_admin_create','user',new_id,details=email,ip_address=self._ip())
                return self.send_json({'ok':True,'user_id':new_id})

            if path == '/api/manager/admin/update':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: admin_id=int(payload.get('id',0))
                except (TypeError,ValueError): admin_id=0
                name=str(payload.get('name','')).strip()[:80]
                email=normalize_email(payload.get('email'))
                password=str(payload.get('password','')).strip()
                if admin_id<1 or len(name)<2 or not email or '@' not in email or (password and not password_is_strong(password,12)):
                    return self.send_json({'ok':False,'error':'invalid_manager_admin','message':'Preencha nome e e-mail válidos. A nova senha, quando informada, deve ter pelo menos 12 caracteres, com maiúscula, minúscula e número.'},400)
                target=conn.execute("SELECT id FROM users WHERE id=? AND company_id=? AND role='manager'",(admin_id,s['company_id'])).fetchone()
                if not target:return self.send_json({'ok':False,'error':'not_found'},404)
                duplicate=conn.execute('SELECT id FROM users WHERE lower(email)=lower(?) AND id<>? LIMIT 1',(email,admin_id)).fetchone()
                if duplicate:return self.send_json({'ok':False,'error':'email_exists','message':'Este e-mail já está cadastrado no sistema.'},409)
                if password:
                    conn.execute('UPDATE users SET name=?,email=?,password_hash=? WHERE id=?',(name,email,hash_password(password),admin_id))
                    conn.execute('DELETE FROM sessions WHERE user_id=?',(admin_id,))
                else:
                    conn.execute('UPDATE users SET name=?,email=? WHERE id=?',(name,email,admin_id))
                audit(conn,s['company_id'],s['user_id'],'manager_admin_update','user',admin_id,details=email,ip_address=self._ip())
                return self.send_json({'ok':True})

            if path == '/api/manager/admin/remove':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: admin_id=int(payload.get('id',0))
                except (TypeError,ValueError): admin_id=0
                if admin_id==int(s['user_id']):return self.send_json({'ok':False,'error':'cannot_remove_self'},409)
                target=conn.execute("SELECT id,email,active FROM users WHERE id=? AND company_id=? AND role='manager'",(admin_id,s['company_id'])).fetchone()
                if not target:return self.send_json({'ok':False,'error':'not_found'},404)
                active_count=conn.execute("SELECT COUNT(*) n FROM users WHERE company_id=? AND role='manager' AND active=1",(s['company_id'],)).fetchone()['n']
                if target['active'] and int(active_count)<=1:return self.send_json({'ok':False,'error':'last_manager'},409)
                conn.execute('UPDATE users SET active=0 WHERE id=?',(admin_id,))
                conn.execute('DELETE FROM sessions WHERE user_id=?',(admin_id,))
                audit(conn,s['company_id'],s['user_id'],'manager_admin_remove','user',admin_id,details=target['email'],ip_address=self._ip())
                return self.send_json({'ok':True})

            if path == '/api/manager/staff/check-email':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                email=normalize_email(payload.get('email'))
                if not email or '@' not in email:
                    return self.send_json({'ok':False,'available':False,'error':'invalid_email','message':'Informe um e-mail válido.'},400)
                configured_admin=normalize_email(os.environ.get('CLUBE_ADMIN_EMAIL',''))
                if configured_admin and email == configured_admin:
                    return self.send_json({
                        'ok':True,'available':False,'error':'admin_email_reserved',
                        'message':'Este e-mail pertence ao administrador geral da plataforma. Utilize outro e-mail para cadastrar um usuário da empresa.'
                    })
                existing=conn.execute('SELECT id,name,active,is_client_admin,campaign_id FROM users WHERE lower(email)=lower(?) LIMIT 1',(email,)).fetchone()
                if existing:
                    return self.send_json({
                        'ok':True,'available':False,'error':'email_exists',
                        'message':'Este e-mail já está cadastrado no sistema. Cada usuário deve possuir um e-mail próprio.'
                    })
                return self.send_json({'ok':True,'available':True,'message':'E-mail disponível para cadastro.'})

            if path == '/api/manager/staff':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                name=str(payload.get('name','')).strip()[:80]
                email=normalize_email(payload.get('email'))
                password=str(payload.get('password','')).strip()
                is_client_admin=1 if payload.get('is_client_admin') else 0
                try: campaign_id=int(payload.get('campaign_id',0))
                except (TypeError,ValueError): campaign_id=0
                if len(name)<2 or not email or '@' not in email or not password_is_strong(password,12) or campaign_id<1:
                    return self.send_json({'ok':False,'error':'invalid_staff','message':'Preencha nome, e-mail, cliente, perfil e uma senha com pelo menos 12 caracteres, com maiúscula, minúscula e número.'},400)
                configured_admin=normalize_email(os.environ.get('CLUBE_ADMIN_EMAIL',''))
                if configured_admin and email == configured_admin:
                    return self.send_json({
                        'ok':False,'error':'admin_email_reserved',
                        'message':'Este e-mail pertence ao administrador geral da plataforma. Utilize outro e-mail para cadastrar um usuário da empresa.'
                    },409)
                existing=conn.execute('SELECT id FROM users WHERE lower(email)=lower(?) LIMIT 1',(email,)).fetchone()
                if existing:
                    return self.send_json({
                        'ok':False,'error':'email_exists',
                        'message':'Este e-mail já está cadastrado no sistema. Cada usuário deve possuir um e-mail próprio.'
                    },409)
                client=conn.execute('SELECT id,name FROM campaigns WHERE id=? AND company_id=? AND active=1',(campaign_id,s['company_id'])).fetchone()
                if not client:
                    return self.send_json({'ok':False,'error':'client_not_found','message':'A empresa selecionada não foi encontrada ou está arquivada.'},404)
                try:
                    new_id=insert_id(conn,'INSERT INTO users(company_id,name,email,password_hash,role,campaign_id,is_client_admin,created_at) VALUES(?,?,?,?,?,?,?,?)',(s['company_id'],name,email,hash_password(password),'attendant',campaign_id,is_client_admin,now_ts()))
                except integrity_errors():
                    return self.send_json({'ok':False,'error':'email_exists','message':'Este e-mail já está cadastrado no sistema. Cada usuário deve possuir um e-mail próprio.'},409)
                audit(conn,s['company_id'],s['user_id'],'staff_create','user',new_id,details=f'{email}:attendant:client={campaign_id}',ip_address=self._ip())
                smtp_cfg=global_email_config(); email_result={'queued':False,'reason':'email_provider_not_configured'}
                if email_configured(smtp_cfg):
                    qid=enqueue_message(conn,None,'attendant_welcome',email,{'name':name,'client_name':client['name']}); email_result={'queued':True,'queue_id':qid}
                    audit(conn,s['company_id'],s['user_id'],'staff_welcome_queued','user',new_id,details=f'queue={qid}',ip_address=self._ip())
                return self.send_json({'ok':True,'user_id':new_id,'client_name':client['name'],'welcome_email':email_result})
            if path == '/api/manager/campaign/delete':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: campaign_id=int(payload.get('campaign_id',0))
                except (TypeError,ValueError): campaign_id=0
                c=conn.execute('SELECT id,name,code FROM campaigns WHERE id=? AND company_id=?',(campaign_id,s['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                conn.execute('UPDATE campaigns SET active=0 WHERE id=? AND company_id=?',(campaign_id,s['company_id']))
                audit(conn,s['company_id'],s['user_id'],'client_delete','campaign',campaign_id,details=c['code'],ip_address=self._ip())
                return self.send_json({'ok':True,'archived_campaign_id':campaign_id})
            if path == '/api/manager/campaign/restore':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: campaign_id=int(payload.get('campaign_id',0))
                except (TypeError,ValueError): campaign_id=0
                c=conn.execute('SELECT id,code FROM campaigns WHERE id=? AND company_id=?',(campaign_id,s['company_id'])).fetchone()
                if not c:return self.send_json({'ok':False,'error':'campaign_not_found'},404)
                conn.execute('UPDATE campaigns SET active=1 WHERE id=? AND company_id=?',(campaign_id,s['company_id']))
                audit(conn,s['company_id'],s['user_id'],'client_restore','campaign',campaign_id,details=c['code'],ip_address=self._ip())
                return self.send_json({'ok':True})
            if path == '/api/manager/staff/delete':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                try: user_id=int(payload.get('user_id',0))
                except (TypeError,ValueError): user_id=0
                if user_id == s['user_id']: return self.send_json({'ok':False,'error':'cannot_delete_self'},409)
                u=conn.execute('SELECT id,name,email,role FROM users WHERE id=? AND company_id=?',(user_id,s['company_id'])).fetchone()
                if not u: return self.send_json({'ok':False,'error':'user_not_found'},404)
                configured_admin=os.environ.get('CLUBE_ADMIN_EMAIL','').strip().lower()
                if configured_admin and u['email'].lower()==configured_admin:
                    return self.send_json({'ok':False,'error':'configured_admin_protected'},409)
                if u['role']=='manager':
                    managers=conn.execute("SELECT COUNT(*) n FROM users WHERE company_id=? AND role='manager' AND active=1",(s['company_id'],)).fetchone()['n']
                    if managers <= 1: return self.send_json({'ok':False,'error':'last_manager_protected'},409)
                audit(conn,s['company_id'],s['user_id'],'staff_delete','user',user_id,details=f"{u['email']}:{u['role']}",ip_address=self._ip())
                conn.execute('DELETE FROM users WHERE id=? AND company_id=?',(user_id,s['company_id']))
                return self.send_json({'ok':True,'deleted_user_id':user_id})
            if path == '/api/manager/block':
                if s['role']!='manager': return self.send_json({'ok':False,'error':'forbidden'},403)
                token,token_error=resolve_member_token(payload.get('token'));
                if token_error:return self.send_json({'ok':False,'error':token_error},410 if token_error=='qr_expired' else 400)
                status='blocked' if payload.get('blocked',True) else 'active'
                m=conn.execute('''SELECT m.* FROM memberships m JOIN campaigns c ON c.id=m.campaign_id WHERE (m.public_id=? OR m.qr_token=?) AND c.company_id=?''',(token,token,s['company_id'])).fetchone()
                if not m: return self.send_json({'ok':False,'error':'membership_not_found'},404)
                conn.execute('UPDATE memberships SET status=? WHERE id=?',(status,m['id']))
                ttype='block' if status=='blocked' else 'unblock'
                conn.execute('''INSERT INTO transactions(membership_id,user_id,branch_id,type,value,previous_progress,new_progress,rewards_delta,ip_address,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)''',(m['id'],s['user_id'],current_branch_id(conn,s['user_id']),ttype,0,m['progress'],m['progress'],0,self._ip(),'manager action',now_ts()))
                audit(conn,s['company_id'],s['user_id'],ttype,'membership',m['public_id'],ip_address=self._ip())
                return self.send_json({'ok':True,'status':status})
            return self.send_json({'ok':False,'error':'not_found'},404)


class SentryHTTPServer(ThreadingHTTPServer):
    """Captura exceções não tratadas das threads HTTP sem alterar a resposta padrão do servidor."""
    def handle_error(self, request, client_address):
        if SENTRY_ENABLED and sentry_sdk is not None:
            try:
                sentry_sdk.capture_exception()
            except Exception as exc:
                print(f'[SENTRY] falha ao registrar exceção: {exc}')
        return super().handle_error(request, client_address)


def main():
    parser=argparse.ArgumentParser(); parser.add_argument('--host',default=os.environ.get('HOST','0.0.0.0')); parser.add_argument('--port',type=int,default=int(os.environ.get('PORT','8000'))); parser.add_argument('--init-only',action='store_true'); args=parser.parse_args()
    init_db(DB_PATH,seed=True)
    ensure_configured_staff(DB_PATH)
    if args.init_only:
        print(f'Database initialized: {DB_PATH}'); return
    threading.Thread(target=background_loop,daemon=True,name='clube-worker').start()
    srv=SentryHTTPServer((args.host,args.port),Handler)
    print(f'Fidelizaê! {VERSION} em http://{args.host}:{args.port}')
    try: srv.serve_forever()
    except KeyboardInterrupt: pass
    finally: srv.server_close()

if __name__=='__main__': main()
